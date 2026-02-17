from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form, Response
from sqlalchemy.orm import Session
from database import get_db, User
from auth import get_current_admin_or_presenter, get_password_hash
import csv
import io
import os
import logging

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/admin", tags=["bulk_operations"])

# Import logging functions
from main import log_admin_action

@router.get("/download-student-template")
async def download_student_template(
    current_user = Depends(get_current_admin_or_presenter)
):
    """Download Excel template for student bulk upload"""
    try:
        try:
            import pandas as pd
            
            sample_data = {
                'Username': ['john_doe', 'jane_smith', 'mike_wilson'],
                'Email': ['john@example.com', 'jane@example.com', 'mike@example.com'],
                'Password': ['password123', 'password456', 'password789'],
                'College': ['MIT University', 'Stanford University', 'Harvard University'],
                'Department': ['Computer Science', 'Engineering', 'Mathematics'],
                'Year': ['2024', '2023', '2025']
            }
            
            df = pd.DataFrame(sample_data)
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Students', index=False)
                
                instructions = pd.DataFrame({
                    'Instructions': [
                        '1. Fill in the student data in the Students sheet',
                        '2. Required: Username, Email, Password, College, Department, Year',
                        '3. Type will be automatically set to Student',
                        '4. Username must be unique',
                        '5. Email must be valid and unique',
                        '6. Password minimum 6 characters',
                        '7. Save as Excel (.xlsx) or CSV (.csv) format',
                        '8. Upload the file using the Bulk Import feature'
                    ]
                })
                instructions.to_excel(writer, sheet_name='Instructions', index=False)
            
            output.seek(0)
            
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "inline; filename=student_template.xlsx"}
            )
            
        except ImportError:
            csv_content = "Username,Email,Password,College,Department,Year\n"
            csv_content += "john_doe,john@example.com,password123,MIT University,Computer Science,2024\n"
            csv_content += "jane_smith,jane@example.com,password456,Stanford University,Engineering,2023\n"
            csv_content += "mike_wilson,mike@example.com,password789,Harvard University,Mathematics,2025\n"
            
            return Response(
                content=csv_content,
                media_type="text/csv",
                headers={"Content-Disposition": "inline; filename=student_template.csv"}
            )
            
    except Exception as e:
        logger.error(f"Download student template error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate student template")

@router.get("/download-faculty-template")
async def download_faculty_template(
    current_user = Depends(get_current_admin_or_presenter)
):
    """Download Excel template for faculty bulk upload"""
    try:
        try:
            import pandas as pd
            
            sample_data = {
                'Username': ['dr_sarah_jones', 'prof_michael_brown', 'dr_lisa_chen'],
                'Email': ['sarah@university.edu', 'michael@university.edu', 'lisa@university.edu'],
                'Password': ['faculty123', 'faculty456', 'faculty789'],
                'College': ['MIT University', 'Stanford University', 'Harvard University'],
                'Department': ['Computer Science', 'Engineering', 'Mathematics'],
                'Experience': ['10', '15', '8'],
                'Designation': ['Associate Professor', 'Professor', 'Assistant Professor'],
                'Specialization': ['Machine Learning, Data Science', 'Software Engineering, AI', 'Statistics, Data Analysis'],
                'Employment_Type': ['Full-time', 'Full-time', 'Visiting'],
                'Joining_Date': ['2020-01-15', '2018-08-20', '2022-09-01']
            }
            
            df = pd.DataFrame(sample_data)
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Faculty', index=False)
                
                instructions = pd.DataFrame({
                    'Instructions': [
                        '1. Fill in the faculty data in the Faculty sheet',
                        '2. Required: Username, Email, Password, College, Department, Experience, Designation, Specialization',
                        '3. Type will be automatically set to Faculty',
                        '4. Optional: Employment_Type (defaults to Full-time), Joining_Date',
                        '5. Username must be unique',
                        '6. Email must be valid and unique',
                        '7. Password minimum 6 characters',
                        '8. Experience must be a number greater than 0',
                        '9. Employment_Type options: Full-time, Visiting, Contract, Part-time',
                        '10. Joining_Date format: YYYY-MM-DD (optional)',
                        '11. Save as Excel (.xlsx) or CSV (.csv) format',
                        '12. Upload the file using the Bulk Import feature'
                    ]
                })
                instructions.to_excel(writer, sheet_name='Instructions', index=False)
            
            output.seek(0)
            
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "inline; filename=faculty_template.xlsx"}
            )
            
        except ImportError:
            csv_content = "Username,Email,Password,College,Department,Experience,Designation,Specialization,Employment_Type,Joining_Date\n"
            csv_content += "dr_sarah_jones,sarah@university.edu,faculty123,MIT University,Computer Science,10,Associate Professor,Machine Learning Data Science,Full-time,2020-01-15\n"
            csv_content += "prof_michael_brown,michael@university.edu,faculty456,Stanford University,Engineering,15,Professor,Software Engineering AI,Full-time,2018-08-20\n"
            csv_content += "dr_lisa_chen,lisa@university.edu,faculty789,Harvard University,Mathematics,8,Assistant Professor,Statistics Data Analysis,Visiting,2022-09-01\n"
            
            return Response(
                content=csv_content,
                media_type="text/csv",
                headers={"Content-Disposition": "inline; filename=faculty_template.csv"}
            )
            
    except Exception as e:
        logger.error(f"Download faculty template error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate faculty template")

@router.post("/users/bulk-upload")
async def bulk_upload_users(
    file: UploadFile = File(...),
    user_type_filter: str = Form("Student"),
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        file_ext = os.path.splitext(file.filename)[1].lower()
        
        if file_ext not in ['.csv', '.xlsx', '.xls']:
            raise HTTPException(status_code=400, detail="Only CSV and Excel files are allowed")
        
        if user_type_filter not in ['Student', 'Faculty']:
            raise HTTPException(status_code=400, detail="Invalid user type filter. Must be 'Student' or 'Faculty'")
        
        content = await file.read()
        records = []
        
        if file_ext == '.csv':
            try:
                csv_content = content.decode('utf-8')
                csv_reader = csv.DictReader(io.StringIO(csv_content))
                records = list(csv_reader)
            except UnicodeDecodeError:
                csv_content = content.decode('utf-8-sig')
                csv_reader = csv.DictReader(io.StringIO(csv_content))
                records = list(csv_reader)
        else:
            try:
                import pandas as pd
                df = pd.read_excel(io.BytesIO(content))
                records = df.to_dict('records')
            except ImportError:
                from openpyxl import load_workbook
                workbook = load_workbook(io.BytesIO(content))
                sheet = workbook.active
                headers = [cell.value for cell in sheet[1] if cell.value]
                records = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if any(cell for cell in row):
                        record = {}
                        for i, value in enumerate(row):
                            if i < len(headers) and headers[i]:
                                record[headers[i]] = value
                        records.append(record)
        
        if not records:
            raise HTTPException(status_code=400, detail="No data found in file. Please check the file format.")
        
        created_users = []
        errors = []
        
        for row_num, row in enumerate(records, 1):
            try:
                username = str(row.get('Username') or row.get('username') or row.get('UserName') or '').strip()
                email = str(row.get('Email') or row.get('email') or '').strip()
                password = str(row.get('Password') or row.get('password') or '').strip()
                user_type = user_type_filter
                college = str(row.get('College') or row.get('college') or '').strip()
                department = str(row.get('Department') or row.get('department') or '').strip()
                year = str(row.get('Year') or row.get('year') or '').strip()
                
                # Faculty-specific fields
                experience = None
                designation = None
                specialization = None
                employment_type = 'Full-time'
                joining_date = None
                
                if user_type_filter == 'Faculty':
                    experience = row.get('Experience') or row.get('experience') or None
                    designation = str(row.get('Designation') or row.get('designation') or '').strip() or None
                    specialization = str(row.get('Specialization') or row.get('specialization') or '').strip() or None
                    employment_type = str(row.get('Employment_Type') or row.get('employment_type') or row.get('EmploymentType') or 'Full-time').strip()
                    joining_date = row.get('Joining_Date') or row.get('joining_date') or row.get('JoiningDate') or None
                
                # Skip empty rows
                if not any([username, email, password]):
                    continue
                
                # Validate required fields
                if not all([username, email, password]):
                    errors.append(f"Row {row_num}: Missing required fields (Username, Email, Password)")
                    continue
                
                # Validate email format
                if '@' not in email or '.' not in email:
                    errors.append(f"Row {row_num}: Invalid email format '{email}'")
                    continue
                
                # Check for duplicates
                if db.query(User).filter(User.username == username).first():
                    errors.append(f"Row {row_num}: Username '{username}' already exists")
                    continue
                
                if db.query(User).filter(User.email == email).first():
                    errors.append(f"Row {row_num}: Email '{email}' already exists")
                    continue
                
                # Faculty-specific validation
                if user_type_filter == 'Faculty':
                    if not designation or designation.strip() == '':
                        errors.append(f"Row {row_num}: Designation is required for faculty")
                        continue
                    
                    if not specialization or specialization.strip() == '':
                        errors.append(f"Row {row_num}: Specialization is required for faculty")
                        continue
                    
                    if experience:
                        try:
                            experience = int(experience)
                            if experience <= 0 or experience > 50:
                                errors.append(f"Row {row_num}: Experience must be between 1 and 50 years")
                                continue
                        except:
                            errors.append(f"Row {row_num}: Experience must be a valid number")
                            continue
                    else:
                        errors.append(f"Row {row_num}: Experience is required for faculty")
                        continue
                
                # Handle joining_date conversion
                joining_date_obj = None
                if joining_date:
                    try:
                        from datetime import datetime
                        if isinstance(joining_date, str):
                            joining_date_obj = datetime.fromisoformat(joining_date)
                        else:
                            joining_date_obj = joining_date
                    except:
                        pass
                
                final_year = year if year else ("N/A" if user_type_filter == "Faculty" else "2024")
                
                # Create user
                user = User(
                    username=username,
                    email=email,
                    password_hash=get_password_hash(password),
                    role=user_type_filter,
                    user_type=user_type_filter,
                    college=college,
                    department=department,
                    year=final_year,
                    experience=experience,
                    designation=designation,
                    specialization=specialization,
                    employment_type=employment_type,
                    joining_date=joining_date_obj
                )
                
                db.add(user)
                created_users.append({
                    'username': username,
                    'email': email,
                    'type': user_type_filter,
                    'college': college
                })
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        if created_users:
            db.commit()
            
            # Log the successful upload
            from database import Admin
            if hasattr(current_admin, 'username') and db.query(Admin).filter(Admin.id == current_admin.id).first():
                log_admin_action(
                    admin_id=current_admin.id,
                    admin_username=current_admin.username,
                    action_type="BULK_UPLOAD",
                    resource_type="USER",
                    details=f"Bulk uploaded {len(created_users)} {user_type_filter.lower()}s from {file.filename}"
                )
        else:
            db.rollback()
        
        return {
            "message": f"Successfully created {len(created_users)} {user_type_filter.lower()}s from {file.filename}",
            "created_users": [u['username'] for u in created_users],
            "errors": errors[:10],
            "total_processed": len(records),
            "success_count": len(created_users),
            "error_count": len(errors),
            "user_type": user_type_filter
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Bulk upload users error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"File upload failed: {str(e)}")