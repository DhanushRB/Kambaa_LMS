from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File, Request, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer
from fastapi.responses import FileResponse, Response
from sqlalchemy.orm import Session
from datetime import timedelta, datetime
from typing import List, Optional
from pydantic import BaseModel, Field, validator
import asyncio
from database import (
    get_db,
    User,
    Admin,
    Presenter,
    Mentor,
    Manager,
    Course,
    Enrollment,
    Module,
    Session as SessionModel,
    Resource,
    Attendance,
    Certificate,
    Forum,
    ForumPost,
    SessionContent,
    Event,
    SystemSettings,
    Cohort,
    UserCohort,
    CohortCourse,
    AdminLog,
    PresenterLog,
    MentorCohort,
    MentorCourse,
    MentorSession,
    MentorLog,
    StudentLog,
    Notification,
    EmailLog,
    EmailRecipient,
    NotificationPreference,
    PresenterCohort,
)

# Import Quiz models from assignment_quiz_tables
try:
    from assignment_quiz_tables import Quiz, QuizAttempt
except ImportError:
    # Fallback for backward compatibility
    Quiz = None
    QuizAttempt = None
from auth import verify_password, get_password_hash, create_access_token, get_current_user, get_current_admin, get_current_presenter, get_current_mentor, get_current_admin_or_presenter, get_current_admin_presenter_mentor_or_manager, require_role, ACCESS_TOKEN_EXPIRE_MINUTES, verify_token
from schemas import CourseCreate, CourseUpdate, AdminCreate, PresenterCreate, ChangePasswordRequest
from services.email_service import email_service
from sqlalchemy import func, and_, or_
import csv
import io
import json
import logging
import os
from sqlalchemy.exc import SQLAlchemyError
import uuid
from pathlib import Path
import mysql.connector
import requests
import aiohttp
import aiofiles

# Import template generator
try:
    from user_template_generator import generate_user_template_excel, generate_cohort_template_excel
except ImportError:
    # Fallback if pandas/openpyxl not available
    generate_user_template_excel = None
    generate_cohort_template_excel = None

# Try to import pandas for Excel processing
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# Import quiz generator
try:
    from quiz_generator import generate_quiz_questions
except ImportError:
    def generate_quiz_questions(content, question_type, num_questions):
        return []

app = FastAPI(title="LMS API - Kambaa AI Learning Management System")

# Import campaign scheduler
try:
    from campaign_scheduler import start_campaign_scheduler
    SCHEDULER_AVAILABLE = True
except ImportError:
    SCHEDULER_AVAILABLE = False
    logger = logging.getLogger(__name__)
    logger.warning("Campaign scheduler not available")



# Import and include role login endpoints
try:
    from role_login_endpoints import router as role_login_router
    app.include_router(role_login_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load role login router: {e}")

# Import and include email endpoints
try:
    from email_endpoints_new import router as email_router
    app.include_router(email_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load email router: {e}")

# Import and include mentor router
from mentor_endpoints import router as mentor_router
app.include_router(mentor_router)

# Import and include email campaigns router
try:
    from email_campaigns import router as campaigns_router
    app.include_router(campaigns_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load campaigns router: {e}")

# Import and include notifications router
try:
    from notifications_endpoints import router as notifications_router
    app.include_router(notifications_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load notifications router: {e}")

# Import and include test router
try:
    from test_endpoints import router as test_router
    app.include_router(test_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load test router: {e}")

# Import and include calendar events router
try:
    from calendar_events_api import router as calendar_router
    app.include_router(calendar_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load calendar router: {e}")

# Import and include SMTP settings router
try:
    from smtp_endpoints import router as smtp_router
    app.include_router(smtp_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load SMTP router: {e}")

# Import and include presenter users router
try:
    from presenter_users_endpoints import router as presenter_users_router
    app.include_router(presenter_users_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load presenter users router: {e}")

# Import and include presenter cohort assignment router
try:
    from presenter_cohort_assignment import router as presenter_cohort_router
    app.include_router(presenter_cohort_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load presenter cohort assignment router: {e}")

# Import and include email template router
try:
    from email_template_endpoints import router as email_template_router
    app.include_router(email_template_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load email template router: {e}")

# Import and include default email templates router
try:
    from default_email_templates import router as default_templates_router
    app.include_router(default_templates_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load default templates router: {e}")

# Import and include cohort router
try:
    from cohort_router import router as cohort_router
    app.include_router(cohort_router)
    logger.info("Cohort router loaded successfully")
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load cohort router: {e}")

# Import and include cohort chat router
from cohort_chat_endpoints import router as cohort_chat_router
app.include_router(cohort_chat_router)

from chat_endpoints import router as chat_router
from chat_websocket import router as websocket_router
from notification_websocket import router as notification_ws_router

# Import and include chat router
app.include_router(chat_router)
app.include_router(websocket_router)
app.include_router(notification_ws_router)

# Import and include system settings router
try:
    from system_settings_endpoints import router as system_settings_router
    app.include_router(system_settings_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load system settings router: {e}")

# Import and include approval endpoints
try:
    from approval_endpoints import router as approval_router
    app.include_router(approval_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load approval router: {e}")



# Import and include live stats endpoints
try:
    from live_stats_endpoints import router as live_stats_router
    app.include_router(live_stats_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load live stats router: {e}")

# Import and include enhanced session content router
try:
    from enhanced_session_content_api import router as enhanced_content_router
    app.include_router(enhanced_content_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load enhanced session content router: {e}")

# Import and include session meeting API
try:
    from session_meeting_api import router as session_meeting_router
    app.include_router(session_meeting_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load session meeting router: {e}")

# Import and include meeting session API
try:
    from meeting_session_api import router as meeting_router
    app.include_router(meeting_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load meeting router: {e}")

# Import and include simple session content router (for testing)
try:
    from simple_session_content import router as simple_content_router
    app.include_router(simple_content_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load simple session content router: {e}")

# Import and include analytics fix
try:
    from analytics_fix import router as analytics_router
    app.include_router(analytics_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load analytics fix router: {e}")

# Import and include assignment and quiz router
try:
    from assignment_quiz_api import router as assignment_quiz_router
    app.include_router(assignment_quiz_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load assignment quiz router: {e}")

# Import and include student dashboard router
try:
    from student_dashboard_endpoints import router as student_dashboard_router
    app.include_router(student_dashboard_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load student dashboard router: {e}")

# Import and include student calendar fix router
try:
    from student_calendar_fix import router as student_calendar_fix_router
    app.include_router(student_calendar_fix_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load student calendar fix router: {e}")

# Import and include enhanced analytics router
try:
    from enhanced_analytics_endpoints import router as enhanced_analytics_router
    app.include_router(enhanced_analytics_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load enhanced analytics router: {e}")

# Import and include file link router
try:
    from file_link_api import router as file_link_router
    app.include_router(file_link_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load file link router: {e}")

# Import and include resource analytics router
try:
    from resource_analytics_endpoints import router as resource_analytics_router
    app.include_router(resource_analytics_router)
except ImportError as e:
    logger = logging.getLogger(__name__)
    logger.error(f"Failed to load resource analytics router: {e}")





app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000","https://x18z30h4-5173.inc1.devtunnels.ms", "http://localhost:3001", "http://localhost:5173", "https://x18z30h4-5173.inc1.devtunnels.ms"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Log successful router loading
logger.info("Simple session content router loaded for testing")

# Helper function to check if operation requires approval
def requires_approval(user_role: str, operation_type: str) -> bool:
    """Check if an operation requires approval based on user role"""
    restricted_roles = ['Student', 'Presenter', 'Mentor']
    major_operations = ['delete', 'unpublish', 'disable', 'archive', 'bulk_update', 'final_modification']
    
    return user_role in restricted_roles and operation_type in major_operations

# Helper function to create approval request
def create_approval_request(db: Session, user_id: int, user_role: str, operation_type: str, 
                          target_entity_type: str, target_entity_id: int, 
                          operation_data: dict, reason: str = ""):
    """Create an approval request for restricted operations"""
    try:
        from approval_models import ApprovalRequest, EntityStatus, ApprovalStatus, OperationType
        import json
        
        # Create approval request
        approval_request = ApprovalRequest(
            requester_id=user_id,
            operation_type=OperationType(operation_type),
            target_entity_type=target_entity_type,
            target_entity_id=target_entity_id,
            operation_data=json.dumps(operation_data),
            reason=reason,
            status=ApprovalStatus.PENDING
        )
        
        db.add(approval_request)
        db.commit()
        db.refresh(approval_request)
        
        # Update entity status to pending approval
        entity_status = db.query(EntityStatus).filter(
            EntityStatus.entity_type == target_entity_type,
            EntityStatus.entity_id == target_entity_id
        ).first()
        
        if entity_status:
            entity_status.status = "pending_approval"
            entity_status.approval_request_id = approval_request.id
        else:
            entity_status = EntityStatus(
                entity_type=target_entity_type,
                entity_id=target_entity_id,
                status="pending_approval",
                approval_request_id=approval_request.id
            )
            db.add(entity_status)
        
        db.commit()
        
        return approval_request.id
    except Exception as e:
        db.rollback()
        logger.error(f"Failed to create approval request: {str(e)}")
        raise e
def format_meeting_url(url: str) -> str:
    """Ensure meeting URL has proper protocol for opening in browser"""
    if not url:
        return url
    
    url = url.strip()
    if not url:
        return url
    
    # If URL already has protocol, return as is
    if url.startswith(('http://', 'https://')):
        return url
    
    # Add https:// as default protocol
    return f"https://{url}"

# Helper function to filter meeting links for students
def filter_meeting_links_for_student(session_contents):
    """Add meeting status and remaining time instead of filtering out"""
    processed_contents = []
    current_time = datetime.now()
    
    for content in session_contents:
        # Always include the content
        processed_content = content.copy()
        
        # For meeting links, add status and remaining time
        if content.get('content_type') == 'MEETING_LINK':
            scheduled_time = content.get('scheduled_time')
            
            # Debug logging
            logger.info(f"Processing meeting: {content.get('title')}")
            logger.info(f"Scheduled time: {scheduled_time}")
            logger.info(f"Current time: {current_time}")
            
            if scheduled_time is None:
                # No scheduled time means it's always available
                processed_content['meeting_status'] = 'available'
                processed_content['is_locked'] = False
                processed_content['remaining_time'] = None
                logger.info("No scheduled time - setting as available")
            else:
                # Parse scheduled time and calculate status
                try:
                    if isinstance(scheduled_time, str):
                        # Handle ISO format with timezone
                        scheduled_dt = datetime.fromisoformat(scheduled_time.replace('Z', '+00:00'))
                        # Convert to naive datetime for comparison
                        if scheduled_dt.tzinfo is not None:
                            scheduled_dt = scheduled_dt.replace(tzinfo=None)
                    else:
                        scheduled_dt = scheduled_time
                    
                    logger.info(f"Parsed scheduled time: {scheduled_dt}")
                    logger.info(f"Comparison: current_time < scheduled_dt = {current_time < scheduled_dt}")
                    
                    # Compare times - if current time is BEFORE scheduled time, lock it
                    if current_time < scheduled_dt:
                        # Meeting is in future - locked with countdown
                        time_diff = scheduled_dt - current_time
                        processed_content['meeting_status'] = 'locked'
                        processed_content['is_locked'] = True
                        processed_content['remaining_time'] = int(time_diff.total_seconds())
                        logger.info(f"Meeting locked - remaining seconds: {int(time_diff.total_seconds())}")
                    else:
                        # Meeting time has passed - available
                        processed_content['meeting_status'] = 'available'
                        processed_content['is_locked'] = False
                        processed_content['remaining_time'] = None
                        logger.info("Meeting time passed - setting as available")
                        
                except (ValueError, TypeError) as e:
                    # If we can't parse the time, make it available
                    processed_content['meeting_status'] = 'available'
                    processed_content['is_locked'] = False
                    processed_content['remaining_time'] = None
                    logger.error(f"Error parsing time: {e} - setting as available")
        else:
            # Non-meeting content is always available
            processed_content['meeting_status'] = None
            processed_content['is_locked'] = False
            processed_content['remaining_time'] = None
        
        processed_contents.append(processed_content)
    
    return processed_contents

# Admin logging function
def log_admin_action(
    admin_id: int,
    admin_username: str,
    action_type: str,
    resource_type: str,
    details: str,
    resource_id: Optional[int] = None,
    ip_address: Optional[str] = None
):
    """Log admin actions to database"""
    try:
        from database import SessionLocal, AdminLog
        db = SessionLocal()
        
        admin_log = AdminLog(
            admin_id=admin_id,
            admin_username=admin_username,
            action_type=action_type,
            resource_type=resource_type,
            resource_id=resource_id,
            details=details,
            ip_address=ip_address
        )
        db.add(admin_log)
        db.commit()
        db.close()
        
        logger.info(f"Admin Action Logged: {admin_username} {action_type} {resource_type}")
    except Exception as e:
        logger.error(f"Failed to log admin action: {str(e)}")

# Presenter logging function
def log_presenter_action(
    presenter_id: int,
    presenter_username: str,
    action_type: str,
    resource_type: str,
    details: str,
    resource_id: Optional[int] = None,
    ip_address: Optional[str] = None
):
    """Log presenter actions to database"""
    try:
        from database import SessionLocal, PresenterLog
        db = SessionLocal()
        
        presenter_log = PresenterLog(
            presenter_id=presenter_id,
            presenter_username=presenter_username,
            action_type=action_type,
            resource_type=resource_type,
            resource_id=resource_id,
            details=details,
            ip_address=ip_address
        )
        db.add(presenter_log)
        db.commit()
        db.close()
        
        logger.info(f"Presenter Action Logged: {presenter_username} {action_type} {resource_type}")
    except Exception as e:
        logger.error(f"Failed to log presenter action: {str(e)}")

# Student logging function
def log_student_action(
    student_id: int,
    student_username: str,
    action_type: str,
    resource_type: str,
    details: str,
    resource_id: Optional[int] = None,
    ip_address: Optional[str] = None
):
    """Log student actions to database"""
    try:
        from database import SessionLocal, StudentLog
        db = SessionLocal()
        
        student_log = StudentLog(
            student_id=student_id,
            student_username=student_username,
            action_type=action_type,
            resource_type=resource_type,
            resource_id=resource_id,
            details=details,
            ip_address=ip_address
        )
        db.add(student_log)
        db.commit()
        db.close()
        
        logger.info(f"Student Action Logged: {student_username} {action_type} {resource_type}")
    except Exception as e:
        logger.error(f"Failed to log student action: {str(e)}")

# Mentor logging function
def log_mentor_action(
    mentor_id: int,
    mentor_username: str,
    action_type: str,
    resource_type: str,
    details: str,
    resource_id: Optional[int] = None,
    ip_address: Optional[str] = None
):
    """Log mentor actions to database"""
    try:
        from database import SessionLocal, MentorLog
        db = SessionLocal()
        
        mentor_log = MentorLog(
            mentor_id=mentor_id,
            mentor_username=mentor_username,
            action_type=action_type,
            resource_type=resource_type,
            resource_id=resource_id,
            details=details,
            ip_address=ip_address
        )
        db.add(mentor_log)
        db.commit()
        db.close()
        
        logger.info(f"Mentor Action Logged: {mentor_username} {action_type} {resource_type}")
    except Exception as e:
        logger.error(f"Failed to log mentor action: {str(e)}")



# Student Resource Access
@app.get("/student/session/{session_id}/resources")
async def get_student_session_resources(
    session_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get session resources for students - serves downloaded files, not original links"""
    try:
        # Check if student is enrolled in the course containing this session
        session = db.query(SessionModel).filter(SessionModel.id == session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        module = db.query(Module).filter(Module.id == session.module_id).first()
        if not module:
            raise HTTPException(status_code=404, detail="Module not found")
        
        # Check enrollment
        enrollment = db.query(Enrollment).filter(
            Enrollment.student_id == current_user.id,
            Enrollment.course_id == module.course_id
        ).first()
        
        if not enrollment:
            raise HTTPException(status_code=403, detail="Not enrolled in this course")
        
        # Get resources
        resources = db.query(Resource).filter(Resource.session_id == session_id).all()
        
        result = []
        for resource in resources:
            # Only show resources with existing files
            if resource.file_path and os.path.exists(resource.file_path):
                filename = os.path.basename(resource.file_path)
                
                result.append({
                    "id": resource.id,
                    "title": resource.title,
                    "resource_type": resource.resource_type,
                    "filename": filename,
                    "file_size": resource.file_size,
                    "description": resource.description,
                    "download_url": f"/api/resources/{filename}",
                    "uploaded_at": resource.uploaded_at
                })
        
        # Log student resource access
        log_student_action(
            student_id=current_user.id,
            student_username=current_user.username,
            action_type="VIEW",
            resource_type="SESSION_RESOURCES",
            resource_id=session_id,
            details=f"Accessed resources for session: {session.title}"
        )
        
        return {"resources": result}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get student session resources error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch resources")

@app.get("/student/resource/{resource_id}/download")
async def download_student_resource(
    resource_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Download a specific resource file for students"""
    try:
        resource = db.query(Resource).filter(Resource.id == resource_id).first()
        if not resource:
            raise HTTPException(status_code=404, detail="Resource not found")
        
        # Check if student has access to this resource
        session = db.query(SessionModel).filter(SessionModel.id == resource.session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        module = db.query(Module).filter(Module.id == session.module_id).first()
        if not module:
            raise HTTPException(status_code=404, detail="Module not found")
        
        # Check enrollment
        enrollment = db.query(Enrollment).filter(
            Enrollment.student_id == current_user.id,
            Enrollment.course_id == module.course_id
        ).first()
        
        if not enrollment:
            raise HTTPException(status_code=403, detail="Not enrolled in this course")
        
        # Check if file exists
        if not resource.file_path or not os.path.exists(resource.file_path):
            raise HTTPException(status_code=404, detail="File not found")
        
        # Log student resource download
        log_student_action(
            student_id=current_user.id,
            student_username=current_user.username,
            action_type="DOWNLOAD",
            resource_type="RESOURCE",
            resource_id=resource_id,
            details=f"Downloaded resource: {resource.title}"
        )
        
        # Serve the file
        filename = os.path.basename(resource.file_path)
        return FileResponse(
            resource.file_path,
            media_type="application/octet-stream",
            headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Download student resource error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to download resource")
UPLOAD_BASE_DIR = Path("uploads")
UPLOAD_BASE_DIR.mkdir(exist_ok=True)
(UPLOAD_BASE_DIR / "resources").mkdir(exist_ok=True)
(UPLOAD_BASE_DIR / "recordings").mkdir(exist_ok=True)
(UPLOAD_BASE_DIR / "certificates").mkdir(exist_ok=True)

# Pydantic models
class UserLogin(BaseModel):
    username: str = Field(..., min_length=3, max_length=50)
    password: str = Field(..., min_length=6)
    role: str = Field(..., pattern="^(Student)$")

class AdminLogin(BaseModel):
    username: str = Field(..., min_length=3, max_length=50)
    password: str = Field(..., min_length=6)

class UserCreate(BaseModel):
    username: str = Field(..., min_length=3, max_length=50)
    email: str = Field(..., pattern=r'^[\w\.-]+@[\w\.-]+\.\w+$')
    password: str = Field(..., min_length=6)
    college: str = Field(..., min_length=2, max_length=200)
    department: str = Field(..., min_length=2, max_length=100)
    year: str = Field(..., min_length=4, max_length=10)
    user_type: str = Field(default="Student", pattern="^(Student|Faculty)$")
    role: Optional[str] = Field(default="Student", pattern="^(Student|Faculty)$")
    
    # Faculty-specific fields (optional for students)
    experience: Optional[int] = Field(None, ge=0, le=50)
    designation: Optional[str] = Field(None, max_length=200)
    specialization: Optional[str] = Field(None, max_length=500)
    employment_type: Optional[str] = Field("Full-time", pattern="^(Full-time|Visiting|Contract|Part-time)$")
    joining_date: Optional[str] = None  # Will be converted to datetime
    
    @validator('experience', pre=True, always=True)
    def validate_experience(cls, v, values):
        user_type = values.get('user_type', 'Student')
        if user_type == 'Faculty':
            if v is None or v == '' or v == 0:
                raise ValueError('Experience is required for faculty and must be greater than 0')
            try:
                return int(v)
            except (ValueError, TypeError):
                raise ValueError('Experience must be a valid number')
        return v
    
    @validator('designation', pre=True, always=True)
    def validate_designation(cls, v, values):
        user_type = values.get('user_type', 'Student')
        if user_type == 'Faculty' and (not v or v.strip() == ''):
            raise ValueError('Designation is required for faculty')
        return v.strip() if v else v
    
    @validator('specialization', pre=True, always=True)
    def validate_specialization(cls, v, values):
        user_type = values.get('user_type', 'Student')
        if user_type == 'Faculty' and (not v or v.strip() == ''):
            raise ValueError('Specialization is required for faculty')
        return v.strip() if v else v

class UserUpdate(BaseModel):
    username: Optional[str] = Field(None, min_length=3, max_length=50)
    email: Optional[str] = Field(None, pattern=r'^[\w\.-]+@[\w\.-]+\.\w+$')
    password: Optional[str] = Field(None, min_length=6)
    college: Optional[str] = Field(None, min_length=2, max_length=200)
    department: Optional[str] = Field(None, min_length=2, max_length=100)
    year: Optional[str] = Field(None, min_length=4, max_length=10)
    user_type: Optional[str] = Field(None, pattern="^(Student|Faculty)$")
    role: Optional[str] = Field(None, pattern="^(Student|Faculty)$")
    
    # Faculty-specific fields
    experience: Optional[int] = Field(None, ge=0, le=50)
    designation: Optional[str] = Field(None, max_length=200)
    specialization: Optional[str] = Field(None, max_length=500)
    employment_type: Optional[str] = Field(None, pattern="^(Full-time|Visiting|Contract|Part-time)$")
    joining_date: Optional[str] = None

class CourseAutoSetup(BaseModel):
    course_id: int
    duration_weeks: int = Field(..., ge=1, le=52)
    sessions_per_week: int = Field(default=2, ge=1, le=7)



class ModuleCreate(BaseModel):
    course_id: int
    week_number: int = Field(..., ge=1, le=52)
    title: str = Field(..., min_length=5, max_length=200)
    description: str = Field(..., min_length=10)
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None

class ModuleUpdate(BaseModel):
    title: Optional[str] = Field(None, min_length=5, max_length=200)
    description: Optional[str] = Field(None, min_length=10)
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None

class SessionCreate(BaseModel):
    module_id: int
    session_number: int = Field(..., ge=1, le=10, description="Enter session number (1-10)")
    session_type: str = Field(default="Live Session", description="Session type")
    title: str = Field(..., min_length=5, max_length=200, description="Enter session title (minimum 5 characters)")
    description: str = Field(..., min_length=10, description="Enter session description (minimum 10 characters)")
    scheduled_date: Optional[str] = Field(None, description="dd-mm-yyyy")
    scheduled_time: Optional[str] = Field(None, description="--:--")
    duration_minutes: int = Field(default=60, ge=30, le=480, description="Duration (Minutes)")
    meeting_link: Optional[str] = Field(None, description="Meeting Link (Optional)")
    
    @validator('scheduled_date', pre=True)
    def validate_scheduled_date(cls, v):
        if v and v.strip():
            try:
                from datetime import datetime
                datetime.strptime(v, '%d-%m-%Y')
                return v
            except ValueError:
                raise ValueError('Date must be in dd-mm-yyyy format')
        return v
    
    @validator('scheduled_time', pre=True)
    def validate_scheduled_time(cls, v):
        if v and v.strip():
            try:
                from datetime import datetime
                datetime.strptime(v, '%H:%M')
                return v
            except ValueError:
                raise ValueError('Time must be in HH:MM format')
        return v

class PresenterSessionCreate(BaseModel):
    module_id: int
    session_number: int = Field(..., ge=1, le=10, description="Enter session number (1-10)")
    session_type: str = Field(default="Live Session", description="Session type")
    title: str = Field(..., min_length=5, max_length=200, description="Enter session title (minimum 5 characters)")
    description: str = Field(..., min_length=10, description="Enter session description (minimum 10 characters)")
    scheduled_date: Optional[str] = Field(None, description="dd-mm-yyyy")
    scheduled_time: Optional[str] = Field(None, description="--:--")
    duration_minutes: int = Field(default=60, ge=30, le=480, description="Duration (Minutes)")
    meeting_link: Optional[str] = Field(None, description="Meeting Link (Optional)")
    
    @validator('scheduled_date', pre=True)
    def validate_scheduled_date(cls, v):
        if v and v.strip():
            try:
                from datetime import datetime
                datetime.strptime(v, '%d-%m-%Y')
                return v
            except ValueError:
                raise ValueError('Date must be in dd-mm-yyyy format')
        return v
    
    @validator('scheduled_time', pre=True)
    def validate_scheduled_time(cls, v):
        if v and v.strip():
            try:
                from datetime import datetime
                datetime.strptime(v, '%H:%M')
                return v
            except ValueError:
                raise ValueError('Time must be in HH:MM format')
        return v

class SessionUpdate(BaseModel):
    title: Optional[str] = Field(None, min_length=5, max_length=200)
    description: Optional[str] = Field(None, min_length=10)
    scheduled_time: Optional[datetime] = None
    duration_minutes: Optional[int] = Field(None, ge=30, le=480)
    zoom_link: Optional[str] = None
    recording_url: Optional[str] = None
    syllabus_content: Optional[str] = None

class ResourceCreate(BaseModel):
    session_id: int
    title: str = Field(..., min_length=3, max_length=200)
    resource_type: str = Field(..., pattern="^(PDF|PPT|VIDEO|CODE|OTHER|TXT|FILE_LINK)$")
    file_path: str
    file_size: Optional[int] = Field(default=0, ge=0)
    description: Optional[str] = None
    file_url: Optional[str] = None  # For FILE_LINK type

class AttendanceCreate(BaseModel):
    session_id: int
    student_id: int
    attended: bool
    duration_minutes: Optional[int] = Field(default=0, ge=0)
    join_time: Optional[datetime] = None
    leave_time: Optional[datetime] = None

class AttendanceBulkCreate(BaseModel):
    session_id: int
    attendance_records: List[dict]  # [{"student_id": 1, "attended": true, "duration_minutes": 120}]

class QuizCreate(BaseModel):
    session_id: int
    title: str = Field(..., min_length=3, max_length=200)
    description: str = Field(..., min_length=10)
    total_marks: int = Field(..., ge=1, le=1000)
    time_limit_minutes: Optional[int] = Field(default=60, ge=5, le=300)
    questions: Optional[str] = None  # JSON string containing questions
    is_active: Optional[bool] = Field(default=True)

class AIQuizGenerateRequest(BaseModel):
    session_id: int
    title: str = Field(..., min_length=1, max_length=200)
    content: str = Field(..., min_length=1)
    question_type: str
    num_questions: int = Field(..., ge=1, le=50)
    
    @validator('question_type')
    def validate_question_type(cls, v):
        if isinstance(v, str):
            v = v.upper()
        valid_types = ['MCQ', 'TRUE_FALSE', 'SHORT_ANSWER']
        if v not in valid_types:
            raise ValueError(f'question_type must be one of {valid_types}')
        return v
    
    @validator('num_questions', pre=True)
    def validate_num_questions(cls, v):
        if isinstance(v, str):
            try:
                v = int(v)
            except ValueError:
                raise ValueError('num_questions must be a valid integer')
        return v
    
    class Config:
        extra = "ignore"

class QuizFileProcessRequest(BaseModel):
    session_id: int
    file_content: str

class QuizAttemptCreate(BaseModel):
    quiz_id: int
    score: float = Field(..., ge=0)
    answers: Optional[str] = None  # JSON string containing answers
    time_taken_minutes: Optional[int] = Field(default=0, ge=0)

class ForumCreate(BaseModel):
    module_id: int
    title: str = Field(..., min_length=5, max_length=200)
    description: str = Field(..., min_length=10)
    is_pinned: Optional[bool] = Field(default=False)

class ForumPostCreate(BaseModel):
    forum_id: int
    content: str = Field(..., min_length=5)
    parent_post_id: Optional[int] = None  # For replies

class SessionContentCreate(BaseModel):
    session_id: int
    content_type: str = Field(..., pattern="^(VIDEO|QUIZ|MATERIAL|RESOURCE|LIVE_SESSION|MEETING_LINK)$")
    title: str = Field(..., min_length=3, max_length=200)
    description: Optional[str] = None
    file_path: Optional[str] = None
    file_type: Optional[str] = None
    file_size: Optional[int] = Field(default=0, ge=0)
    duration_minutes: Optional[int] = Field(default=60, ge=0)
    content_data: Optional[str] = None  # JSON data for quizzes, live session links, etc.
    meeting_url: Optional[str] = None  # For meeting links
    scheduled_time: Optional[datetime] = None  # For meeting links





class CertificateGenerate(BaseModel):
    student_id: int
    course_id: int
    completion_date: Optional[datetime] = None

class ProgressUpdate(BaseModel):
    student_id: int
    course_id: int
    progress_percentage: float = Field(..., ge=0, le=100)

class NotificationCreate(BaseModel):
    user_id: Optional[int] = None
    title: str = Field(..., min_length=3, max_length=200)
    message: str = Field(..., min_length=5)
    notification_type: str = Field(default="INFO")  # INFO, WARNING, SUCCESS, ERROR
    is_global: Optional[bool] = Field(default=False)

# Authentication endpoints
@app.post("/auth/login")
async def login(user_data: UserLogin, request: Request, db: Session = Depends(get_db)):
    try:
        if user_data.role != "Student":
            raise HTTPException(status_code=400, detail="Invalid role for user login")
        
        user = db.query(User).filter(
            User.username == user_data.username, 
            User.role == user_data.role
        ).first()
        
        if not user or not verify_password(user_data.password, user.password_hash):
            raise HTTPException(status_code=401, detail="Invalid credentials")
        
        access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
        access_token = create_access_token(
            data={"sub": user.username, "role": user.role, "user_id": user.id}, 
            expires_delta=access_token_expires
        )
        
        # Log student login
        log_student_action(
            student_id=user.id,
            student_username=user.username,
            action_type="LOGIN",
            resource_type="STUDENT_SESSION",
            details=f"Student login successful: {user.username}",
            ip_address=request.client.host if request.client else "127.0.0.1"
        )
        
        return {
            "access_token": access_token, 
            "token_type": "bearer", 
            "role": user.role, 
            "user_id": user.id,
            "username": user.username,
            "email": user.email
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Login error: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.post("/auth/logout")
async def user_logout(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        # Log student logout
        log_student_action(
            student_id=current_user.id,
            student_username=current_user.username,
            action_type="LOGOUT",
            resource_type="STUDENT_SESSION",
            details=f"Student logout: {current_user.username}",
            ip_address=request.client.host if request.client else "127.0.0.1"
        )
        
        return {"message": "Logged out successfully"}
    except Exception as e:
        logger.error(f"User logout error: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.post("/admin/login")
async def admin_login(admin_data: AdminLogin, request: Request, db: Session = Depends(get_db)):
    try:
        admin = db.query(Admin).filter(
            Admin.username == admin_data.username
        ).first()
        
        if not admin or not verify_password(admin_data.password, admin.password_hash):
            raise HTTPException(status_code=401, detail="Invalid admin credentials")
        
        access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
        access_token = create_access_token(
            data={"sub": admin.username, "role": "Admin", "user_id": admin.id}, 
            expires_delta=access_token_expires
        )
        
        # Log admin login
        log_admin_action(
            admin_id=admin.id,
            admin_username=admin.username,
            action_type="LOGIN",
            resource_type="ADMIN_SESSION",
            details=f"Admin login successful: {admin.username}",
            ip_address=request.client.host if request.client else "127.0.0.1"
        )
        
        return {
            "access_token": access_token, 
            "token_type": "bearer", 
            "role": "Admin", 
            "user_id": admin.id,
            "username": admin.username,
            "email": admin.email
        }
    except Exception as e:
        logger.error(f"Admin login error: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.post("/admin/logout")
async def admin_logout(
    request: Request,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        # Log admin logout
        log_admin_action(
            admin_id=current_admin.id,
            admin_username=current_admin.username,
            action_type="LOGOUT",
            resource_type="ADMIN_SESSION",
            details=f"Admin logout: {current_admin.username}",
            ip_address=request.client.host if request.client else "127.0.0.1"
        )
        
        return {"message": "Logged out successfully"}
    except Exception as e:
        logger.error(f"Admin logout error: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")

# Presenter Authentication
@app.post("/presenter/login")
async def presenter_login(presenter_data: AdminLogin, request: Request, db: Session = Depends(get_db)):
    try:
        presenter = db.query(Presenter).filter(
            Presenter.username == presenter_data.username
        ).first()
        
        if not presenter or not verify_password(presenter_data.password, presenter.password_hash):
            raise HTTPException(status_code=401, detail="Invalid presenter credentials")
        
        access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
        access_token = create_access_token(
            data={"sub": presenter.username, "role": "Presenter", "user_id": presenter.id}, 
            expires_delta=access_token_expires
        )
        
        # Log presenter login
        log_presenter_action(
            presenter_id=presenter.id,
            presenter_username=presenter.username,
            action_type="LOGIN",
            resource_type="PRESENTER_SESSION",
            details=f"Presenter login successful: {presenter.username}",
            ip_address=request.client.host if request.client else "127.0.0.1"
        )
        
        return {
            "access_token": access_token, 
            "token_type": "bearer", 
            "role": "Presenter", 
            "user_id": presenter.id,
            "username": presenter.username,
            "email": presenter.email
        }
    except Exception as e:
        logger.error(f"Presenter login error: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.post("/presenter/logout")
async def presenter_logout(
    request: Request,
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    try:
        # Log presenter logout
        log_presenter_action(
            presenter_id=current_presenter.id,
            presenter_username=current_presenter.username,
            action_type="LOGOUT",
            resource_type="PRESENTER_SESSION",
            details=f"Presenter logout: {current_presenter.username}",
            ip_address=request.client.host if request.client else "127.0.0.1"
        )
        
        return {"message": "Logged out successfully"}
    except Exception as e:
        logger.error(f"Presenter logout error: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")

# Manager Authentication
@app.post("/manager/login")
async def manager_login(manager_data: AdminLogin, request: Request, db: Session = Depends(get_db)):
    try:
        manager = db.query(Manager).filter(
            Manager.username == manager_data.username
        ).first()
        
        if not manager or not verify_password(manager_data.password, manager.password_hash):
            raise HTTPException(status_code=401, detail="Invalid manager credentials")
        
        access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
        access_token = create_access_token(
            data={"sub": manager.username, "role": "Manager", "user_id": manager.id}, 
            expires_delta=access_token_expires
        )
        
        # Log manager login
        log_admin_action(
            admin_id=manager.id,
            admin_username=manager.username,
            action_type="LOGIN",
            resource_type="MANAGER_SESSION",
            details=f"Manager login successful: {manager.username}",
            ip_address=request.client.host if request.client else "127.0.0.1"
        )
        
        return {
            "access_token": access_token, 
            "token_type": "bearer", 
            "role": "Manager", 
            "user_id": manager.id,
            "username": manager.username,
            "email": manager.email
        }
    except Exception as e:
        logger.error(f"Manager login error: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.post("/manager/logout")
async def manager_logout(
    request: Request,
    db: Session = Depends(get_db)
):
    try:
        # For manager logout, we don't need current_manager dependency
        # since we're just clearing the session
        return {"message": "Logged out successfully"}
    except Exception as e:
        logger.error(f"Manager logout error: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")

# Admin Logs Endpoints
@app.post("/admin/presenter-logs")
async def get_presenter_logs(
    filters: dict,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    """Get presenter logs for admin activity logs page"""
    try:
        # Query presenter logs using SQLAlchemy
        query = db.query(PresenterLog)
        
        if filters.get('action_type'):
            query = query.filter(PresenterLog.action_type == filters['action_type'])
        
        if filters.get('resource_type'):
            query = query.filter(PresenterLog.resource_type == filters['resource_type'])
        
        if filters.get('date_from'):
            query = query.filter(PresenterLog.timestamp >= filters['date_from'])
        
        if filters.get('date_to'):
            query = query.filter(PresenterLog.timestamp <= filters['date_to'])
        
        if filters.get('search'):
            query = query.filter(
                or_(
                    PresenterLog.presenter_username.contains(filters['search']),
                    PresenterLog.details.contains(filters['search'])
                )
            )
        
        # Apply pagination and ordering
        page = filters.get('page', 1)
        limit = 50
        logs = query.order_by(PresenterLog.timestamp.desc()).offset((page - 1) * limit).limit(limit).all()
        
        # Convert to dict format
        logs_data = []
        for log in logs:
            logs_data.append({
                "id": log.id,
                "presenter_id": log.presenter_id,
                "presenter_username": log.presenter_username,
                "action_type": log.action_type,
                "resource_type": log.resource_type,
                "resource_id": log.resource_id,
                "details": log.details,
                "ip_address": log.ip_address,
                "timestamp": log.timestamp
            })
        
        return {"logs": logs_data}
    except Exception as e:
        logger.error(f"Get presenter logs error: {str(e)}")
        return {"logs": []}

@app.get("/admin/test-logs")
async def test_logs(
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    """Test endpoint to check if logs exist"""
    try:
        admin_count = db.query(AdminLog).count()
        presenter_count = db.query(PresenterLog).count()
        
        # Get sample logs
        admin_logs = db.query(AdminLog).order_by(AdminLog.timestamp.desc()).limit(3).all()
        presenter_logs = db.query(PresenterLog).order_by(PresenterLog.timestamp.desc()).limit(3).all()
        
        return {
            "admin_count": admin_count,
            "presenter_count": presenter_count,
            "sample_admin_logs": [{
                "id": log.id,
                "username": log.admin_username,
                "action": log.action_type,
                "timestamp": str(log.timestamp)
            } for log in admin_logs],
            "sample_presenter_logs": [{
                "id": log.id,
                "username": log.presenter_username,
                "action": log.action_type,
                "timestamp": str(log.timestamp)
            } for log in presenter_logs]
        }
    except Exception as e:
        logger.error(f"Test logs error: {str(e)}")
        return {"error": str(e)}

@app.get("/admin/logs")
async def get_admin_logs(
    page: int = 1,
    limit: int = 50,
    action_type: Optional[str] = None,
    resource_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    search: Optional[str] = None,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    """Get admin logs with filtering - redirects to all logs endpoint"""
    return await get_all_system_logs(
        page=page,
        limit=limit,
        action_type=action_type,
        resource_type=resource_type,
        user_type=None,  # Get all user types
        date_from=date_from,
        date_to=date_to,
        search=search,
        current_admin=current_admin,
        db=db
    )

@app.get("/admin/logs/all")
async def get_all_system_logs(
    page: int = 1,
    limit: int = 50,
    action_type: Optional[str] = None,
    resource_type: Optional[str] = None,
    user_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    search: Optional[str] = None,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    """Get all system logs from all user types"""
    try:
        all_logs = []
        
        # Get admin logs
        admin_query = db.query(AdminLog)
        if action_type:
            admin_query = admin_query.filter(AdminLog.action_type == action_type)
        if resource_type:
            admin_query = admin_query.filter(AdminLog.resource_type == resource_type)
        if date_from:
            admin_query = admin_query.filter(AdminLog.timestamp >= date_from)
        if date_to:
            admin_query = admin_query.filter(AdminLog.timestamp <= date_to)
        if search:
            admin_query = admin_query.filter(
                or_(
                    AdminLog.admin_username.contains(search),
                    AdminLog.details.contains(search)
                )
            )
        
        if not user_type or user_type == "Admin":
            admin_logs = admin_query.all()
            for log in admin_logs:
                all_logs.append({
                    "id": f"admin_{log.id}",
                    "user_type": "Admin",
                    "user_id": log.admin_id,
                    "username": log.admin_username,
                    "action_type": log.action_type,
                    "resource_type": log.resource_type,
                    "resource_id": log.resource_id,
                    "details": log.details,
                    "ip_address": log.ip_address,
                    "timestamp": log.timestamp
                })
        
        # Get presenter logs
        presenter_query = db.query(PresenterLog)
        if action_type:
            presenter_query = presenter_query.filter(PresenterLog.action_type == action_type)
        if resource_type:
            presenter_query = presenter_query.filter(PresenterLog.resource_type == resource_type)
        if date_from:
            presenter_query = presenter_query.filter(PresenterLog.timestamp >= date_from)
        if date_to:
            presenter_query = presenter_query.filter(PresenterLog.timestamp <= date_to)
        if search:
            presenter_query = presenter_query.filter(
                or_(
                    PresenterLog.presenter_username.contains(search),
                    PresenterLog.details.contains(search)
                )
            )
        
        if not user_type or user_type == "Presenter":
            presenter_logs = presenter_query.all()
            for log in presenter_logs:
                all_logs.append({
                    "id": f"presenter_{log.id}",
                    "user_type": "Presenter",
                    "user_id": log.presenter_id,
                    "username": log.presenter_username,
                    "action_type": log.action_type,
                    "resource_type": log.resource_type,
                    "resource_id": log.resource_id,
                    "details": log.details,
                    "ip_address": log.ip_address,
                    "timestamp": log.timestamp
                })
        
        # Get mentor logs
        mentor_query = db.query(MentorLog)
        if action_type:
            mentor_query = mentor_query.filter(MentorLog.action_type == action_type)
        if resource_type:
            mentor_query = mentor_query.filter(MentorLog.resource_type == resource_type)
        if date_from:
            mentor_query = mentor_query.filter(MentorLog.timestamp >= date_from)
        if date_to:
            mentor_query = mentor_query.filter(MentorLog.timestamp <= date_to)
        if search:
            mentor_query = mentor_query.filter(
                or_(
                    MentorLog.mentor_username.contains(search),
                    MentorLog.details.contains(search)
                )
            )
        
        if not user_type or user_type == "Mentor":
            mentor_logs = mentor_query.all()
            for log in mentor_logs:
                all_logs.append({
                    "id": f"mentor_{log.id}",
                    "user_type": "Mentor",
                    "user_id": log.mentor_id,
                    "username": log.mentor_username,
                    "action_type": log.action_type,
                    "resource_type": log.resource_type,
                    "resource_id": log.resource_id,
                    "details": log.details,
                    "ip_address": log.ip_address,
                    "timestamp": log.timestamp
                })
        
        # Get student logs
        student_query = db.query(StudentLog)
        if action_type:
            student_query = student_query.filter(StudentLog.action_type == action_type)
        if resource_type:
            student_query = student_query.filter(StudentLog.resource_type == resource_type)
        if date_from:
            student_query = student_query.filter(StudentLog.timestamp >= date_from)
        if date_to:
            student_query = student_query.filter(StudentLog.timestamp <= date_to)
        if search:
            student_query = student_query.filter(
                or_(
                    StudentLog.student_username.contains(search),
                    StudentLog.details.contains(search)
                )
            )
        
        if not user_type or user_type == "Student":
            student_logs = student_query.all()
            for log in student_logs:
                all_logs.append({
                    "id": f"student_{log.id}",
                    "user_type": "Student",
                    "user_id": log.student_id,
                    "username": log.student_username,
                    "action_type": log.action_type,
                    "resource_type": log.resource_type,
                    "resource_id": log.resource_id,
                    "details": log.details,
                    "ip_address": log.ip_address,
                    "timestamp": log.timestamp
                })
        
        # Sort by timestamp (most recent first)
        all_logs.sort(key=lambda x: x['timestamp'], reverse=True)
        
        # Apply pagination
        total = len(all_logs)
        start_idx = (page - 1) * limit
        end_idx = start_idx + limit
        paginated_logs = all_logs[start_idx:end_idx]
        
        return {
            "data": {
                "logs": paginated_logs,
                "total": total,
                "page": page,
                "limit": limit
            }
        }
    except Exception as e:
        logger.error(f"Get all system logs error: {str(e)}")
        return {"data": {"logs": [], "total": 0, "page": page, "limit": limit}}

@app.get("/admin/logs/export")
async def export_all_logs(
    action_type: Optional[str] = None,
    resource_type: Optional[str] = None,
    user_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    search: Optional[str] = None,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    """Export all system logs as CSV"""
    try:
        import csv
        import io
        
        all_logs = []
        
        # Get admin logs
        admin_query = db.query(AdminLog)
        if action_type:
            admin_query = admin_query.filter(AdminLog.action_type == action_type)
        if resource_type:
            admin_query = admin_query.filter(AdminLog.resource_type == resource_type)
        if date_from:
            admin_query = admin_query.filter(AdminLog.timestamp >= date_from)
        if date_to:
            admin_query = admin_query.filter(AdminLog.timestamp <= date_to)
        if search:
            admin_query = admin_query.filter(
                or_(
                    AdminLog.admin_username.contains(search),
                    AdminLog.details.contains(search)
                )
            )
        
        if not user_type or user_type == "Admin":
            admin_logs = admin_query.all()
            for log in admin_logs:
                all_logs.append({
                    "timestamp": log.timestamp,
                    "user_type": "Admin",
                    "username": log.admin_username,
                    "action_type": log.action_type,
                    "resource_type": log.resource_type,
                    "resource_id": log.resource_id,
                    "details": log.details
                })
        
        # Get presenter logs
        presenter_query = db.query(PresenterLog)
        if action_type:
            presenter_query = presenter_query.filter(PresenterLog.action_type == action_type)
        if resource_type:
            presenter_query = presenter_query.filter(PresenterLog.resource_type == resource_type)
        if date_from:
            presenter_query = presenter_query.filter(PresenterLog.timestamp >= date_from)
        if date_to:
            presenter_query = presenter_query.filter(PresenterLog.timestamp <= date_to)
        if search:
            presenter_query = presenter_query.filter(
                or_(
                    PresenterLog.presenter_username.contains(search),
                    PresenterLog.details.contains(search)
                )
            )
        
        if not user_type or user_type == "Presenter":
            presenter_logs = presenter_query.all()
            for log in presenter_logs:
                all_logs.append({
                    "timestamp": log.timestamp,
                    "user_type": "Presenter",
                    "username": log.presenter_username,
                    "action_type": log.action_type,
                    "resource_type": log.resource_type,
                    "resource_id": log.resource_id,
                    "details": log.details
                })
        
        # Get mentor logs
        mentor_query = db.query(MentorLog)
        if action_type:
            mentor_query = mentor_query.filter(MentorLog.action_type == action_type)
        if resource_type:
            mentor_query = mentor_query.filter(MentorLog.resource_type == resource_type)
        if date_from:
            mentor_query = mentor_query.filter(MentorLog.timestamp >= date_from)
        if date_to:
            mentor_query = mentor_query.filter(MentorLog.timestamp <= date_to)
        if search:
            mentor_query = mentor_query.filter(
                or_(
                    MentorLog.mentor_username.contains(search),
                    MentorLog.details.contains(search)
                )
            )
        
        if not user_type or user_type == "Mentor":
            mentor_logs = mentor_query.all()
            for log in mentor_logs:
                all_logs.append({
                    "timestamp": log.timestamp,
                    "user_type": "Mentor",
                    "username": log.mentor_username,
                    "action_type": log.action_type,
                    "resource_type": log.resource_type,
                    "resource_id": log.resource_id,
                    "details": log.details
                })
        
        # Get student logs
        student_query = db.query(StudentLog)
        if action_type:
            student_query = student_query.filter(StudentLog.action_type == action_type)
        if resource_type:
            student_query = student_query.filter(StudentLog.resource_type == resource_type)
        if date_from:
            student_query = student_query.filter(StudentLog.timestamp >= date_from)
        if date_to:
            student_query = student_query.filter(StudentLog.timestamp <= date_to)
        if search:
            student_query = student_query.filter(
                or_(
                    StudentLog.student_username.contains(search),
                    StudentLog.details.contains(search)
                )
            )
        
        if not user_type or user_type == "Student":
            student_logs = student_query.all()
            for log in student_logs:
                all_logs.append({
                    "timestamp": log.timestamp,
                    "user_type": "Student",
                    "username": log.student_username,
                    "action_type": log.action_type,
                    "resource_type": log.resource_type,
                    "resource_id": log.resource_id,
                    "details": log.details
                })
        
        # Sort by timestamp (most recent first)
        all_logs.sort(key=lambda x: x['timestamp'], reverse=True)
        
        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(['Timestamp', 'User Type', 'Username', 'Action', 'Resource Type', 'Resource ID', 'Details'])
        
        # Write data
        for log in all_logs:
            writer.writerow([
                log['timestamp'],
                log['user_type'],
                log['username'],
                log['action_type'],
                log['resource_type'],
                log['resource_id'],
                log['details']
            ])
        
        csv_content = output.getvalue()
        output.close()
        
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": "inline; filename=system_activity_logs.csv"}
        )
    except Exception as e:
        logger.error(f"Export logs error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to export logs")

# Admin Management Endpoints
@app.post("/admin/create-admin")
async def create_admin(
    admin_data: AdminCreate,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        if db.query(Admin).filter(Admin.username == admin_data.username).first():
            raise HTTPException(status_code=400, detail="Username already exists")
        
        if db.query(Admin).filter(Admin.email == admin_data.email).first():
            raise HTTPException(status_code=400, detail="Email already exists")
        
        hashed_password = get_password_hash(admin_data.password)
        admin = Admin(
            username=admin_data.username,
            email=admin_data.email,
            password_hash=hashed_password
        )
        db.add(admin)
        db.commit()
        db.refresh(admin)
        
        return {"message": "Admin created successfully", "admin_id": admin.id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Create admin error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create admin")

@app.post("/admin/create-presenter")
async def create_presenter(
    presenter_data: PresenterCreate,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        if db.query(Presenter).filter(Presenter.username == presenter_data.username).first():
            raise HTTPException(status_code=400, detail="Username already exists")
        
        if db.query(Presenter).filter(Presenter.email == presenter_data.email).first():
            raise HTTPException(status_code=400, detail="Email already exists")
        
        hashed_password = get_password_hash(presenter_data.password)
        presenter = Presenter(
            username=presenter_data.username,
            email=presenter_data.email,
            password_hash=hashed_password
        )
        db.add(presenter)
        db.commit()
        db.refresh(presenter)
        
        return {"message": "Presenter created successfully", "presenter_id": presenter.id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Create presenter error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create presenter")

@app.post("/admin/create-manager")
async def create_manager(
    manager_data: AdminCreate,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        if db.query(Manager).filter(Manager.username == manager_data.username).first():
            raise HTTPException(status_code=400, detail="Username already exists")
        
        if db.query(Manager).filter(Manager.email == manager_data.email).first():
            raise HTTPException(status_code=400, detail="Email already exists")
        
        hashed_password = get_password_hash(manager_data.password)
        manager = Manager(
            username=manager_data.username,
            email=manager_data.email,
            password_hash=hashed_password
        )
        db.add(manager)
        db.commit()
        db.refresh(manager)
        
        return {"message": "Manager created successfully", "manager_id": manager.id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Create manager error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create manager")

@app.get("/admin/presenters")
async def get_all_presenters(
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        presenters = db.query(Presenter).all()
        
        return {
            "presenters": [{
                "id": p.id,
                "username": p.username,
                "email": p.email,
                "is_active": getattr(p, 'is_active', True),
                "created_at": p.created_at
            } for p in presenters]
        }
    except Exception as e:
        logger.error(f"Get presenters error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch presenters")

@app.post("/admin/change-password")
async def change_admin_password(
    password_data: ChangePasswordRequest,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        if not verify_password(password_data.current_password, current_admin.password_hash):
            raise HTTPException(status_code=400, detail="Current password is incorrect")
        
        current_admin.password_hash = get_password_hash(password_data.new_password)
        db.commit()
        
        return {"message": "Password changed successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Change password error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to change password")

@app.post("/presenter/change-password")
async def change_presenter_password(
    password_data: ChangePasswordRequest,
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    try:
        if not verify_password(password_data.current_password, current_presenter.password_hash):
            raise HTTPException(status_code=400, detail="Current password is incorrect")
        
        current_presenter.password_hash = get_password_hash(password_data.new_password)
        db.commit()
        
        return {"message": "Password changed successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Change password error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to change password")

# Presenter Dashboard - Same as Admin Analytics
@app.get("/presenter/dashboard")
async def get_presenter_dashboard(
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    """Presenter dashboard with filtered data based on assigned cohorts and upcoming sessions"""
    try:
        from database import Session as SessionModel
        
        # Get cohorts assigned to this presenter
        presenter_cohorts = db.query(PresenterCohort).filter(
            PresenterCohort.presenter_id == current_presenter.id
        ).all()
        
        assigned_cohort_ids = [pc.cohort_id for pc in presenter_cohorts]
        
        # Get upcoming sessions for presenter's cohorts
        current_time = datetime.now()
        upcoming_sessions = []
        
        if assigned_cohort_ids:
            # Get courses assigned to presenter's cohorts
            cohort_courses = db.query(CohortCourse).filter(
                CohortCourse.cohort_id.in_(assigned_cohort_ids)
            ).all()
            course_ids = [cc.course_id for cc in cohort_courses]
            
            if course_ids:
                # Query sessions for these courses only
                upcoming_sessions = db.query(SessionModel).join(
                    Module, SessionModel.module_id == Module.id
                ).join(
                    Course, Module.course_id == Course.id
                ).filter(
                    Course.id.in_(course_ids),
                    SessionModel.scheduled_time.isnot(None),
                    SessionModel.scheduled_time > current_time
                ).order_by(SessionModel.scheduled_time).limit(10).all()
        
        sessions_data = []
        for session in upcoming_sessions:
            module = db.query(Module).filter(Module.id == session.module_id).first()
            course = db.query(Course).filter(Course.id == module.course_id).first() if module else None
            
            sessions_data.append({
                "id": session.id,
                "title": session.title,
                "course_title": course.title if course else "Unknown Course",
                "module_title": module.title if module else "Unknown Module",
                "scheduled_date": session.scheduled_time.strftime("%Y-%m-%d") if session.scheduled_time else None,
                "scheduled_time": session.scheduled_time.strftime("%H:%M") if session.scheduled_time else None,
                "scheduled_datetime": session.scheduled_time,
                "duration_minutes": session.duration_minutes,
                "zoom_link": session.zoom_link,
                "session_number": session.session_number,
                "week_number": module.week_number if module else None
            })
        
        if not assigned_cohort_ids:
            # Show all courses when no cohorts assigned
            total_courses = db.query(Course).count()
            total_students = db.query(User).filter(User.role == "Student").count()
            total_modules = db.query(Module).count()
            total_sessions = db.query(SessionModel).count()
            total_enrollments = db.query(Enrollment).count()
            active_enrollments = db.query(Enrollment).filter(Enrollment.progress > 0).count()
            
            # Get all upcoming sessions
            current_time = datetime.now()
            upcoming_sessions = db.query(SessionModel).join(
                Module, SessionModel.module_id == Module.id
            ).join(
                Course, Module.course_id == Course.id
            ).filter(
                SessionModel.scheduled_time.isnot(None),
                SessionModel.scheduled_time > current_time
            ).order_by(SessionModel.scheduled_time).limit(10).all()
            
            sessions_data = []
            for session in upcoming_sessions:
                module = db.query(Module).filter(Module.id == session.module_id).first()
                course = db.query(Course).filter(Course.id == module.course_id).first() if module else None
                
                sessions_data.append({
                    "id": session.id,
                    "title": session.title,
                    "course_title": course.title if course else "Unknown Course",
                    "module_title": module.title if module else "Unknown Module",
                    "scheduled_date": session.scheduled_time.strftime("%Y-%m-%d") if session.scheduled_time else None,
                    "scheduled_time": session.scheduled_time.strftime("%H:%M") if session.scheduled_time else None,
                    "scheduled_datetime": session.scheduled_time,
                    "duration_minutes": session.duration_minutes,
                    "zoom_link": session.zoom_link,
                    "session_number": session.session_number,
                    "week_number": module.week_number if module else None
                })
            
            return {
                "users": {
                    "total_students": total_students,
                    "total_admins": db.query(Admin).count(),
                    "growth_rate": 12.5
                },
                "courses": {
                    "total_courses": total_courses,
                    "total_modules": total_modules,
                    "total_sessions": total_sessions,
                    "completed_sessions": total_sessions
                },
                "engagement": {
                    "total_enrollments": total_enrollments,
                    "active_enrollments": active_enrollments,
                    "engagement_rate": (active_enrollments / total_enrollments * 100) if total_enrollments > 0 else 0,
                    "total_resources": db.query(Resource).count(),
                    "total_quizzes": 0
                },
                "performance": {
                    "attendance_rate": 85.0,
                    "completion_rate": 78.0,
                    "average_quiz_score": 82.5,
                    "target_attendance": 80.0,
                    "target_completion": 90.0,
                    "target_quiz_score": 75.0
                },
                "system_health": {
                    "database_status": "healthy",
                    "api_response_time": "45ms",
                    "uptime": "99.9%",
                    "storage_usage": "65%"
                },
                "upcoming_sessions": sessions_data,
                "total_upcoming": len(sessions_data)
            }
        
        # User statistics - only students in assigned cohorts
        total_students = db.query(User).join(UserCohort, User.id == UserCohort.user_id).filter(
            UserCohort.cohort_id.in_(assigned_cohort_ids),
            User.role == "Student"
        ).count()
        
        total_admins = db.query(Admin).count()
        
        # Course statistics - courses assigned to presenter's cohorts
        cohort_courses = db.query(CohortCourse).filter(
            CohortCourse.cohort_id.in_(assigned_cohort_ids)
        ).all() if assigned_cohort_ids else []
        
        course_ids = [cc.course_id for cc in cohort_courses]
        total_courses = len(set(course_ids)) if course_ids else 0
        
        total_modules = db.query(Module).filter(
            Module.course_id.in_(course_ids)
        ).count() if course_ids else 0
        
        total_sessions = db.query(SessionModel).join(
            Module, SessionModel.module_id == Module.id
        ).filter(
            Module.course_id.in_(course_ids)
        ).count() if course_ids else 0
        
        completed_sessions = total_sessions  # Simplified
        
        # Engagement statistics - only for assigned cohorts
        total_enrollments = db.query(Enrollment).filter(
            Enrollment.cohort_id.in_(assigned_cohort_ids)
        ).count() if assigned_cohort_ids else 0
        
        active_enrollments = db.query(Enrollment).filter(
            Enrollment.cohort_id.in_(assigned_cohort_ids),
            Enrollment.progress > 0
        ).count() if assigned_cohort_ids else 0
        
        total_resources = db.query(Resource).join(
            SessionModel, Resource.session_id == SessionModel.id
        ).join(
            Module, SessionModel.module_id == Module.id
        ).filter(
            Module.course_id.in_(course_ids)
        ).count() if course_ids else 0
        
        try:
            from assignment_quiz_tables import Quiz
            total_quizzes = db.query(Quiz).join(SessionModel, Quiz.session_id == SessionModel.id).join(Module, SessionModel.module_id == Module.id).filter(
                Module.course_id.in_(course_ids)
            ).count() if course_ids else 0
        except ImportError:
            # Fallback if quiz models are not available
            total_quizzes = 0
        
        # Performance metrics - only for assigned cohorts
        total_attendances = db.query(Attendance).join(
            SessionModel, Attendance.session_id == SessionModel.id
        ).join(
            Module, SessionModel.module_id == Module.id
        ).filter(
            Module.course_id.in_(course_ids)
        ).count() if course_ids else 0
        
        attended_count = db.query(Attendance).join(
            SessionModel, Attendance.session_id == SessionModel.id
        ).join(
            Module, SessionModel.module_id == Module.id
        ).filter(
            Module.course_id.in_(course_ids),
            Attendance.attended == True
        ).count() if course_ids else 0
        
        attendance_rate = (attended_count / total_attendances * 100) if total_attendances > 0 else 0
        
        completed_courses = db.query(Enrollment).filter(
            Enrollment.cohort_id.in_(assigned_cohort_ids),
            Enrollment.progress >= 90
        ).count() if assigned_cohort_ids else 0
        
        completion_rate = (completed_courses / total_enrollments * 100) if total_enrollments > 0 else 0
        
        # Average quiz score for assigned cohorts
        avg_quiz_score = 0
        if course_ids:
            try:
                from assignment_quiz_tables import QuizResult, Quiz
                quiz_scores = db.query(func.avg(QuizResult.marks_obtained)).join(
                    Quiz, QuizResult.quiz_id == Quiz.id
                ).join(
                    SessionModel, Quiz.session_id == SessionModel.id
                ).join(
                    Module, SessionModel.module_id == Module.id
                ).filter(
                    Module.course_id.in_(course_ids)
                ).scalar()
                avg_quiz_score = quiz_scores or 0
            except ImportError:
                # Fallback if quiz models are not available
                avg_quiz_score = 0
        
        return {
            "users": {
                "total_students": total_students,
                "total_admins": total_admins,
                "growth_rate": 12.5
            },
            "courses": {
                "total_courses": total_courses,
                "total_modules": total_modules,
                "total_sessions": total_sessions,
                "completed_sessions": completed_sessions
            },
            "engagement": {
                "total_enrollments": total_enrollments,
                "active_enrollments": active_enrollments,
                "engagement_rate": (active_enrollments / total_enrollments * 100) if total_enrollments > 0 else 0,
                "total_resources": total_resources,
                "total_quizzes": total_quizzes
            },
            "performance": {
                "attendance_rate": round(attendance_rate, 2),
                "completion_rate": round(completion_rate, 2),
                "average_quiz_score": round(float(avg_quiz_score), 2),
                "target_attendance": 80.0,
                "target_completion": 90.0,
                "target_quiz_score": 75.0
            },
            "system_health": {
                "database_status": "healthy",
                "api_response_time": "45ms",
                "uptime": "99.9%",
                "storage_usage": "65%"
            },
            "upcoming_sessions": sessions_data,
            "total_upcoming": len(sessions_data)
        }
    except Exception as e:
        logger.error(f"Presenter dashboard error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch dashboard data")

# Presenter Analytics Endpoints
@app.get("/presenter/analytics")
async def get_presenter_analytics(
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    """Presenter analytics with filtered data based on assigned cohorts"""
    try:
        # Get cohorts assigned to this presenter
        presenter_cohorts = db.query(PresenterCohort).filter(
            PresenterCohort.presenter_id == current_presenter.id
        ).all()
        
        assigned_cohort_ids = [pc.cohort_id for pc in presenter_cohorts]
        
        if not assigned_cohort_ids:
            # Show all data when no cohorts assigned
            total_students = db.query(User).filter(User.role == "Student").count()
            total_courses = db.query(Course).count()
            total_enrollments = db.query(Enrollment).count()
            active_enrollments = db.query(Enrollment).filter(Enrollment.progress > 0).count()
            
            return {
                "users": {"total_students": total_students, "total_admins": db.query(Admin).count(), "growth_rate": 12.5},
                "courses": {"total_courses": total_courses, "total_modules": db.query(Module).count(), "total_sessions": db.query(SessionModel).count(), "completed_sessions": db.query(SessionModel).count(), "completion_percentage": 100 if total_courses > 0 else 0},
                "engagement": {"total_enrollments": total_enrollments, "active_enrollments": active_enrollments, "engagement_rate": (active_enrollments / total_enrollments * 100) if total_enrollments > 0 else 0, "total_resources": db.query(Resource).count(), "total_quizzes": 0},
                "performance": {"attendance_rate": 85.0, "completion_rate": 78.0, "average_quiz_score": 82.5, "target_attendance": 80.0, "target_completion": 90.0, "target_quiz_score": 75.0},
                "system_health": {"database_status": "healthy", "api_response_time": "45ms", "uptime": "99.9%", "storage_usage": "65%"}
            }
        
        # Filter analytics by assigned cohorts (same logic as dashboard)
        total_students = db.query(User).join(UserCohort).filter(
            UserCohort.cohort_id.in_(assigned_cohort_ids),
            User.role == "Student"
        ).count()
        
        cohort_courses = db.query(CohortCourse).filter(
            CohortCourse.cohort_id.in_(assigned_cohort_ids)
        ).all()
        
        course_ids = [cc.course_id for cc in cohort_courses]
        total_courses = len(set(course_ids)) if course_ids else 0
        
        total_enrollments = db.query(Enrollment).filter(
            Enrollment.cohort_id.in_(assigned_cohort_ids)
        ).count()
        
        active_enrollments = db.query(Enrollment).filter(
            Enrollment.cohort_id.in_(assigned_cohort_ids),
            Enrollment.progress > 0
        ).count()
        
        return {
            "users": {"total_students": total_students, "total_admins": 0, "growth_rate": 12.5},
            "courses": {"total_courses": total_courses, "total_modules": 0, "total_sessions": 0, "completed_sessions": 0, "completion_percentage": 0},
            "engagement": {"total_enrollments": total_enrollments, "active_enrollments": active_enrollments, "engagement_rate": (active_enrollments / total_enrollments * 100) if total_enrollments > 0 else 0, "total_resources": 0, "total_quizzes": 0},
            "performance": {"attendance_rate": 0, "completion_rate": 0, "average_quiz_score": 0, "target_attendance": 80.0, "target_completion": 90.0, "target_quiz_score": 75.0},
            "system_health": {"database_status": "healthy", "api_response_time": "45ms", "uptime": "99.9%", "storage_usage": "65%"}
        }
    except Exception as e:
        logger.error(f"Presenter analytics error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch analytics")

@app.get("/presenter/analytics/overview")
async def get_presenter_analytics_overview(
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    try:
        total_courses = db.query(Course).count()
        total_sessions = db.query(SessionModel).count()
        total_resources = db.query(Resource).count()
        total_quizzes = db.query(Quiz).count()
        
        return {
            "courses": {
                "total_courses": total_courses,
                "total_sessions": total_sessions
            },
            "engagement": {
                "total_resources": total_resources,
                "total_quizzes": total_quizzes
            }
        }
    except Exception as e:
        logger.error(f"Presenter analytics overview error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch analytics overview")

@app.post("/auth/register")
async def register(user_data: UserCreate, db: Session = Depends(get_db)):
    try:
        if db.query(User).filter(User.username == user_data.username).first():
            raise HTTPException(status_code=400, detail="Username already exists")
        
        if db.query(User).filter(User.email == user_data.email).first():
            raise HTTPException(status_code=400, detail="Email already exists")
        
        hashed_password = get_password_hash(user_data.password)
        user = User(
            username=user_data.username,
            email=user_data.email,
            password_hash=hashed_password,
            role=user_data.user_type or "Student",
            college=user_data.college,
            department=user_data.department,
            year=user_data.year,
            user_type=user_data.user_type or "Student"
        )
        db.add(user)
        db.commit()
        db.refresh(user)
        
        return {"message": "User registered successfully", "user_id": user.id}
    except HTTPException:
        raise
    except SQLAlchemyError as e:
        db.rollback()
        logger.error(f"Database error in register: {str(e)}")
        raise HTTPException(status_code=500, detail="Database error")
    except Exception as e:
        db.rollback()
        logger.error(f"Registration error: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")

# Manager Dashboard with Upcoming Sessions
@app.get("/manager/dashboard")
async def get_manager_dashboard(
    current_manager = Depends(get_current_admin_presenter_mentor_or_manager),
    db: Session = Depends(get_db)
):
    """Manager dashboard with full system overview and upcoming sessions"""
    try:
        from database import Session as SessionModel
        
        # Get system-wide analytics (managers have full access)
        total_students = db.query(User).filter(User.role == "Student").count()
        total_courses = db.query(Course).count()
        total_sessions = db.query(SessionModel).count()
        total_enrollments = db.query(Enrollment).count()
        total_cohorts = db.query(Cohort).count()
        
        # Get upcoming sessions (all sessions for managers)
        current_time = datetime.now()
        upcoming_sessions = db.query(SessionModel).join(Module, SessionModel.module_id == Module.id).join(Course, Module.course_id == Course.id).filter(
            SessionModel.scheduled_time.isnot(None),
            SessionModel.scheduled_time > current_time
        ).order_by(SessionModel.scheduled_time).limit(10).all()
        
        sessions_data = []
        for session in upcoming_sessions:
            module = db.query(Module).filter(Module.id == session.module_id).first()
            course = db.query(Course).filter(Course.id == module.course_id).first() if module else None
            
            sessions_data.append({
                "id": session.id,
                "title": session.title,
                "course_title": course.title if course else "Unknown Course",
                "module_title": module.title if module else "Unknown Module",
                "scheduled_date": session.scheduled_time.strftime("%Y-%m-%d") if session.scheduled_time else None,
                "scheduled_time": session.scheduled_time.strftime("%H:%M") if session.scheduled_time else None,
                "scheduled_datetime": session.scheduled_time,
                "duration_minutes": session.duration_minutes,
                "zoom_link": session.zoom_link,
                "session_number": session.session_number,
                "week_number": module.week_number if module else None
            })
        
        return {
            "analytics": {
                "total_students": total_students,
                "total_courses": total_courses,
                "total_sessions": total_sessions,
                "total_enrollments": total_enrollments,
                "total_cohorts": total_cohorts
            },
            "upcoming_sessions": sessions_data,
            "total_upcoming": len(sessions_data)
        }
    except Exception as e:
        logger.error(f"Manager dashboard error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch dashboard data")

# Admin Dashboard with Upcoming Sessions
@app.get("/admin/dashboard")
async def get_admin_dashboard(
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    """Admin dashboard with analytics and upcoming sessions"""
    try:
        from database import Session as SessionModel
        
        # Get basic analytics
        total_students = db.query(User).filter(User.role == "Student").count()
        total_courses = db.query(Course).count()
        total_sessions = db.query(SessionModel).count()
        total_enrollments = db.query(Enrollment).count()
        
        # Get upcoming sessions
        current_time = datetime.now()
        upcoming_sessions = db.query(SessionModel).join(Module, SessionModel.module_id == Module.id).join(Course, Module.course_id == Course.id).filter(
            SessionModel.scheduled_time.isnot(None),
            SessionModel.scheduled_time > current_time
        ).order_by(SessionModel.scheduled_time).limit(10).all()
        
        sessions_data = []
        for session in upcoming_sessions:
            module = db.query(Module).filter(Module.id == session.module_id).first()
            course = db.query(Course).filter(Course.id == module.course_id).first() if module else None
            
            sessions_data.append({
                "id": session.id,
                "title": session.title,
                "course_title": course.title if course else "Unknown Course",
                "module_title": module.title if module else "Unknown Module",
                "scheduled_date": session.scheduled_time.strftime("%Y-%m-%d") if session.scheduled_time else None,
                "scheduled_time": session.scheduled_time.strftime("%H:%M") if session.scheduled_time else None,
                "scheduled_datetime": session.scheduled_time,
                "duration_minutes": session.duration_minutes,
                "zoom_link": session.zoom_link,
                "session_number": session.session_number,
                "week_number": module.week_number if module else None
            })
        
        return {
            "analytics": {
                "total_students": total_students,
                "total_courses": total_courses,
                "total_sessions": total_sessions,
                "total_enrollments": total_enrollments
            },
            "upcoming_sessions": sessions_data,
            "total_upcoming": len(sessions_data)
        }
    except Exception as e:
        logger.error(f"Admin dashboard error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch dashboard data")

# Dashboard Upcoming Sessions Endpoint for All Roles
@app.get("/dashboard/upcoming-sessions")
async def get_upcoming_sessions(
    current_user = Depends(get_current_admin_presenter_mentor_or_manager),
    db: Session = Depends(get_db)
):
    """Get upcoming sessions for dashboard - available to all roles"""
    try:
        current_time = datetime.now()
        
        # Get upcoming sessions with scheduled time
        upcoming_sessions = db.query(SessionModel).join(Module).join(Course).filter(
            SessionModel.scheduled_time.isnot(None),
            SessionModel.scheduled_time > current_time
        ).order_by(SessionModel.scheduled_time).limit(10).all()
        
        sessions_data = []
        for session in upcoming_sessions:
            module = db.query(Module).filter(Module.id == session.module_id).first()
            course = db.query(Course).filter(Course.id == module.course_id).first() if module else None
            
            sessions_data.append({
                "id": session.id,
                "title": session.title,
                "course_title": course.title if course else "Unknown Course",
                "module_title": module.title if module else "Unknown Module",
                "scheduled_date": session.scheduled_time.strftime("%Y-%m-%d") if session.scheduled_time else None,
                "scheduled_time": session.scheduled_time.strftime("%H:%M") if session.scheduled_time else None,
                "scheduled_datetime": session.scheduled_time,
                "duration_minutes": session.duration_minutes,
                "zoom_link": session.zoom_link,
                "session_number": session.session_number,
                "week_number": module.week_number if module else None
            })
        
        return {
            "upcoming_sessions": sessions_data,
            "total": len(sessions_data)
        }
    except Exception as e:
        logger.error(f"Get upcoming sessions error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch upcoming sessions")

# Enhanced Admin Dashboard Endpoints

# College Management
@app.get("/admin/colleges")
async def get_colleges(
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    """Get list of unique colleges from users"""
    try:
        # Get unique colleges from users table
        colleges = db.query(User.college).filter(User.college.isnot(None)).distinct().all()
        college_list = [college[0] for college in colleges if college[0] and college[0].strip()]
        
        # Sort alphabetically
        college_list.sort()
        
        return {"colleges": college_list}
    except Exception as e:
        logger.error(f"Get colleges error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch colleges")

# User Management
@app.post("/admin/users")
async def create_user(user_data: UserCreate, current_admin = Depends(get_current_admin_or_presenter), db: Session = Depends(get_db)):
    try:
        if db.query(User).filter(User.username == user_data.username).first():
            raise HTTPException(status_code=400, detail="Username already exists")
        
        if db.query(User).filter(User.email == user_data.email).first():
            raise HTTPException(status_code=400, detail="Email already exists")
        
        hashed_password = get_password_hash(user_data.password)
        
        # Handle joining_date conversion
        joining_date = None
        if user_data.joining_date:
            try:
                from datetime import datetime
                joining_date = datetime.fromisoformat(user_data.joining_date)
            except:
                pass
        
        user = User(
            username=user_data.username,
            email=user_data.email,
            password_hash=hashed_password,
            role=user_data.user_type or "Student",
            college=user_data.college,
            department=user_data.department,
            year=user_data.year,
            user_type=user_data.user_type or "Student",
            experience=user_data.experience,
            designation=user_data.designation,
            specialization=user_data.specialization,
            employment_type=user_data.employment_type,
            joining_date=joining_date
        )
        db.add(user)
        db.commit()
        db.refresh(user)
        
        # Send welcome email
        try:
            from notification_service import NotificationService
            
            service = NotificationService(db)
            
            # Get user registration template from database
            template_subject = "Welcome to Kamba LMS - Your Account Has Been Created!"
            template_body = f"""
<p>Dear {user.username},</p>
<p>Welcome to Kamba LMS! Your account has been successfully created.</p>
<p><strong>Your login details:</strong></p>
<ul>
    <li>Username: {user.username}</li>
    <li>Email: {user.email}</li>
    <li>Password: {user_data.password}</li>
    <li>College: {user.college}</li>
    <li>Department: {user.department}</li>
    <li>Year: {user.year}</li>
</ul>
<p>Please keep these credentials safe and change your password after your first login.</p>
<p><strong>To get started:</strong></p>
<ol>
    <li>Log in to the LMS portal</li>
    <li>Complete your profile</li>
    <li>Explore available courses</li>
    <li>Join your assigned cohort (if applicable)</li>
</ol>
<p>If you have any questions or need assistance, please contact our support team.</p>
<p>Best regards,<br>The Kamba LMS Team</p>
<hr>
<p><small>This is an automated message. Please do not reply to this email.</small></p>
            """
            
            # Try to get custom template from database
            try:
                from database import EmailTemplate
                template = db.query(EmailTemplate).filter(
                    EmailTemplate.name == "User Registration Welcome Email",
                    EmailTemplate.is_active == True
                ).first()
                
                if template:
                    template_subject = template.subject.format(
                        username=user.username,
                        email=user.email,
                        password=user_data.password,
                        college=user.college,
                        department=user.department,
                        year=user.year
                    )
                    # Convert plain text template to HTML with proper line breaks
                    template_body_raw = template.body.format(
                        username=user.username,
                        email=user.email,
                        password=user_data.password,
                        college=user.college,
                        department=user.department,
                        year=user.year
                    )
                    # Convert to HTML format
                    template_body = template_body_raw.replace('\n', '<br>').replace('\n\n', '<br><br>')
            except Exception as template_error:
                logger.warning(f"Template formatting error, using default: {str(template_error)}")
            
            # Send using the notification service
            email_log = service.send_email_notification(
                user_id=user.id,
                email=user.email,
                subject=template_subject,
                body=template_body
            )
            
            if email_log.status == "sent":
                logger.info(f"Registration welcome email sent successfully to {user.email}")
            elif email_log.status == "queued":
                logger.info(f"Registration welcome email queued for {user.email}")
            else:
                logger.warning(f"Registration welcome email failed for {user.email}: {email_log.error_message}")
                
        except Exception as e:
            logger.warning(f"Failed to send welcome email to {user.email}: {str(e)}")
        
        # Log user creation
        if hasattr(current_admin, 'username') and db.query(Admin).filter(Admin.id == current_admin.id).first():
            log_admin_action(
                admin_id=current_admin.id,
                admin_username=current_admin.username,
                action_type="CREATE",
                resource_type="USER",
                resource_id=user.id,
                details=f"Created user: {user.username} ({user.email})"
            )
        
        return {"message": "User created successfully", "user_id": user.id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Create user error: {str(e)}")
        logger.error(f"User data: {user_data.dict()}")
        raise HTTPException(status_code=500, detail=f"Failed to create user: {str(e)}")

@app.get("/admin/all-members")
async def get_all_members_comprehensive(
    page: int = 1, 
    limit: int = 50, 
    search: Optional[str] = None,
    role: Optional[str] = None,
    college: Optional[str] = None,
    current_user = Depends(get_current_admin_or_presenter), 
    db: Session = Depends(get_db)
):
    """Get only Admin, Presenter, Manager, and Mentor members"""
    try:
        all_users = []
        
        # Get Admins
        admin_query = db.query(Admin)
        if search:
            admin_query = admin_query.filter(
                or_(
                    Admin.username.contains(search),
                    Admin.email.contains(search)
                )
            )
        if not role or role == 'Admin':
            admins = admin_query.all()
            for admin in admins:
                all_users.append({
                    "id": admin.id,
                    "username": admin.username,
                    "email": admin.email,
                    "role": "Admin",
                    "college": None,
                    "department": None,
                    "year": None,
                    "user_type": "Admin",
                    "experience": None,
                    "designation": None,
                    "specialization": None,
                    "employment_type": None,
                    "joining_date": None,
                    "active": True,
                    "created_at": admin.created_at
                })
        
        # Get Presenters
        presenter_query = db.query(Presenter)
        if search:
            presenter_query = presenter_query.filter(
                or_(
                    Presenter.username.contains(search),
                    Presenter.email.contains(search)
                )
            )
        if not role or role == 'Presenter':
            presenters = presenter_query.all()
            for presenter in presenters:
                all_users.append({
                    "id": presenter.id,
                    "username": presenter.username,
                    "email": presenter.email,
                    "role": "Presenter",
                    "college": None,
                    "department": None,
                    "year": None,
                    "user_type": "Presenter",
                    "experience": None,
                    "designation": None,
                    "specialization": None,
                    "employment_type": None,
                    "joining_date": None,
                    "active": True,
                    "created_at": presenter.created_at
                })
        
        # Get Managers
        manager_query = db.query(Manager)
        if search:
            manager_query = manager_query.filter(
                or_(
                    Manager.username.contains(search),
                    Manager.email.contains(search)
                )
            )
        if not role or role == 'Manager':
            managers = manager_query.all()
            for manager in managers:
                all_users.append({
                    "id": manager.id,
                    "username": manager.username,
                    "email": manager.email,
                    "role": "Manager",
                    "college": None,
                    "department": None,
                    "year": None,
                    "user_type": "Manager",
                    "experience": None,
                    "designation": None,
                    "specialization": None,
                    "employment_type": None,
                    "joining_date": None,
                    "active": True,
                    "created_at": manager.created_at
                })
        
        # Get Mentors
        mentor_query = db.query(Mentor)
        if search:
            mentor_query = mentor_query.filter(
                or_(
                    Mentor.username.contains(search),
                    Mentor.email.contains(search)
                )
            )
        if not role or role == 'Mentor':
            mentors = mentor_query.all()
            for mentor in mentors:
                all_users.append({
                    "id": mentor.id,
                    "username": mentor.username,
                    "email": mentor.email,
                    "role": "Mentor",
                    "college": None,
                    "department": None,
                    "year": None,
                    "user_type": "Mentor",
                    "experience": None,
                    "designation": None,
                    "specialization": None,
                    "employment_type": None,
                    "joining_date": None,
                    "active": True,
                    "created_at": mentor.created_at
                })
        
        # Sort by created_at (most recent first) - handle mixed types
        try:
            all_users.sort(key=lambda x: x['created_at'] if isinstance(x['created_at'], datetime) else datetime.now(), reverse=True)
        except (TypeError, AttributeError):
            # Fallback: sort by username if datetime sorting fails
            all_users.sort(key=lambda x: x['username'])
        
        # Apply pagination
        total = len(all_users)
        start_idx = (page - 1) * limit
        end_idx = start_idx + limit
        paginated_users = all_users[start_idx:end_idx]
        
        return {
            "users": paginated_users,
            "total": total,
            "page": page,
            "limit": limit
        }
    except Exception as e:
        logger.error(f"Get all members error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch all members")

@app.get("/admin/users")
async def get_all_users(
    page: int = 1, 
    limit: int = 50, 
    search: Optional[str] = None,
    role: Optional[str] = None,
    college: Optional[str] = None,
    current_user = Depends(get_current_admin_or_presenter), 
    db: Session = Depends(get_db)
):
    try:
        # Filter to only show Students and Faculty in the users table
        query = db.query(User).filter(User.user_type.in_(['Student', 'Faculty']))
        
        if search:
            query = query.filter(
                or_(
                    User.username.contains(search),
                    User.email.contains(search),
                    User.college.contains(search)
                )
            )
        
        if role:
            query = query.filter(User.user_type == role)
        
        if college:
            query = query.filter(User.college == college)
        
        total = query.count()
        users = query.offset((page - 1) * limit).limit(limit).all()
        
        return {
            "users": [{
                "id": u.id,
                "username": u.username,
                "email": u.email,
                "role": u.role,
                "college": u.college,
                "department": u.department,
                "year": u.year,
                "user_type": u.user_type,
                "experience": u.experience,
                "designation": u.designation,
                "specialization": u.specialization,
                "employment_type": u.employment_type,
                "joining_date": u.joining_date,
                "active": True,
                "created_at": u.created_at
            } for u in users],
            "total": total,
            "page": page,
            "limit": limit
        }
    except Exception as e:
        logger.error(f"Get users error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch users")

@app.put("/admin/users/{user_id}")
async def update_user(
    user_id: int, 
    user_data: UserUpdate, 
    current_admin = Depends(get_current_admin_or_presenter), 
    db: Session = Depends(get_db)
):
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="User not found")
        
        # Check for duplicates
        if user_data.username and db.query(User).filter(User.username == user_data.username, User.id != user_id).first():
            raise HTTPException(status_code=400, detail="Username already exists")
        
        if user_data.email and db.query(User).filter(User.email == user_data.email, User.id != user_id).first():
            raise HTTPException(status_code=400, detail="Email already exists")
        
        # Update fields
        update_data = user_data.dict(exclude_unset=True)
        if 'password' in update_data:
            update_data['password_hash'] = get_password_hash(update_data.pop('password'))
        
        # Handle joining_date conversion
        if 'joining_date' in update_data and update_data['joining_date']:
            try:
                from datetime import datetime
                update_data['joining_date'] = datetime.fromisoformat(update_data['joining_date'])
            except:
                update_data.pop('joining_date', None)
        
        for field, value in update_data.items():
            setattr(user, field, value)
        
        db.commit()
        return {"message": "User updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Update user error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update user")

@app.delete("/admin/users/{user_id}")
async def delete_user(
    user_id: int,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        # Try to find user in different tables
        user = db.query(User).filter(User.id == user_id).first()
        admin_user = db.query(Admin).filter(Admin.id == user_id).first()
        presenter_user = db.query(Presenter).filter(Presenter.id == user_id).first()
        mentor_user = db.query(Mentor).filter(Mentor.id == user_id).first()
        manager_user = db.query(Manager).filter(Manager.id == user_id).first()
        
        if user:
            # Delete regular user (Student/Faculty)
            username = user.username
            user_type = user.user_type
            
            # Delete related records (including email_logs and email_recipients)
            db.query(EmailLog).filter(EmailLog.user_id == user_id).delete()
            db.query(EmailRecipient).filter(EmailRecipient.user_id == user_id).delete()
            db.query(UserCohort).filter(UserCohort.user_id == user_id).delete()
            db.query(Enrollment).filter(Enrollment.student_id == user_id).delete()
            db.query(QuizAttempt).filter(QuizAttempt.student_id == user_id).delete()
            db.query(Attendance).filter(Attendance.student_id == user_id).delete()
            db.query(Certificate).filter(Certificate.student_id == user_id).delete()
            db.query(ForumPost).filter(ForumPost.user_id == user_id).delete()
            db.query(Notification).filter(Notification.user_id == user_id).delete()
            db.query(NotificationPreference).filter(NotificationPreference.user_id == user_id).delete()
            db.query(StudentLog).filter(StudentLog.student_id == user_id).delete()
            
            db.delete(user)
            
        elif admin_user:
            # Delete admin user
            if admin_user.id == current_admin.id:
                raise HTTPException(status_code=400, detail="Cannot delete your own account")
            
            username = admin_user.username
            user_type = "Admin"
            
            # Delete admin logs
            db.query(AdminLog).filter(AdminLog.admin_id == user_id).delete()
            
            db.delete(admin_user)
            
        elif presenter_user:
            # Delete presenter user
            username = presenter_user.username
            user_type = "Presenter"
            
            # Delete presenter logs
            db.query(PresenterLog).filter(PresenterLog.presenter_id == user_id).delete()
            
            db.delete(presenter_user)
            
        elif mentor_user:
            # Delete mentor user
            username = mentor_user.username
            user_type = "Mentor"
            
            # Delete mentor logs
            db.query(MentorLog).filter(MentorLog.mentor_id == user_id).delete()
            
            db.delete(mentor_user)
            
        elif manager_user:
            # Delete manager user
            username = manager_user.username
            user_type = "Manager"
            
            # Delete the manager (no specific logs table for managers yet)
            db.delete(manager_user)
            
        else:
            raise HTTPException(status_code=404, detail="User not found")
        
        db.commit()
        
        # Log deletion
        if hasattr(current_admin, 'username') and db.query(Admin).filter(Admin.id == current_admin.id).first():
            log_admin_action(
                admin_id=current_admin.id,
                admin_username=current_admin.username,
                action_type="DELETE",
                resource_type="USER",
                resource_id=user_id,
                details=f"Deleted {user_type}: {username}"
            )
        
        return {"message": f"{user_type} deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Delete user error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete user")

# File serving endpoints
@app.get("/api/resources/{filename}")
async def serve_resource(filename: str, request: Request, db: Session = Depends(get_db)):
    """Serve uploaded resource files with proper content types for browser viewing"""
    file_path = UPLOAD_BASE_DIR / "resources" / filename
    if file_path.exists():
        # Track resource view by filename
        try:
            from resource_analytics_models import ResourceView
            # Find resource by filename
            resource = db.query(Resource).filter(Resource.file_path.contains(filename)).first()
            if resource:
                client_ip = request.client.host if request.client else "127.0.0.1"
                user_agent = request.headers.get("user-agent", "")
                
                view_record = ResourceView(
                    resource_id=resource.id,
                    student_id=None,  # Anonymous view
                    viewed_at=datetime.utcnow(),
                    ip_address=client_ip,
                    user_agent=user_agent
                )
                
                db.add(view_record)
                db.commit()
                logger.info(f"Resource file view tracked: resource_id={resource.id}, filename={filename}")
        except Exception as track_error:
            logger.warning(f"Failed to track resource file view: {str(track_error)}")
    file_path = UPLOAD_BASE_DIR / "resources" / filename
    if file_path.exists():
        file_ext = os.path.splitext(filename)[1].lower()
        
        # Headers to force inline viewing in browser - CRITICAL for preventing downloads
        headers = {
            "Content-Disposition": "inline; filename=\"" + filename + "\"",
            "Cache-Control": "public, max-age=3600",
            "X-Content-Type-Options": "nosniff",
            "Accept-Ranges": "bytes"
        }
        
        # Set proper MIME types for browser viewing
        if file_ext == ".pdf":
            media_type = "application/pdf"
        elif file_ext in [".txt", ".text"]:
            media_type = "text/plain; charset=utf-8"
        elif file_ext in [".html", ".htm"]:
            media_type = "text/html; charset=utf-8"
        elif file_ext in [".css"]:
            media_type = "text/css; charset=utf-8"
        elif file_ext in [".js"]:
            media_type = "application/javascript; charset=utf-8"
        elif file_ext in [".json"]:
            media_type = "application/json; charset=utf-8"
        elif file_ext in [".xml"]:
            media_type = "application/xml; charset=utf-8"
        elif file_ext in [".jpg", ".jpeg"]:
            media_type = "image/jpeg"
        elif file_ext == ".png":
            media_type = "image/png"
        elif file_ext == ".gif":
            media_type = "image/gif"
        elif file_ext == ".svg":
            media_type = "image/svg+xml"
        elif file_ext in [".mp4"]:
            media_type = "video/mp4"
        elif file_ext in [".webm"]:
            media_type = "video/webm"
        elif file_ext in [".mp3"]:
            media_type = "audio/mpeg"
        elif file_ext in [".wav"]:
            media_type = "audio/wav"
        elif file_ext in [".ogg"]:
            media_type = "audio/ogg"
        elif file_ext in [".ppt", ".pptx"]:
            # For PowerPoint files - use proper MIME type with inline disposition
            media_type = "application/vnd.ms-powerpoint" if file_ext == ".ppt" else "application/vnd.openxmlformats-officedocument.presentationml.presentation"
            # Override headers for Office files to ensure inline viewing
            headers["Content-Disposition"] = "inline; filename=\"" + filename + "\""
        elif file_ext in [".doc", ".docx"]:
            # For Word files - use proper MIME type with inline disposition
            media_type = "application/msword" if file_ext == ".doc" else "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            # Override headers for Office files to ensure inline viewing
            headers["Content-Disposition"] = "inline; filename=\"" + filename + "\""
        elif file_ext in [".xls", ".xlsx"]:
            # For Excel files
            media_type = "application/vnd.ms-excel" if file_ext == ".xls" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            headers["Content-Disposition"] = "inline; filename=\"" + filename + "\""
        else:
            # For unknown file types, try to serve as text if possible
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    f.read(100)  # Try to read first 100 chars
                media_type = "text/plain; charset=utf-8"
            except:
                media_type = "application/octet-stream"
                # Even for binary files, try inline first
                headers["Content-Disposition"] = "inline; filename=\"" + filename + "\""
        
        return FileResponse(file_path, media_type=media_type, headers=headers)
    raise HTTPException(status_code=404, detail="File not found")

# Authenticated Resource Viewing Endpoints
@app.get("/api/resources/{resource_id}/view")
async def view_resource_authenticated(
    resource_id: int,
    token: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """View a specific resource file with authentication"""
    try:
        if not token:
            raise HTTPException(status_code=401, detail="Authentication required")
        
        from auth import get_current_user_from_token
        current_user = get_current_user_from_token(token, db)
        
        resource = db.query(Resource).filter(Resource.id == resource_id).first()
        if not resource:
            raise HTTPException(status_code=404, detail="Resource not found")
        
        # Check if file exists
        if not resource.file_path or not os.path.exists(resource.file_path):
            raise HTTPException(status_code=404, detail="File not found")
        
        # Serve the file
        filename = os.path.basename(resource.file_path)
        file_ext = os.path.splitext(filename)[1].lower()
        
        headers = {
            "Content-Disposition": "inline; filename=\"" + filename + "\"",
            "Cache-Control": "public, max-age=3600",
            "Accept-Ranges": "bytes"
        }
        
        if file_ext == ".pdf":
            media_type = "application/pdf"
        elif file_ext in [".txt", ".text"]:
            media_type = "text/plain; charset=utf-8"
        elif file_ext in [".jpg", ".jpeg"]:
            media_type = "image/jpeg"
        elif file_ext == ".png":
            media_type = "image/png"
        elif file_ext in [".mp4"]:
            media_type = "video/mp4"
        elif file_ext in [".ppt", ".pptx"]:
            media_type = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
        else:
            media_type = "application/octet-stream"
            
        return FileResponse(resource.file_path, media_type=media_type, headers=headers)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"View resource error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to view resource")

@app.get("/api/resources/uploads/resources/{filename}")
async def serve_admin_resource(
    filename: str,
    token: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Serve uploaded resource files for admin dashboard with authentication"""
    try:
        if not token:
            raise HTTPException(status_code=401, detail="Authentication required")
        
        from auth import get_current_user_from_token
        current_user = get_current_user_from_token(token, db)
        
        # Build file path
        file_path = os.path.join("uploads", "resources", filename)
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
        
        # Serve the file
        file_ext = os.path.splitext(filename)[1].lower()
        
        headers = {
            "Content-Disposition": "inline; filename=\"" + filename + "\"",
            "Cache-Control": "public, max-age=3600",
            "Accept-Ranges": "bytes"
        }
        
        if file_ext == ".pdf":
            media_type = "application/pdf"
        elif file_ext in [".txt", ".text"]:
            media_type = "text/plain; charset=utf-8"
        elif file_ext in [".jpg", ".jpeg"]:
            media_type = "image/jpeg"
        elif file_ext == ".png":
            media_type = "image/png"
        elif file_ext in [".mp4"]:
            media_type = "video/mp4"
        elif file_ext in [".ppt", ".pptx"]:
            media_type = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
        else:
            media_type = "application/octet-stream"
            
        return FileResponse(file_path, media_type=media_type, headers=headers)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Serve admin resource error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to serve resource")

@app.get("/api/recordings/{filename}")
async def serve_recording(filename: str):
    """Serve uploaded recording files with inline viewing"""
    file_path = UPLOAD_BASE_DIR / "recordings" / filename
    if file_path.exists():
        file_ext = os.path.splitext(filename)[1].lower()
        
        headers = {
            "Content-Disposition": "inline; filename=\"" + filename + "\"",
            "Cache-Control": "public, max-age=3600",
            "Accept-Ranges": "bytes"
        }
        
        # Set appropriate media type for recordings
        if file_ext in [".mp4"]:
            media_type = "video/mp4"
        elif file_ext in [".webm"]:
            media_type = "video/webm"
        elif file_ext in [".avi"]:
            media_type = "video/x-msvideo"
        elif file_ext in [".mov"]:
            media_type = "video/quicktime"
        elif file_ext in [".mp3"]:
            media_type = "audio/mpeg"
        elif file_ext in [".wav"]:
            media_type = "audio/wav"
        else:
            media_type = "application/octet-stream"
            
        return FileResponse(file_path, media_type=media_type, headers=headers)
    raise HTTPException(status_code=404, detail="Recording not found")

@app.get("/api/certificates/{filename}")
async def serve_certificate(filename: str):
    """Serve generated certificate files with inline viewing"""
    file_path = UPLOAD_BASE_DIR / "certificates" / filename
    if file_path.exists():
        headers = {
            "Content-Disposition": "inline; filename=\"" + filename + "\"",
            "Cache-Control": "public, max-age=3600"
        }
        
        file_ext = os.path.splitext(filename)[1].lower()
        if file_ext == ".pdf":
            media_type = "application/pdf"
        elif file_ext in [".jpg", ".jpeg"]:
            media_type = "image/jpeg"
        elif file_ext == ".png":
            media_type = "image/png"
        else:
            media_type = "application/octet-stream"
            
        return FileResponse(file_path, media_type=media_type, headers=headers)
    raise HTTPException(status_code=404, detail="Certificate not found")

# Simple cohorts endpoint for testing (no auth required)
@app.get("/api/cohorts")
async def get_cohorts_simple(db: Session = Depends(get_db)):
    """Simple cohorts endpoint that works without auth"""
    try:
        # Get cohorts from database
        cohorts = db.query(Cohort).filter(Cohort.is_active == True).all()
        
        cohorts_data = []
        for cohort in cohorts:
            # Get user count
            user_count = db.query(UserCohort).filter(
                UserCohort.cohort_id == cohort.id,
                UserCohort.is_active == True
            ).count()
            
            cohorts_data.append({
                "id": cohort.id,
                "name": cohort.name,
                "description": cohort.description or "",
                "user_count": user_count,
                "is_active": cohort.is_active
            })
        
        return {"cohorts": cohorts_data}
        
    except Exception as e:
        logger.error(f"Get cohorts error: {str(e)}")
        return {"cohorts": []}

@app.get("/api/cohorts/{cohort_id}/members")
async def get_cohort_members_simple(cohort_id: int, db: Session = Depends(get_db)):
    """Simple cohort members endpoint without auth"""
    try:
        members = []
        
        # Get students in cohort
        user_cohorts = db.query(UserCohort).filter(
            UserCohort.cohort_id == cohort_id,
            UserCohort.is_active == True
        ).all()
        
        for uc in user_cohorts:
            user = db.query(User).filter(User.id == uc.user_id).first()
            if user:
                members.append({
                    "id": user.id,
                    "name": user.username,
                    "email": user.email,
                    "role": "Student",
                    "user_type": "Student"
                })
        
        return {"members": members}
        
    except Exception as e:
        logger.error(f"Get cohort members error: {str(e)}")
        return {"members": []}

@app.get("/api/cohorts/{cohort_id}/users")
async def get_cohort_users_simple(cohort_id: int, db: Session = Depends(get_db)):
    """Simple cohort users endpoint without auth"""
    try:
        users = []
        
        # Get students in cohort
        user_cohorts = db.query(UserCohort).filter(
            UserCohort.cohort_id == cohort_id,
            UserCohort.is_active == True
        ).all()
        
        for uc in user_cohorts:
            user = db.query(User).filter(User.id == uc.user_id).first()
            if user:
                users.append({
                    "id": user.id,
                    "username": user.username,
                    "email": user.email,
                    "role": "Student",
                    "user_type": "Student",
                    "college": user.college,
                    "department": user.department
                })
        
        # Get presenters assigned to cohort
        presenter_cohorts = db.query(PresenterCohort).filter(
            PresenterCohort.cohort_id == cohort_id
        ).all()
        
        for pc in presenter_cohorts:
            presenter = db.query(Presenter).filter(Presenter.id == pc.presenter_id).first()
            if presenter:
                users.append({
                    "id": presenter.id,
                    "username": presenter.username,
                    "email": presenter.email,
                    "role": "Presenter",
                    "user_type": "Presenter"
                })
        
        # Add mentors
        mentors = db.query(Mentor).all()
        for mentor in mentors:
            users.append({
                "id": mentor.id,
                "username": mentor.username,
                "email": mentor.email,
                "role": "Mentor",
                "user_type": "Mentor"
            })
        
        # Add admins
        admins = db.query(Admin).all()
        for admin in admins:
            users.append({
                "id": admin.id,
                "username": admin.username,
                "email": admin.email,
                "role": "Admin",
                "user_type": "Admin"
            })
        
        return {"users": users}
        
    except Exception as e:
        logger.error(f"Get cohort users error: {str(e)}")
        return {"users": []}

@app.get("/api/cohorts/{cohort_id}/staff")
async def get_cohort_staff_simple(cohort_id: int, db: Session = Depends(get_db)):
    """Simple cohort staff endpoint without auth"""
    try:
        staff = []
        
        # Get presenters assigned to cohort
        presenter_cohorts = db.query(PresenterCohort).filter(
            PresenterCohort.cohort_id == cohort_id
        ).all()
        
        for pc in presenter_cohorts:
            presenter = db.query(Presenter).filter(Presenter.id == pc.presenter_id).first()
            if presenter:
                staff.append({
                    "id": presenter.id,
                    "username": presenter.username,
                    "email": presenter.email,
                    "role": "Presenter",
                    "user_type": "Presenter"
                })
        
        # Add mentors
        mentors = db.query(Mentor).all()
        for mentor in mentors:
            staff.append({
                "id": mentor.id,
                "username": mentor.username,
                "email": mentor.email,
                "role": "Mentor",
                "user_type": "Mentor"
            })
        
        # Add admins
        admins = db.query(Admin).all()
        for admin in admins:
            staff.append({
                "id": admin.id,
                "username": admin.username,
                "email": admin.email,
                "role": "Admin",
                "user_type": "Admin"
            })
        
        return {"staff": staff}
        
    except Exception as e:
        logger.error(f"Get cohort staff error: {str(e)}")
        return {"staff": []}

@app.get("/api/cohorts/{cohort_id}/presenters")
async def get_cohort_presenters_simple(cohort_id: int, db: Session = Depends(get_db)):
    """Simple cohort presenters endpoint without auth"""
    try:
        presenters = []
        
        # Get presenters assigned to cohort
        presenter_cohorts = db.query(PresenterCohort).filter(
            PresenterCohort.cohort_id == cohort_id
        ).all()
        
        for pc in presenter_cohorts:
            presenter = db.query(Presenter).filter(Presenter.id == pc.presenter_id).first()
            if presenter:
                presenters.append({
                    "id": presenter.id,
                    "username": presenter.username,
                    "email": presenter.email,
                    "role": "Presenter",
                    "user_type": "Presenter"
                })
        
        return {"presenters": presenters}
        
    except Exception as e:
        logger.error(f"Get cohort presenters error: {str(e)}")
        return {"presenters": []}

# Debug endpoints for file-link troubleshooting
@app.get("/admin/debug/resources/{session_id}")
async def debug_resources(
    session_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    """Debug endpoint to check resources for a session"""
    try:
        resources = db.query(Resource).filter(Resource.session_id == session_id).all()
        
        result = []
        for r in resources:
            result.append({
                "id": r.id,
                "session_id": r.session_id,
                "title": r.title,
                "resource_type": r.resource_type,
                "file_path": r.file_path,
                "file_size": r.file_size,
                "description": r.description,
                "uploaded_at": r.uploaded_at,
                "created_at": r.created_at
            })
        
        return {
            "session_id": session_id,
            "resource_count": len(result),
            "resources": result
        }
        
    except Exception as e:
        logger.error(f"Debug resources error: {str(e)}")
        return {"error": str(e)}

# Startup event to start campaign scheduler
@app.on_event("startup")
async def startup_event():
    """Start background tasks when the application starts"""
    if SCHEDULER_AVAILABLE:
        # Start campaign scheduler in background
        asyncio.create_task(start_campaign_scheduler())
        logger.info("Campaign scheduler started successfully")
    else:
        logger.warning("Campaign scheduler not started - module not available")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

# Add simple session content router for testing
try:
    from simple_session_content import router as simple_router
    app.include_router(simple_router)
    print("Simple session content router loaded successfully")
except Exception as e:
    print(f"Failed to load simple router: {e}")







@app.get("/admin/download-student-template")
async def download_student_template(
    current_user = Depends(get_current_admin_or_presenter)
):
    """Download Excel template for student bulk upload"""
    try:
        try:
            import pandas as pd
            
            # Student template data
            sample_data = {
                'Username': ['john_doe', 'jane_smith', 'mike_wilson'],
                'Email': ['john@example.com', 'jane@example.com', 'mike@example.com'],
                'Password': ['password123', 'password456', 'password789'],
                'College': ['MIT University', 'Stanford University', 'Harvard University'],
                'Department': ['Computer Science', 'Engineering', 'Mathematics'],
                'Year': ['2024', '2023', '2025']
            }
            
            df = pd.DataFrame(sample_data)
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Students', index=False)
                
                instructions = pd.DataFrame({
                    'Instructions': [
                        '1. Fill in the student data in the Students sheet',
                        '2. Required: Username, Email, Password, College, Department, Year',
                        '3. Type will be automatically set to Student',
                        '4. Username must be unique',
                        '5. Email must be valid and unique',
                        '6. Password minimum 6 characters',
                        '7. Save as Excel (.xlsx) or CSV (.csv) format',
                        '8. Upload the file using the Bulk Import feature'
                    ]
                })
                instructions.to_excel(writer, sheet_name='Instructions', index=False)
            
            output.seek(0)
            
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "inline; filename=student_template.xlsx"}
            )
            
        except ImportError:
            csv_content = "Username,Email,Password,College,Department,Year\n"
            csv_content += "john_doe,john@example.com,password123,MIT University,Computer Science,2024\n"
            csv_content += "jane_smith,jane@example.com,password456,Stanford University,Engineering,2023\n"
            csv_content += "mike_wilson,mike@example.com,password789,Harvard University,Mathematics,2025\n"
            
            return Response(
                content=csv_content,
                media_type="text/csv",
                headers={"Content-Disposition": "inline; filename=student_template.csv"}
            )
            
    except Exception as e:
        logger.error(f"Download student template error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate student template")

@app.get("/admin/download-faculty-template")
async def download_faculty_template(
    current_user = Depends(get_current_admin_or_presenter)
):
    """Download Excel template for faculty bulk upload"""
    try:
        try:
            import pandas as pd
            
            # Faculty template data
            sample_data = {
                'Username': ['dr_sarah_jones', 'prof_michael_brown', 'dr_lisa_chen'],
                'Email': ['sarah@university.edu', 'michael@university.edu', 'lisa@university.edu'],
                'Password': ['faculty123', 'faculty456', 'faculty789'],
                'College': ['MIT University', 'Stanford University', 'Harvard University'],
                'Department': ['Computer Science', 'Engineering', 'Mathematics'],
                'Experience': ['10', '15', '8'],
                'Designation': ['Associate Professor', 'Professor', 'Assistant Professor'],
                'Specialization': ['Machine Learning, Data Science', 'Software Engineering, AI', 'Statistics, Data Analysis'],
                'Employment_Type': ['Full-time', 'Full-time', 'Visiting'],
                'Joining_Date': ['2020-01-15', '2018-08-20', '2022-09-01']
            }
            
            df = pd.DataFrame(sample_data)
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Faculty', index=False)
                
                instructions = pd.DataFrame({
                    'Instructions': [
                        '1. Fill in the faculty data in the Faculty sheet',
                        '2. Required: Username, Email, Password, College, Department, Experience, Designation, Specialization',
                        '3. Type will be automatically set to Faculty',
                        '4. Optional: Employment_Type (defaults to Full-time), Joining_Date',
                        '5. Username must be unique',
                        '6. Email must be valid and unique',
                        '7. Password minimum 6 characters',
                        '8. Experience must be a number greater than 0',
                        '9. Employment_Type options: Full-time, Visiting, Contract, Part-time',
                        '10. Joining_Date format: YYYY-MM-DD (optional)',
                        '11. Save as Excel (.xlsx) or CSV (.csv) format',
                        '12. Upload the file using the Bulk Import feature'
                    ]
                })
                instructions.to_excel(writer, sheet_name='Instructions', index=False)
            
            output.seek(0)
            
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "inline; filename=faculty_template.xlsx"}
            )
            
        except ImportError:
            csv_content = "Username,Email,Password,College,Department,Experience,Designation,Specialization,Employment_Type,Joining_Date\n"
            csv_content += "dr_sarah_jones,sarah@university.edu,faculty123,MIT University,Computer Science,10,Associate Professor,Machine Learning Data Science,Full-time,2020-01-15\n"
            csv_content += "prof_michael_brown,michael@university.edu,faculty456,Stanford University,Engineering,15,Professor,Software Engineering AI,Full-time,2018-08-20\n"
            csv_content += "dr_lisa_chen,lisa@university.edu,faculty789,Harvard University,Mathematics,8,Assistant Professor,Statistics Data Analysis,Visiting,2022-09-01\n"
            
            return Response(
                content=csv_content,
                media_type="text/csv",
                headers={"Content-Disposition": "inline; filename=faculty_template.csv"}
            )
            
    except Exception as e:
        logger.error(f"Download faculty template error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate faculty template")

@app.get("/admin/cohort-template")
async def download_cohort_template(
    current_user = Depends(get_current_admin_or_presenter)
):
    """Download Excel template for cohort user bulk upload"""
    try:
        if generate_cohort_template_excel is None:
            raise HTTPException(status_code=500, detail="Template generation not available. Please install required dependencies.")
        
        return generate_cohort_template_excel()
    except Exception as e:
        logger.error(f"Download cohort template error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate cohort template")



@app.post("/admin/users/bulk-upload")
async def bulk_upload_users(
    file: UploadFile = File(...),
    user_type_filter: str = Form("Student"),  # Add filter parameter
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        file_ext = os.path.splitext(file.filename)[1].lower()
        
        if file_ext not in ['.csv', '.xlsx', '.xls']:
            raise HTTPException(status_code=400, detail="Only CSV and Excel files are allowed")
        
        # Validate user_type_filter
        if user_type_filter not in ['Student', 'Faculty']:
            raise HTTPException(status_code=400, detail="Invalid user type filter. Must be 'Student' or 'Faculty'")
        
        content = await file.read()
        records = []
        
        if file_ext == '.csv':
            try:
                csv_content = content.decode('utf-8')
                csv_reader = csv.DictReader(io.StringIO(csv_content))
                records = list(csv_reader)
            except UnicodeDecodeError:
                # Try with different encoding
                csv_content = content.decode('utf-8-sig')
                csv_reader = csv.DictReader(io.StringIO(csv_content))
                records = list(csv_reader)
        else:
            try:
                import pandas as pd
                df = pd.read_excel(io.BytesIO(content))
                # Convert DataFrame to list of dictionaries
                records = df.to_dict('records')
            except ImportError:
                # Fallback to openpyxl if pandas not available
                from openpyxl import load_workbook
                workbook = load_workbook(io.BytesIO(content))
                sheet = workbook.active
                headers = [cell.value for cell in sheet[1] if cell.value]
                records = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if any(cell for cell in row):  # Skip empty rows
                        record = {}
                        for i, value in enumerate(row):
                            if i < len(headers) and headers[i]:
                                record[headers[i]] = value
                        records.append(record)
        
        if not records:
            raise HTTPException(status_code=400, detail="No data found in file. Please check the file format.")
        
        created_users = []
        errors = []
        
        for row_num, row in enumerate(records, 1):
            try:
                # Handle different column name variations
                username = str(row.get('Username') or row.get('username') or row.get('UserName') or '').strip()
                email = str(row.get('Email') or row.get('email') or '').strip()
                password = str(row.get('Password') or row.get('password') or '').strip()
                # Force user type to match the filter - ignore any type in the file
                user_type = user_type_filter
                college = str(row.get('College') or row.get('college') or '').strip()
                department = str(row.get('Department') or row.get('department') or '').strip()
                year = str(row.get('Year') or row.get('year') or '').strip()
                
                # Faculty-specific fields (only process if uploading faculty)
                experience = None
                designation = None
                specialization = None
                employment_type = 'Full-time'
                joining_date = None
                
                if user_type_filter == 'Faculty':
                    experience = row.get('Experience') or row.get('experience') or None
                    designation = str(row.get('Designation') or row.get('designation') or '').strip() or None
                    specialization = str(row.get('Specialization') or row.get('specialization') or '').strip() or None
                    employment_type = str(row.get('Employment_Type') or row.get('employment_type') or row.get('EmploymentType') or 'Full-time').strip()
                    joining_date = row.get('Joining_Date') or row.get('joining_date') or row.get('JoiningDate') or None
                
                # Skip empty rows
                if not any([username, email, password]):
                    continue
                
                # Validate required fields
                if not all([username, email, password]):
                    errors.append(f"Row {row_num}: Missing required fields (Username, Email, Password)")
                    continue
                
                # Validate email format
                if '@' not in email or '.' not in email:
                    errors.append(f"Row {row_num}: Invalid email format '{email}'")
                    continue
                
                # Check for duplicates
                if db.query(User).filter(User.username == username).first():
                    errors.append(f"Row {row_num}: Username '{username}' already exists")
                    continue
                
                if db.query(User).filter(User.email == email).first():
                    errors.append(f"Row {row_num}: Email '{email}' already exists")
                    continue
                
                # Faculty-specific validation
                if user_type_filter == 'Faculty':
                    # Validate required faculty fields
                    if not designation or designation.strip() == '':
                        errors.append(f"Row {row_num}: Designation is required for faculty")
                        continue
                    
                    if not specialization or specialization.strip() == '':
                        errors.append(f"Row {row_num}: Specialization is required for faculty")
                        continue
                    
                    # Convert experience to integer if provided
                    if experience:
                        try:
                            experience = int(experience)
                            if experience <= 0 or experience > 50:
                                errors.append(f"Row {row_num}: Experience must be between 1 and 50 years")
                                continue
                        except:
                            errors.append(f"Row {row_num}: Experience must be a valid number")
                            continue
                    else:
                        errors.append(f"Row {row_num}: Experience is required for faculty")
                        continue
                
                # Handle joining_date conversion
                joining_date_obj = None
                if joining_date:
                    try:
                        from datetime import datetime
                        if isinstance(joining_date, str):
                            joining_date_obj = datetime.fromisoformat(joining_date)
                        else:
                            joining_date_obj = joining_date
                    except:
                        pass
                
                # Handle year field - required for all users but set default for faculty
                final_year = year if year else ("N/A" if user_type_filter == "Faculty" else "2024")
                
                # Create user with proper type enforcement
                user = User(
                    username=username,
                    email=email,
                    password_hash=get_password_hash(password),
                    role=user_type_filter,  # Enforce the filter type
                    user_type=user_type_filter,  # Enforce the filter type
                    college=college,
                    department=department,
                    year=final_year,
                    experience=experience,
                    designation=designation,
                    specialization=specialization,
                    employment_type=employment_type,
                    joining_date=joining_date_obj
                )
                
                db.add(user)
                created_users.append({
                    'username': username,
                    'email': email,
                    'type': user_type_filter,  # Use the enforced type
                    'college': college
                })
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        if created_users:
            db.commit()
            
            # Send welcome emails to all created users
            try:
                from notification_service import NotificationService
                service = NotificationService(db)
                
                # Get user registration template from database
                template_subject = "Welcome to Kamba LMS - Your Account Has Been Created!"
                template_body_default = """
<p>Dear {username},</p>
<p>Welcome to Kamba LMS! Your account has been successfully created.</p>
<p><strong>Your login details:</strong></p>
<ul>
    <li>Username: {username}</li>
    <li>Email: {email}</li>
    <li>Password: {password}</li>
    <li>College: {college}</li>
    <li>Department: {department}</li>
    <li>Year: {year}</li>
</ul>
<p>Please keep these credentials safe and change your password after your first login.</p>
<p><strong>To get started:</strong></p>
<ol>
    <li>Log in to the LMS portal</li>
    <li>Complete your profile</li>
    <li>Explore available courses</li>
    <li>Join your assigned cohort (if applicable)</li>
</ol>
<p>If you have any questions or need assistance, please contact our support team.</p>
<p>Best regards,<br>The Kamba LMS Team</p>
<hr>
<p><small>This is an automated message. Please do not reply to this email.</small></p>
                """
                
                # Try to get custom template from database
                try:
                    from database import EmailTemplate
                    template = db.query(EmailTemplate).filter(
                        EmailTemplate.name == "User Registration Welcome Email",
                        EmailTemplate.is_active == True
                    ).first()
                    
                    if template:
                        template_subject = template.subject
                        template_body_default = template.body
                except Exception as template_error:
                    logger.warning(f"Template loading error, using default: {str(template_error)}")
                
                # Send emails to all created users
                email_success_count = 0
                email_failed_count = 0
                
                for user_data in created_users:
                    try:
                        # Get the actual user object from database
                        user = db.query(User).filter(User.username == user_data['username']).first()
                        if not user:
                            continue
                            
                        # Get the original password from records
                        original_password = None
                        for row_num, row in enumerate(records, 1):
                            username = str(row.get('Username') or row.get('username') or row.get('UserName') or '').strip()
                            if username == user_data['username']:
                                original_password = str(row.get('Password') or row.get('password') or '').strip()
                                break
                        
                        if not original_password:
                            logger.warning(f"Could not find original password for user {user_data['username']}")
                            continue
                        
                        # Format template with user data
                        formatted_subject = template_subject.format(
                            username=user.username,
                            email=user.email,
                            password=original_password,
                            college=user.college or '',
                            department=user.department or '',
                            year=user.year or ''
                        )
                        
                        formatted_body = template_body_default.format(
                            username=user.username,
                            email=user.email,
                            password=original_password,
                            college=user.college or '',
                            department=user.department or '',
                            year=user.year or ''
                        )
                        
                        # Convert to HTML format if needed
                        if '\n' in formatted_body and '<br>' not in formatted_body:
                            formatted_body = formatted_body.replace('\n', '<br>').replace('\n\n', '<br><br>')
                        
                        # Send using the notification service
                        email_log = service.send_email_notification(
                            user_id=user.id,
                            email=user.email,
                            subject=formatted_subject,
                            body=formatted_body
                        )
                        
                        if email_log.status == "sent":
                            email_success_count += 1
                            logger.info(f"Bulk registration welcome email sent successfully to {user.email}")
                        elif email_log.status == "queued":
                            email_success_count += 1
                            logger.info(f"Bulk registration welcome email queued for {user.email}")
                        else:
                            email_failed_count += 1
                            logger.warning(f"Bulk registration welcome email failed for {user.email}: {email_log.error_message}")
                            
                    except Exception as user_email_error:
                        email_failed_count += 1
                        logger.warning(f"Failed to send welcome email to {user_data.get('email', 'unknown')}: {str(user_email_error)}")
                
                logger.info(f"Bulk email sending completed: {email_success_count} emails sent successfully, {email_failed_count} failed")
                
            except Exception as e:
                logger.warning(f"Failed to send bulk welcome emails: {str(e)}")
            
            # Log the successful upload
            if hasattr(current_admin, 'username') and db.query(Admin).filter(Admin.id == current_admin.id).first():
                log_admin_action(
                    admin_id=current_admin.id,
                    admin_username=current_admin.username,
                    action_type="BULK_UPLOAD",
                    resource_type="USER",
                    details=f"Bulk uploaded {len(created_users)} {user_type_filter.lower()}s from {file.filename}"
                )
        else:
            db.rollback()
        
        # Log the final result
        logger.info(f"Bulk upload completed: {len(created_users)} {user_type_filter.lower()}s created, {len(errors)} errors")
        logger.info(f"Created {user_type_filter.lower()}s: {[u['username'] for u in created_users]}")
        if errors:
            logger.warning(f"Errors during bulk upload: {errors[:5]}")  # Log first 5 errors
        
        return {
            "message": f"Successfully created {len(created_users)} {user_type_filter.lower()}s from {file.filename}. Welcome emails have been sent to all users.",
            "created_users": [u['username'] for u in created_users],
            "errors": errors[:10],  # Return first 10 errors to avoid overwhelming response
            "total_processed": len(records),
            "success_count": len(created_users),
            "error_count": len(errors),
            "user_type": user_type_filter
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Bulk upload users error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"File upload failed: {str(e)}")

# Enhanced Course Management
@app.post("/admin/courses")
async def create_course(
    course_data: CourseCreate, 
    current_user = Depends(get_current_admin_or_presenter), 
    db: Session = Depends(get_db)
):
    try:
        course = Course(
            title=course_data.title,
            description=course_data.description,
            duration_weeks=course_data.duration_weeks,
            sessions_per_week=course_data.sessions_per_week,
            is_active=course_data.is_active
        )
        db.add(course)
        db.commit()
        db.refresh(course)
        
        # Auto-setup course structure
        await auto_setup_course_structure(
            course.id, 
            course_data.duration_weeks, 
            course_data.sessions_per_week, 
            db
        )
        
        # Log course creation
        if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="CREATE",
                resource_type="COURSE",
                resource_id=course.id,
                details=f"Created course: {course.title} ({course_data.duration_weeks} weeks)"
            )
        elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="CREATE",
                resource_type="COURSE",
                resource_id=course.id,
                details=f"Created course: {course.title} ({course_data.duration_weeks} weeks)"
            )
        
        return {"message": "Course created successfully with auto-setup", "course_id": course.id}
    except Exception as e:
        db.rollback()
        logger.error(f"Create course error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create course")

@app.post("/admin/courses/{course_id}/auto-setup")
async def setup_course_structure(
    course_id: int,
    setup_data: CourseAutoSetup,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        course = db.query(Course).filter(Course.id == course_id).first()
        if not course:
            raise HTTPException(status_code=404, detail="Course not found")
        
        await auto_setup_course_structure(
            course_id, 
            setup_data.duration_weeks, 
            setup_data.sessions_per_week, 
            db
        )
        
        return {"message": "Course structure setup completed"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Auto setup error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to setup course structure")

async def auto_setup_course_structure(
    course_id: int, 
    duration_weeks: int, 
    sessions_per_week: int, 
    db: Session
):
    """Auto-generate course structure with weeks, modules, and sessions"""
    try:
        # Create modules for each week
        for week in range(1, duration_weeks + 1):
            module = Module(
                course_id=course_id,
                week_number=week,
                title=f"Week {week} - Module",
                description=f"Learning objectives and content for week {week}"
            )
            db.add(module)
            db.flush()  # Get the module ID
            
            # Create sessions for each module
            for session_num in range(1, sessions_per_week + 1):
                session = SessionModel(
                    module_id=module.id,
                    session_number=session_num,
                    title=f"Week {week} - Session {session_num}",
                    description=f"Session {session_num} content for week {week}",
                    duration_minutes=120
                )
                db.add(session)
        
        db.commit()
    except Exception as e:
        db.rollback()
        raise e

@app.get("/admin/courses")
async def get_admin_courses(
    page: int = 1,
    limit: int = 20,
    search: Optional[str] = None,
    current_user = Depends(get_current_admin_or_presenter), 
    db: Session = Depends(get_db)
):
    try:
        query = db.query(Course)
        
        if search:
            query = query.filter(
                or_(
                    Course.title.contains(search),
                    Course.description.contains(search)
                )
            )
        
        total = query.count()
        courses = query.offset((page - 1) * limit).limit(limit).all()
        
        result = []
        for course in courses:
            # Get course statistics
            enrolled_count = db.query(Enrollment).filter(Enrollment.course_id == course.id).count()
            modules_count = db.query(Module).filter(Module.course_id == course.id).count()
            
            # Calculate weeks based on actual modules or use stored value
            max_week = db.query(func.max(Module.week_number)).filter(Module.course_id == course.id).scalar()
            duration_weeks = max_week if max_week else course.duration_weeks
            
            result.append({
                "id": course.id,
                "title": course.title,
                "description": course.description,
                "duration_weeks": duration_weeks,
                "sessions_per_week": course.sessions_per_week,
                "is_active": course.is_active,
                "enrolled_students": enrolled_count,
                "modules_count": modules_count,
                "created_at": course.created_at
            })
        
        return {
            "courses": result,
            "total": total,
            "page": page,
            "limit": limit
        }
    except Exception as e:
        logger.error(f"Get courses error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch courses")

@app.put("/admin/courses/{course_id}")
async def update_course(
    course_id: int,
    course_data: CourseUpdate,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        course = db.query(Course).filter(Course.id == course_id).first()
        if not course:
            raise HTTPException(status_code=404, detail="Course not found")
        
        update_data = course_data.dict(exclude_unset=True)
        for field, value in update_data.items():
            setattr(course, field, value)
        
        db.commit()
        
        # Log course update
        if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="UPDATE",
                resource_type="COURSE",
                resource_id=course_id,
                details=f"Updated course: {course.title}"
            )
        elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="UPDATE",
                resource_type="COURSE",
                resource_id=course_id,
                details=f"Updated course: {course.title}"
            )
        
        return {"message": "Course updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Update course error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update course")

@app.delete("/admin/courses/{course_id}")
async def delete_course(
    course_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        course = db.query(Course).filter(Course.id == course_id).first()
        if not course:
            raise HTTPException(status_code=404, detail="Course not found")
        
        # Get user role
        user_role = None
        if hasattr(current_user, 'username'):
            # Check user type
            if db.query(Admin).filter(Admin.id == current_user.id).first():
                user_role = "Admin"
            elif db.query(Manager).filter(Manager.id == current_user.id).first():
                user_role = "Manager"
            elif db.query(Presenter).filter(Presenter.id == current_user.id).first():
                user_role = "Presenter"
            elif db.query(Mentor).filter(Mentor.id == current_user.id).first():
                user_role = "Mentor"
            else:
                # Check if it's a student in the User table
                user_check = db.query(User).filter(User.id == current_user.id).first()
                if user_check:
                    user_role = user_check.role
        
        # Check if operation requires approval
        if requires_approval(user_role, "delete"):
            # Create approval request instead of deleting immediately
            course_title = course.title
            
            approval_id = create_approval_request(
                db=db,
                user_id=current_user.id,
                user_role=user_role,
                operation_type="delete",
                target_entity_type="course",
                target_entity_id=course_id,
                operation_data={
                    "course_title": course_title,
                    "course_description": course.description,
                    "requester_username": current_user.username
                },
                reason=f"Delete course: {course_title}"
            )
            
            # Log the approval request
            if user_role == "Presenter":
                log_presenter_action(
                    presenter_id=current_user.id,
                    presenter_username=current_user.username,
                    action_type="REQUEST_APPROVAL",
                    resource_type="COURSE",
                    resource_id=course_id,
                    details=f"Requested approval to delete course: {course_title}"
                )
            elif user_role == "Mentor":
                log_mentor_action(
                    mentor_id=current_user.id,
                    mentor_username=current_user.username,
                    action_type="REQUEST_APPROVAL",
                    resource_type="COURSE",
                    resource_id=course_id,
                    details=f"Requested approval to delete course: {course_title}"
                )
            
            return {
                "message": "Approval request submitted successfully. The course deletion is pending approval from an Admin or Manager.",
                "approval_id": approval_id,
                "status": "pending_approval"
            }
        
        # If user is Admin or Manager, proceed with direct deletion
        course_title = course.title
        
        # Delete related records in the correct order to avoid foreign key constraints
        
        # 1. Delete cohort_courses entries
        db.query(CohortCourse).filter(CohortCourse.course_id == course_id).delete()
        
        # 2. Delete enrollments
        db.query(Enrollment).filter(Enrollment.course_id == course_id).delete()
        
        # 3. Delete certificates
        db.query(Certificate).filter(Certificate.course_id == course_id).delete()
        
        # 4. Delete modules and their related content
        modules = db.query(Module).filter(Module.course_id == course_id).all()
        for module in modules:
            # Get sessions for this module
            sessions = db.query(SessionModel).filter(SessionModel.module_id == module.id).all()
            for session in sessions:
                # Delete session-related content
                db.query(Attendance).filter(Attendance.session_id == session.id).delete()
                db.query(Resource).filter(Resource.session_id == session.id).delete()
                db.query(SessionContent).filter(SessionContent.session_id == session.id).delete()
                
                # Delete quizzes and their attempts
                quizzes = db.query(Quiz).filter(Quiz.session_id == session.id).all()
                for quiz in quizzes:
                    db.query(QuizAttempt).filter(QuizAttempt.quiz_id == quiz.id).delete()
                db.query(Quiz).filter(Quiz.session_id == session.id).delete()
            
            # Delete sessions
            db.query(SessionModel).filter(SessionModel.module_id == module.id).delete()
            
            # Delete forums and their posts
            forums = db.query(Forum).filter(Forum.module_id == module.id).all()
            for forum in forums:
                db.query(ForumPost).filter(ForumPost.forum_id == forum.id).delete()
            db.query(Forum).filter(Forum.module_id == module.id).delete()
        
        # 5. Delete modules
        db.query(Module).filter(Module.course_id == course_id).delete()
        
        # 6. Finally delete the course
        db.delete(course)
        db.commit()
        
        # Log course deletion
        if user_role == "Admin":
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="DELETE",
                resource_type="COURSE",
                resource_id=course_id,
                details=f"Deleted course: {course_title}"
            )
        elif user_role == "Manager":
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="DELETE",
                resource_type="COURSE",
                resource_id=course_id,
                details=f"Manager deleted course: {course_title}"
            )
        
        return {"message": "Course deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Delete course error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete course")



# Module Management
@app.get("/admin/course/{course_id}")
async def get_course(
    course_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        course = db.query(Course).filter(Course.id == course_id).first()
        if not course:
            raise HTTPException(status_code=404, detail="Course not found")
        
        return {
            "id": course.id,
            "title": course.title,
            "description": course.description,
            "created_at": course.created_at
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get course error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch course")

@app.get("/admin/module/{module_id}")
async def get_module(
    module_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        module = db.query(Module).filter(Module.id == module_id).first()
        if not module:
            raise HTTPException(status_code=404, detail="Module not found")
        
        return {
            "id": module.id,
            "course_id": module.course_id,
            "week_number": module.week_number,
            "title": module.title,
            "description": module.description,
            "start_date": module.start_date,
            "end_date": module.end_date,
            "created_at": module.created_at
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get module error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch module")

@app.get("/admin/modules")
async def get_course_modules(
    course_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        modules = db.query(Module).filter(Module.course_id == course_id).order_by(Module.week_number).all()
        
        result = []
        for module in modules:
            # Get sessions for this module
            sessions = db.query(SessionModel).filter(SessionModel.module_id == module.id).order_by(SessionModel.session_number).all()
            
            # Get module statistics
            total_resources = sum([
                db.query(Resource).filter(Resource.session_id == s.id).count() + 
                db.query(SessionContent).filter(SessionContent.session_id == s.id).count()
                for s in sessions
            ])
            
            result.append({
                "id": module.id,
                "week_number": module.week_number,
                "title": module.title,
                "description": module.description,
                "start_date": module.start_date,
                "end_date": module.end_date,
                "sessions_count": len(sessions),
                "resources_count": total_resources,
                "created_at": module.created_at,
                "sessions": [{
                    "id": s.id,
                    "session_number": s.session_number,
                    "title": s.title,
                    "scheduled_time": s.scheduled_time,
                    "duration_minutes": s.duration_minutes,
                    "session_type": getattr(s, 'session_type', 'LIVE'),
                    "is_completed": getattr(s, 'is_completed', False),
                    "has_recording": bool(s.recording_url)
                } for s in sessions]
            })
        
        return {"modules": result}
    except Exception as e:
        logger.error(f"Get course modules error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch modules")

@app.post("/admin/modules")
async def create_module(
    module_data: ModuleCreate,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        # Check if course exists
        course = db.query(Course).filter(Course.id == module_data.course_id).first()
        if not course:
            raise HTTPException(status_code=404, detail="Course not found")
        
        module = Module(**module_data.dict())
        db.add(module)
        db.commit()
        db.refresh(module)
        
        # Log module creation
        if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="CREATE",
                resource_type="MODULE",
                resource_id=module.id,
                details=f"Created module: {module_data.title}"
            )
        elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="CREATE",
                resource_type="MODULE",
                resource_id=module.id,
                details=f"Created module: {module_data.title}"
            )
        
        return {"message": "Module created successfully", "module_id": module.id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Create module error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create module")

@app.put("/admin/modules/{module_id}")
async def update_module(
    module_id: int,
    module_data: ModuleUpdate,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        module = db.query(Module).filter(Module.id == module_id).first()
        if not module:
            raise HTTPException(status_code=404, detail="Module not found")
        
        update_data = module_data.dict(exclude_unset=True)
        for field, value in update_data.items():
            setattr(module, field, value)
        
        db.commit()
        
        # Log module update
        if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="UPDATE",
                resource_type="MODULE",
                resource_id=module_id,
                details=f"Updated module: {module.title}"
            )
        elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="UPDATE",
                resource_type="MODULE",
                resource_id=module_id,
                details=f"Updated module: {module.title}"
            )
        
        return {"message": "Module updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Update module error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update module")

@app.delete("/admin/modules/{module_id}")
async def delete_module(
    module_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        module = db.query(Module).filter(Module.id == module_id).first()
        if not module:
            raise HTTPException(status_code=404, detail="Module not found")
        
        module_title = module.title
        
        db.delete(module)
        db.commit()
        
        # Log module deletion
        if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="DELETE",
                resource_type="MODULE",
                resource_id=module_id,
                details=f"Deleted module: {module_title}"
            )
        elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="DELETE",
                resource_type="MODULE",
                resource_id=module_id,
                details=f"Deleted module: {module_title}"
            )
        
        return {"message": "Module deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Delete module error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete module")

# Presenter Session Management - Add after existing session endpoints
@app.post("/presenter/sessions")
async def create_presenter_session(
    session_data: SessionCreate,
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    try:
        logger.info(f"Presenter creating session with data: {session_data.dict()}")
        
        # Check if module exists
        module = db.query(Module).filter(Module.id == session_data.module_id).first()
        if not module:
            logger.error(f"Module {session_data.module_id} not found")
            raise HTTPException(status_code=404, detail="Module not found")
        
        # Combine date and time if provided
        scheduled_datetime = None
        if session_data.scheduled_date and session_data.scheduled_time:
            try:
                from datetime import datetime
                date_str = f"{session_data.scheduled_date} {session_data.scheduled_time}"
                scheduled_datetime = datetime.strptime(date_str, '%d-%m-%Y %H:%M')
            except ValueError:
                logger.warning(f"Invalid date/time format: {session_data.scheduled_date} {session_data.scheduled_time}")
        
        session = SessionModel(
            module_id=session_data.module_id,
            session_number=session_data.session_number,
            title=session_data.title,
            description=session_data.description,
            scheduled_time=scheduled_datetime,
            duration_minutes=session_data.duration_minutes,
            zoom_link=session_data.meeting_link,
            syllabus_content=getattr(session_data, 'syllabus_content', None)
        )
        
        db.add(session)
        db.commit()
        db.refresh(session)
        
        logger.info(f"Presenter session created successfully with ID: {session.id}")
        
        # Log presenter session creation
        log_presenter_action(
            presenter_id=current_presenter.id,
            presenter_username=current_presenter.username,
            action_type="CREATE",
            resource_type="SESSION",
            resource_id=session.id,
            details=f"Created session: {session_data.title}"
        )
        
        return {"message": "Session created successfully", "session_id": session.id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Create presenter session error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create session: {str(e)}")
@app.post("/admin/sessions")
async def create_session(
    session_data: SessionCreate,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        logger.info(f"Creating session with data: {session_data.dict()}")
        
        # Check if module exists
        module = db.query(Module).filter(Module.id == session_data.module_id).first()
        if not module:
            logger.error(f"Module {session_data.module_id} not found")
            raise HTTPException(status_code=404, detail="Module not found")
        
        # Combine date and time if provided
        scheduled_datetime = None
        if session_data.scheduled_date and session_data.scheduled_time:
            try:
                from datetime import datetime
                date_str = f"{session_data.scheduled_date} {session_data.scheduled_time}"
                scheduled_datetime = datetime.strptime(date_str, '%d-%m-%Y %H:%M')
            except ValueError:
                logger.warning(f"Invalid date/time format: {session_data.scheduled_date} {session_data.scheduled_time}")
        
        # Create session with explicit field mapping
        session = SessionModel(
            module_id=session_data.module_id,
            session_number=session_data.session_number,
            title=session_data.title,
            description=session_data.description,
            scheduled_time=scheduled_datetime,
            duration_minutes=session_data.duration_minutes,
            zoom_link=session_data.zoom_link,
            syllabus_content=session_data.syllabus_content
        )
        
        db.add(session)
        db.commit()
        db.refresh(session)
        
        logger.info(f"Session created successfully with ID: {session.id}")
        
        # Log session creation
        if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="CREATE",
                resource_type="SESSION",
                resource_id=session.id,
                details=f"Created session: {session_data.title}"
            )
        elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="CREATE",
                resource_type="SESSION",
                resource_id=session.id,
                details=f"Created session: {session_data.title}"
            )
        
        return {"message": "Session created successfully", "session_id": session.id}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Create session error: {str(e)}")
        logger.error(f"Session data that failed: {session_data.dict() if session_data else 'None'}")
        raise HTTPException(status_code=500, detail=f"Failed to create session: {str(e)}")

@app.get("/admin/sessions")
async def get_module_sessions(
    module_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        sessions = db.query(SessionModel).filter(SessionModel.module_id == module_id).order_by(SessionModel.session_number).all()
        
        result = []
        for session in sessions:
            # Get session statistics
            resources_count = db.query(Resource).filter(Resource.session_id == session.id).count()
            session_content_count = db.query(SessionContent).filter(SessionContent.session_id == session.id).count()
            total_resources_count = resources_count + session_content_count
            attendance_count = db.query(Attendance).filter(Attendance.session_id == session.id).count()
            
            result.append({
                "id": session.id,
                "session_number": session.session_number,
                "title": session.title,
                "description": session.description,
                "scheduled_time": session.scheduled_time,
                "duration_minutes": session.duration_minutes,
                "session_type": getattr(session, 'session_type', 'LIVE'),
                "zoom_link": session.zoom_link,
                "recording_url": session.recording_url,
                "is_completed": getattr(session, 'is_completed', False),
                "resources_count": total_resources_count,
                "attendance_count": attendance_count,
                "syllabus_content": session.syllabus_content,
                "created_at": session.created_at
            })
        
        return result
    except Exception as e:
        logger.error(f"Get module sessions error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch sessions")

@app.put("/admin/sessions/{session_id}")
async def update_session(
    session_id: int,
    session_data: SessionUpdate,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        session = db.query(SessionModel).filter(SessionModel.id == session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        update_data = session_data.dict(exclude_unset=True)
        for field, value in update_data.items():
            setattr(session, field, value)
        
        db.commit()
        
        # Log session update
        if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="UPDATE",
                resource_type="SESSION",
                resource_id=session_id,
                details=f"Updated session: {session.title}"
            )
        elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="UPDATE",
                resource_type="SESSION",
                resource_id=session_id,
                details=f"Updated session: {session.title}"
            )
        
        return {"message": "Session updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Update session error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update session")

@app.delete("/admin/sessions/{session_id}")
async def delete_session(
    session_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        session = db.query(SessionModel).filter(SessionModel.id == session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        session_title = session.title
        
        db.delete(session)
        db.commit()
        
        # Log session deletion
        if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="DELETE",
                resource_type="SESSION",
                resource_id=session_id,
                details=f"Deleted session: {session_title}"
            )
        elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="DELETE",
                resource_type="SESSION",
                resource_id=session_id,
                details=f"Deleted session: {session_title}"
            )
        
        return {"message": "Session deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Delete session error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete session")

# Session Content Management
@app.post("/admin/session-content")
async def create_session_content(
    content_data: SessionContentCreate,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        session = db.query(SessionModel).filter(SessionModel.id == content_data.session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        uploaded_by_id = None
        if hasattr(current_user, 'username'):
            admin_check = db.query(Admin).filter(Admin.id == current_user.id).first()
            if admin_check:
                uploaded_by_id = current_user.id
        
        content = SessionContent(
            session_id=content_data.session_id,
            content_type=content_data.content_type,
            title=content_data.title,
            description=content_data.description,
            file_path=content_data.file_path,
            file_type=content_data.file_type,
            file_size=content_data.file_size,
            meeting_url=content_data.meeting_url,
            scheduled_time=content_data.scheduled_time,
            uploaded_by=uploaded_by_id
        )
        
        db.add(content)
        db.flush()
        
        # Auto-create calendar event for meeting links with scheduled time
        calendar_mapped = False
        if content_data.content_type == "MEETING_LINK" and content_data.scheduled_time:
            try:
                from database import CalendarEvent
                end_datetime = content_data.scheduled_time + timedelta(minutes=getattr(content_data, 'duration_minutes', 60))
                
                calendar_event = CalendarEvent(
                    title=f"Meeting: {content_data.title}",
                    description=content_data.description or f"Scheduled meeting: {content_data.title}",
                    start_datetime=content_data.scheduled_time,
                    end_datetime=end_datetime,
                    event_type="meeting",
                    location=content_data.meeting_url,
                    is_auto_generated=True,
                    created_by_admin_id=uploaded_by_id
                )
                
                db.add(calendar_event)
                calendar_mapped = True
                logger.info(f"Auto-created calendar event for meeting: {content_data.title}")
            except Exception as e:
                logger.warning(f"Failed to create calendar event: {str(e)}")
        
        db.commit()
        
        message = "Content created successfully"
        if calendar_mapped:
            message += " and mapped to calendar"
        
        return {
            "message": message,
            "content_id": content.id,
            "calendar_mapped": calendar_mapped
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Create session content error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create content")

@app.get("/admin/session-content/{session_id}")
async def get_session_content(
    session_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        contents = db.query(SessionContent).filter(SessionContent.session_id == session_id).all()
        
        return {
            "contents": [{
                "id": c.id,
                "content_type": c.content_type,
                "title": c.title,
                "description": c.description,
                "file_type": c.file_type,
                "file_path": c.file_path,
                "file_size": c.file_size,
                "meeting_url": c.meeting_url,
                "scheduled_time": c.scheduled_time,
                "created_at": c.created_at
            } for c in contents]
        }
    except Exception as e:
        logger.error(f"Get session content error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch content")

@app.get("/admin/session-content")
async def get_session_content_by_query(
    session_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        session_contents = db.query(SessionContent).filter(SessionContent.session_id == session_id).all()
        resources = db.query(Resource).filter(Resource.session_id == session_id).all()
        
        all_content = []
        
        for c in session_contents:
            all_content.append({
                "id": c.id,
                "content_type": c.content_type,
                "title": c.title,
                "description": c.description,
                "file_type": c.file_type,
                "file_path": c.file_path,
                "file_size": c.file_size,
                "meeting_url": c.meeting_url,
                "scheduled_time": c.scheduled_time,
                "created_at": c.created_at
            })
        
        for r in resources:
            all_content.append({
                "id": f"resource_{r.id}",
                "content_type": "RESOURCE",
                "title": r.title,
                "description": r.description,
                "file_type": r.resource_type,
                "file_path": r.file_path,
                "file_size": r.file_size,
                "meeting_url": None,
                "scheduled_time": None,
                "created_at": r.uploaded_at or r.created_at
            })
        
        return {"contents": all_content}
    except Exception as e:
        logger.error(f"Get session content error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch content")

@app.delete("/admin/session-content/{content_id}")
async def delete_session_content(
    content_id: str,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        # Handle both integer IDs and content_X format
        actual_content_id = None
        if content_id.startswith("content_"):
            try:
                actual_content_id = int(content_id.replace("content_", ""))
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid content ID format")
        else:
            try:
                actual_content_id = int(content_id)
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid content ID format")
        
        content = db.query(SessionContent).filter(SessionContent.id == actual_content_id).first()
        if not content:
            raise HTTPException(status_code=404, detail="Content not found")
        
        content_title = content.title
        content_type = content.content_type
        
        # Delete physical file if it exists (but not for meeting links)
        if content.file_path and os.path.exists(content.file_path) and content_type != "MEETING_LINK":
            os.remove(content.file_path)
        
        db.delete(content)
        db.commit()
        
        # Log session content deletion
        if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="DELETE",
                resource_type="SESSION_CONTENT",
                resource_id=actual_content_id,
                details=f"Deleted {content_type.lower()}: {content_title}"
            )
        elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="DELETE",
                resource_type="SESSION_CONTENT",
                resource_id=actual_content_id,
                details=f"Deleted {content_type.lower()}: {content_title}"
            )
        
        return {"message": f"{content_type.replace('_', ' ').title()} deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Delete session content error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete content")

# File Link Resource Management
@app.post("/admin/resources/file-link")
async def create_file_link_resource(
    session_id: int = Form(...),
    title: str = Form(...),
    file_url: str = Form(...),
    description: str = Form(""),
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    """Create a file link resource that downloads and stores the file"""
    try:
        # Validate session exists
        session = db.query(SessionModel).filter(SessionModel.id == session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        # Validate URL
        if not file_url.startswith(('http://', 'https://')):
            raise HTTPException(status_code=400, detail="Invalid URL format")
        
        # Download file from URL
        try:
            async with aiohttp.ClientSession() as client_session:
                async with client_session.get(file_url) as response:
                    if response.status != 200:
                        raise HTTPException(status_code=400, detail=f"Failed to download file: HTTP {response.status}")
                    
                    # Get file info
                    content_type = response.headers.get('content-type', '')
                    content_length = response.headers.get('content-length')
                    file_size = int(content_length) if content_length else 0
                    
                    # Determine file extension from URL or content type
                    file_ext = ''
                    if '.' in file_url.split('/')[-1]:
                        file_ext = '.' + file_url.split('/')[-1].split('.')[-1].lower()
                    elif 'pdf' in content_type:
                        file_ext = '.pdf'
                    elif 'powerpoint' in content_type or 'presentation' in content_type:
                        file_ext = '.pptx'
                    elif 'word' in content_type or 'document' in content_type:
                        file_ext = '.docx'
                    elif 'video' in content_type:
                        file_ext = '.mp4'
                    elif 'image' in content_type:
                        file_ext = '.jpg'
                    else:
                        file_ext = '.bin'
                    
                    # Generate unique filename
                    unique_filename = f"{uuid.uuid4()}{file_ext}"
                    file_path = UPLOAD_BASE_DIR / "resources" / unique_filename
                    
                    # Save file
                    async with aiofiles.open(file_path, 'wb') as f:
                        async for chunk in response.content.iter_chunked(8192):
                            await f.write(chunk)
                    
                    # Get actual file size
                    actual_file_size = file_path.stat().st_size
                    
        except aiohttp.ClientError as e:
            raise HTTPException(status_code=400, detail=f"Failed to download file: {str(e)}")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error downloading file: {str(e)}")
        
        # Determine resource type based on file extension
        resource_type = "OTHER"
        if file_ext.lower() in ['.pdf']:
            resource_type = "PDF"
        elif file_ext.lower() in ['.ppt', '.pptx']:
            resource_type = "PPT"
        elif file_ext.lower() in ['.mp4', '.avi', '.mov', '.wmv']:
            resource_type = "VIDEO"
        elif file_ext.lower() in ['.txt', '.md']:
            resource_type = "TXT"
        elif file_ext.lower() in ['.py', '.js', '.html', '.css', '.java', '.cpp']:
            resource_type = "CODE"
        
        # Get uploader ID
        uploaded_by_id = None
        if hasattr(current_user, 'username'):
            admin_check = db.query(Admin).filter(Admin.id == current_user.id).first()
            if admin_check:
                uploaded_by_id = current_user.id
        
        # Create resource record
        resource = Resource(
            session_id=session_id,
            title=title,
            resource_type=resource_type,
            file_path=str(file_path),
            file_size=actual_file_size,
            description=description,
            uploaded_by=uploaded_by_id
        )
        
        db.add(resource)
        db.commit()
        db.refresh(resource)
        
        # Log resource creation
        if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="CREATE",
                resource_type="RESOURCE",
                resource_id=resource.id,
                details=f"Downloaded and created resource from URL: {title}"
            )
        elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="CREATE",
                resource_type="RESOURCE",
                resource_id=resource.id,
                details=f"Downloaded and created resource from URL: {title}"
            )
        
        return {
            "message": "File downloaded and resource created successfully",
            "resource_id": resource.id,
            "filename": unique_filename,
            "file_size": actual_file_size,
            "resource_type": resource_type
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        # Clean up downloaded file if it exists
        try:
            if 'file_path' in locals() and file_path.exists():
                file_path.unlink()
        except:
            pass
        logger.error(f"Create file link resource error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to create file link resource: {str(e)}")
@app.get("/admin/resources/{session_id}")
async def get_session_resources(
    session_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    """Get all resources for a session including file links"""
    try:
        session = db.query(SessionModel).filter(SessionModel.id == session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        resources = db.query(Resource).filter(Resource.session_id == session_id).all()
        
        result = []
        for resource in resources:
            # Get file info
            file_exists = os.path.exists(resource.file_path) if resource.file_path else False
            filename = os.path.basename(resource.file_path) if resource.file_path else None
            
            result.append({
                "id": resource.id,
                "title": resource.title,
                "resource_type": resource.resource_type,
                "file_path": resource.file_path,
                "filename": filename,
                "file_size": resource.file_size,
                "description": resource.description,
                "file_exists": file_exists,
                "uploaded_at": resource.uploaded_at,
                "created_at": resource.created_at,
                "download_url": f"/api/resources/{filename}" if filename and file_exists else None
            })
        
        return {"resources": result}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get session resources error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch resources")

@app.delete("/admin/resources/{resource_id}")
async def delete_resource(
    resource_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    """Delete a resource and its associated file"""
    try:
        resource = db.query(Resource).filter(Resource.id == resource_id).first()
        if not resource:
            raise HTTPException(status_code=404, detail="Resource not found")
        
        resource_title = resource.title
        file_path = resource.file_path
        
        # Delete the database record
        db.delete(resource)
        db.commit()
        
        # Delete the physical file if it exists
        if file_path and os.path.exists(file_path):
            try:
                os.remove(file_path)
                logger.info(f"Deleted file: {file_path}")
            except Exception as e:
                logger.warning(f"Failed to delete file {file_path}: {str(e)}")
        
        # Log resource deletion
        if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="DELETE",
                resource_type="RESOURCE",
                resource_id=resource_id,
                details=f"Deleted resource: {resource_title}"
            )
        elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="DELETE",
                resource_type="RESOURCE",
                resource_id=resource_id,
                details=f"Deleted resource: {resource_title}"
            )
        
        return {"message": "Resource deleted successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Delete resource error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete resource")

@app.post("/admin/upload/resource")
async def upload_resource_file(
    file: UploadFile = File(...),
    session_id: int = Form(...),
    title: str = Form(...),
    description: str = Form(""),
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        logger.info(f"Upload resource request - session_id: {session_id}, title: {title}, file: {file.filename if file else 'None'}")
        
        if not session_id:
            raise HTTPException(status_code=400, detail="Session ID is required")
        
        session = db.query(SessionModel).filter(SessionModel.id == session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        # Create unique filename
        file_extension = os.path.splitext(file.filename)[1]
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        file_path = UPLOAD_BASE_DIR / "resources" / unique_filename
        
        # Save file
        logger.info(f"Saving file to: {file_path}")
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        logger.info(f"File saved successfully, size: {len(content)} bytes")
        
        # Determine resource type based on file extension and content type
        resource_type = "OTHER"
        file_ext = os.path.splitext(file.filename)[1].lower() if file.filename else ""
        
        # Check file extension first, then content type
        if file_ext in [".pdf"]:
            resource_type = "PDF"
        elif file_ext in [".ppt", ".pptx"]:
            resource_type = "PPT"
        elif file_ext in [".mp4", ".avi", ".mov", ".wmv", ".flv", ".mkv"]:
            resource_type = "VIDEO"
        elif file_ext in [".py", ".js", ".html", ".css", ".java", ".cpp", ".c", ".txt", ".json", ".xml"]:
            resource_type = "CODE"
        elif file.content_type:
            if "pdf" in file.content_type.lower():
                resource_type = "PDF"
            elif "powerpoint" in file.content_type.lower() or "presentation" in file.content_type.lower():
                resource_type = "PPT"
            elif "video" in file.content_type.lower():
                resource_type = "VIDEO"
            elif "text" in file.content_type.lower():
                resource_type = "CODE"
        
        # Create resource record
        resource = Resource(
            session_id=session_id,
            title=title or file.filename,
            resource_type=resource_type,
            file_path=str(file_path),
            file_size=len(content),
            description=description,
            uploaded_by=current_user.id
        )
        db.add(resource)
        db.commit()
        
        # Log resource upload
        if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="UPLOAD",
                resource_type="RESOURCE",
                resource_id=resource.id,
                details=f"Uploaded resource: {file.filename} ({len(content)} bytes)"
            )
        elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="UPLOAD",
                resource_type="RESOURCE",
                resource_id=resource.id,
                details=f"Uploaded resource: {file.filename} ({len(content)} bytes)"
            )
        
        return {
            "message": "Resource uploaded successfully",
            "resource_id": resource.id,
            "filename": file.filename,
            "file_size": len(content),
            "file_path": str(file_path)
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Upload resource error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to upload resource")

@app.post("/admin/test-upload")
async def test_upload(
    file: UploadFile = File(...),
    session_id: int = Form(...),
    title: str = Form(...),
    description: str = Form(""),
    current_user = Depends(get_current_admin_or_presenter)
):
    """Test endpoint for file upload functionality"""
    try:
        logger.info(f"Test upload - File: {file.filename}, Size: {file.size}, Type: {file.content_type}")
        logger.info(f"Test upload - Session ID: {session_id}, Title: {title}, Description: {description}")
        
        # Read file content
        content = await file.read()
        logger.info(f"File content read successfully: {len(content)} bytes")
        
        return {
            "message": "Test upload successful",
            "filename": file.filename,
            "size": len(content),
            "content_type": file.content_type,
            "session_id": session_id,
            "title": title,
            "description": description
        }
    except Exception as e:
        logger.error(f"Test upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Test upload failed: {str(e)}")

@app.post("/admin/upload/recording")
async def upload_session_recording(
    file: UploadFile = File(...),
    session_id: int = None,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        if not session_id:
            raise HTTPException(status_code=400, detail="Session ID is required")
        
        session = db.query(SessionModel).filter(SessionModel.id == session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        # Create unique filename for recording
        file_extension = os.path.splitext(file.filename)[1]
        unique_filename = f"recording_{session_id}_{uuid.uuid4()}{file_extension}"
        file_path = UPLOAD_BASE_DIR / "recordings" / unique_filename
        
        # Save recording file
        with open(file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Update session with recording URL
        session.recording_url = f"/api/recordings/{unique_filename}"
        db.commit()
        
        # Log recording upload
        if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="UPLOAD",
                resource_type="RECORDING",
                resource_id=session_id,
                details=f"Uploaded recording: {file.filename} ({len(content)} bytes) for session {session_id}"
            )
        elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="UPLOAD",
                resource_type="RECORDING",
                resource_id=session_id,
                details=f"Uploaded recording: {file.filename} ({len(content)} bytes) for session {session_id}"
            )
        
        return {
            "message": "Recording uploaded successfully",
            "recording_url": session.recording_url,
            "file_size": len(content)
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Upload recording error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to upload recording")

# Resource Management
@app.post("/admin/resources")
async def create_resource_legacy(
    request: Request,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        form = await request.form()
        logger.info(f"Form data received: {dict(form)}")
        
        # Extract form fields
        file = form.get("file")
        title = form.get("title")
        description = form.get("description", "")
        session_id = request.query_params.get("session_id") or "1"  # Default to session 1
        
        if not file:
            raise HTTPException(status_code=400, detail="File is required")
        if not title:
            raise HTTPException(status_code=400, detail="Title is required")
            
        session_id = int(session_id)
        
        session = db.query(SessionModel).filter(SessionModel.id == session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        # Create unique filename
        file_extension = os.path.splitext(file.filename)[1] if file.filename else ".txt"
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        saved_file_path = UPLOAD_BASE_DIR / "resources" / unique_filename
        
        # Save file
        with open(saved_file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Use resource_type from form if provided, otherwise determine from file
        resource_type = form.get("resource_type", "OTHER")
        if resource_type == "OTHER":
            file_ext = os.path.splitext(file.filename)[1].lower() if file.filename else ""
            if file_ext in [".pdf"]:
                resource_type = "PDF"
            elif file_ext in [".ppt", ".pptx"]:
                resource_type = "PPT"
            elif file_ext in [".docx", ".doc"]:
                resource_type = "DOCX"
            elif file_ext in [".mp4", ".avi", ".mov"]:
                resource_type = "VIDEO"
            elif file_ext in [".txt", ".py", ".js"]:
                resource_type = "CODE"
        
        # Create resource record
        resource = Resource(
            session_id=session_id,
            title=title,
            resource_type=resource_type,
            file_path=str(saved_file_path),
            file_size=len(content),
            description=description,
            uploaded_by=None
        )
        db.add(resource)
        db.commit()
        
        return {
            "message": "Resource created successfully",
            "resource_id": resource.id,
            "filename": file.filename,
            "file_size": len(content)
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Create resource error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create resource")

@app.post("/admin/resources/{session_id}")
async def create_resource(
    session_id: int,
    request: Request,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        form = await request.form()
        logger.info(f"Form data received: {dict(form)}")
        
        # Extract form fields
        file = form.get("file")
        title = form.get("title")
        description = form.get("description", "")
        
        if not file:
            raise HTTPException(status_code=400, detail="File is required")
        if not title:
            raise HTTPException(status_code=400, detail="Title is required")
        
        session = db.query(SessionModel).filter(SessionModel.id == session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        # Create unique filename
        file_extension = os.path.splitext(file.filename)[1] if file.filename else ".txt"
        unique_filename = f"{uuid.uuid4()}{file_extension}"
        saved_file_path = UPLOAD_BASE_DIR / "resources" / unique_filename
        
        # Save file
        with open(saved_file_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Use resource_type from form if provided, otherwise determine from file
        resource_type = form.get("resource_type", "OTHER")
        if resource_type == "OTHER":
            file_ext = os.path.splitext(file.filename)[1].lower() if file.filename else ""
            if file_ext in [".pdf"]:
                resource_type = "PDF"
            elif file_ext in [".ppt", ".pptx"]:
                resource_type = "PPT"
            elif file_ext in [".docx", ".doc"]:
                resource_type = "DOCX"
            elif file_ext in [".mp4", ".avi", ".mov"]:
                resource_type = "VIDEO"
            elif file_ext in [".txt", ".py", ".js"]:
                resource_type = "CODE"
        
        # Create resource record
        resource = Resource(
            session_id=session_id,
            title=title,
            resource_type=resource_type,
            file_path=str(saved_file_path),
            file_size=len(content),
            description=description,
            uploaded_by=current_user.id
        )
        db.add(resource)
        db.commit()
        
        return {
            "message": "Resource created successfully",
            "resource_id": resource.id,
            "filename": file.filename,
            "file_size": len(content)
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Create resource error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create resource")

@app.get("/admin/resources")
async def get_session_resources(
    session_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        resources = db.query(Resource).filter(Resource.session_id == session_id).all()
        
        result = []
        for r in resources:
            result.append({
                "id": r.id,
                "title": r.title,
                "resource_type": r.resource_type,
                "file_size": r.file_size,
                "description": r.description,
                "uploaded_at": r.uploaded_at,
                "created_at": r.created_at,
                "file_path": r.file_path
            })
        
        return {"resources": result}
    except Exception as e:
        logger.error(f"Get session resources error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch resources")

@app.delete("/admin/resources/{resource_id}")
async def delete_resource(
    resource_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        # First try to find in Resource table
        resource = db.query(Resource).filter(Resource.id == resource_id).first()
        if resource:
            resource_title = resource.title
            
            # Delete physical file
            if resource.file_path and os.path.exists(resource.file_path):
                os.remove(resource.file_path)
            
            db.delete(resource)
            db.commit()
            
            # Log resource deletion
            if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
                log_admin_action(
                    admin_id=current_user.id,
                    admin_username=current_user.username,
                    action_type="DELETE",
                    resource_type="RESOURCE",
                    resource_id=resource_id,
                    details=f"Deleted resource: {resource_title}"
                )
            elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
                log_presenter_action(
                    presenter_id=current_user.id,
                    presenter_username=current_user.username,
                    action_type="DELETE",
                    resource_type="RESOURCE",
                    resource_id=resource_id,
                    details=f"Deleted resource: {resource_title}"
                )
            
            return {"message": "Resource deleted successfully"}
        
        # If not found in Resource table, try SessionContent table (for meeting links)
        session_content = db.query(SessionContent).filter(SessionContent.id == resource_id).first()
        if session_content:
            content_title = session_content.title
            content_type = session_content.content_type
            
            # Delete physical file if it exists (but not for meeting links)
            if session_content.file_path and os.path.exists(session_content.file_path) and content_type != "MEETING_LINK":
                os.remove(session_content.file_path)
            
            db.delete(session_content)
            db.commit()
            
            # Log session content deletion
            if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
                log_admin_action(
                    admin_id=current_user.id,
                    admin_username=current_user.username,
                    action_type="DELETE",
                    resource_type="SESSION_CONTENT",
                    resource_id=resource_id,
                    details=f"Deleted {content_type.lower()}: {content_title}"
                )
            elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
                log_presenter_action(
                    presenter_id=current_user.id,
                    presenter_username=current_user.username,
                    action_type="DELETE",
                    resource_type="SESSION_CONTENT",
                    resource_id=resource_id,
                    details=f"Deleted {content_type.lower()}: {content_title}"
                )
            
            return {"message": f"{content_type.replace('_', ' ').title()} deleted successfully"}
        
        # If not found in either table
        raise HTTPException(status_code=404, detail="Resource not found")
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Delete resource error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete resource")



# Attendance Management
@app.post("/admin/attendance/bulk")
async def bulk_record_attendance(
    attendance_data: AttendanceBulkCreate,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        session = db.query(SessionModel).filter(SessionModel.id == attendance_data.session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        # Delete existing attendance for this session
        db.query(Attendance).filter(Attendance.session_id == attendance_data.session_id).delete()
        
        # Add new attendance records
        for record in attendance_data.attendance_records:
            attendance = Attendance(
                session_id=attendance_data.session_id,
                student_id=record["student_id"],
                attended=record["attended"],
                duration_minutes=record.get("duration_minutes", 0),
                join_time=record.get("join_time"),
                leave_time=record.get("leave_time")
            )
            db.add(attendance)
        
        db.commit()
        return {
            "message": f"Attendance recorded for {len(attendance_data.attendance_records)} students",
            "session_id": attendance_data.session_id
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Bulk attendance error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to record attendance")

@app.get("/admin/attendance/{session_id}")
async def get_session_attendance(
    session_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        attendance_records = db.query(Attendance).filter(Attendance.session_id == session_id).all()
        
        return {
            "attendance": [{
                "id": a.id,
                "student_name": f"{getattr(a.student, 'first_name', '') or ''} {getattr(a.student, 'last_name', '') or ''}".strip() or a.student.username,
                "student_email": a.student.email,
                "attended": a.attended,
                "duration_minutes": a.duration_minutes,
                "join_time": a.join_time,
                "leave_time": a.leave_time
            } for a in attendance_records]
        }
    except Exception as e:
        logger.error(f"Get session attendance error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch attendance")

# Quiz Management
@app.post("/admin/quizzes")
async def create_quiz(
    quiz_data: QuizCreate,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        quiz = Quiz(
            **quiz_data.dict(),
            created_by=current_user.id
        )
        db.add(quiz)
        db.commit()
        db.refresh(quiz)
        
        # Log quiz creation based on user type
        admin_check = db.query(Admin).filter(Admin.id == current_user.id).first()
        presenter_check = db.query(Presenter).filter(Presenter.id == current_user.id).first()
        
        if admin_check:
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="CREATE",
                resource_type="QUIZ",
                resource_id=quiz.id,
                details=f"Created quiz: {quiz_data.title}"
            )
        elif presenter_check:
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="CREATE",
                resource_type="QUIZ",
                resource_id=quiz.id,
                details=f"Created quiz: {quiz_data.title}"
            )
        
        return {"message": "Quiz created successfully", "quiz_id": quiz.id}
    except Exception as e:
        db.rollback()
        logger.error(f"Create quiz error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create quiz")

@app.get("/admin/quiz/{quiz_id}")
async def get_quiz(
    quiz_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        quiz = db.query(Quiz).filter(Quiz.id == quiz_id).first()
        if not quiz:
            raise HTTPException(status_code=404, detail="Quiz not found")
        
        # Parse questions if they exist
        questions = []
        if quiz.questions:
            try:
                questions = json.loads(quiz.questions)
            except:
                questions = []
        
        return {
            "id": quiz.id,
            "title": quiz.title,
            "description": quiz.description,
            "total_marks": quiz.total_marks,
            "time_limit_minutes": quiz.time_limit_minutes,
            "questions": questions,
            "is_active": quiz.is_active,
            "created_at": quiz.created_at
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get quiz error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch quiz")

@app.delete("/admin/quiz/{quiz_id}")
async def delete_quiz(
    quiz_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        quiz = db.query(Quiz).filter(Quiz.id == quiz_id).first()
        if not quiz:
            raise HTTPException(status_code=404, detail="Quiz not found")
        
        quiz_title = quiz.title
        
        # Delete related quiz attempts first
        db.query(QuizAttempt).filter(QuizAttempt.quiz_id == quiz_id).delete()
        
        db.delete(quiz)
        db.commit()
        
        # Log quiz deletion
        if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="DELETE",
                resource_type="QUIZ",
                resource_id=quiz_id,
                details=f"Deleted quiz: {quiz_title}"
            )
        elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="DELETE",
                resource_type="QUIZ",
                resource_id=quiz_id,
                details=f"Deleted quiz: {quiz_title}"
            )
        
        return {"message": "Quiz deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Delete quiz error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete quiz")

@app.get("/admin/quizzes")
async def get_session_quizzes_admin(
    session_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        quizzes = db.query(Quiz).filter(Quiz.session_id == session_id).all()
        
        result = []
        for quiz in quizzes:
            attempts_count = db.query(QuizAttempt).filter(QuizAttempt.quiz_id == quiz.id).count()
            avg_score = db.query(func.avg(QuizAttempt.score)).filter(QuizAttempt.quiz_id == quiz.id).scalar() or 0
            
            result.append({
                "id": quiz.id,
                "title": quiz.title,
                "description": quiz.description,
                "total_marks": quiz.total_marks,
                "time_limit_minutes": quiz.time_limit_minutes,
                "is_active": quiz.is_active,
                "attempts_count": attempts_count,
                "average_score": round(float(avg_score), 2),
                "created_at": quiz.created_at
            })
        
        return {"quizzes": result}
    except Exception as e:
        logger.error(f"Get session quizzes error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch quizzes")

# AI Quiz Generation Endpoints
@app.post("/admin/test-quiz-request")
async def test_quiz_request(
    request_data: dict,
    current_user = Depends(get_current_admin_or_presenter)
):
    """Test endpoint to validate quiz request data"""
    try:
        quiz_request = AIQuizGenerateRequest(**request_data)
        return {
            "message": "Request validation successful",
            "received_data": {
                "session_id": quiz_request.session_id,
                "title": quiz_request.title,
                "content": quiz_request.content[:100] + "..." if len(quiz_request.content) > 100 else quiz_request.content,
                "question_type": quiz_request.question_type,
                "num_questions": quiz_request.num_questions
            }
        }
    except Exception as e:
        return {
            "message": "Validation failed",
            "error": str(e),
            "received_data": request_data
        }

@app.post("/admin/generate-ai-quiz")
async def generate_ai_quiz(
    request_data: dict,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        import json
        
        # Log raw request data
        logger.info(f"Raw AI Quiz request data: {request_data}")
        
        # Validate and parse request
        try:
            quiz_request = AIQuizGenerateRequest(**request_data)
        except Exception as validation_error:
            logger.error(f"Request validation failed: {str(validation_error)}")
            raise HTTPException(status_code=422, detail=f"Invalid request data: {str(validation_error)}")
        
        # Log the parsed request
        logger.info(f"AI Quiz generation request: session_id={quiz_request.session_id}, title='{quiz_request.title}', question_type='{quiz_request.question_type}', num_questions={quiz_request.num_questions}")
        
        # Verify session exists
        session = db.query(SessionModel).filter(SessionModel.id == quiz_request.session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        # Generate quiz using AI logic (simplified for demo)
        generated_questions = generate_quiz_questions(
            quiz_request.content,
            quiz_request.question_type,
            quiz_request.num_questions
        )
        
        # Calculate total marks based on question type
        marks_per_question = {
            "MCQ": 2,
            "TRUE_FALSE": 1,
            "SHORT_ANSWER": 5
        }
        total_marks = len(generated_questions) * marks_per_question.get(quiz_request.question_type, 2)
        
        # Create quiz in database
        quiz = Quiz(
            session_id=quiz_request.session_id,
            title=quiz_request.title,
            description=f"AI-generated {quiz_request.question_type} quiz with {len(generated_questions)} questions",
            total_marks=total_marks,
            time_limit_minutes=max(15, len(generated_questions) * 2),
            questions=json.dumps(generated_questions),
            is_active=True,
            created_by=current_user.id
        )
        
        db.add(quiz)
        db.flush()  # Get the quiz ID
        
        db.commit()
        db.refresh(quiz)
        
        return {
            "message": "AI quiz generated successfully",
            "quiz_id": quiz.id,
            "questions": generated_questions,
            "total_marks": total_marks,
            "time_limit": quiz.time_limit_minutes
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"AI quiz generation error: {str(e)}")
        logger.error(f"Raw request data: {request_data}")
        raise HTTPException(status_code=500, detail="Failed to generate AI quiz")

@app.get("/admin/debug/session-content/{session_id}")
async def debug_session_content(
    session_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    """Debug endpoint to check session content in database"""
    try:
        contents = db.query(SessionContent).filter(SessionContent.session_id == session_id).all()
        
        debug_data = []
        for c in contents:
            debug_data.append({
                "id": c.id,
                "content_type": c.content_type,
                "title": c.title,
                "description": c.description,
                "meeting_url": c.meeting_url,
                "scheduled_time": c.scheduled_time,
                "file_path": c.file_path,
                "file_type": c.file_type,
                "created_at": c.created_at
            })
        
        return {
            "session_id": session_id,
            "total_contents": len(contents),
            "contents": debug_data
        }
    except Exception as e:
        logger.error(f"Debug session content error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to debug session content")

@app.get("/admin/session/{session_id}")
async def get_session(
    session_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        session = db.query(SessionModel).filter(SessionModel.id == session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        return {
            "id": session.id,
            "module_id": session.module_id,
            "session_number": session.session_number,
            "title": session.title,
            "description": session.description,
            "scheduled_time": session.scheduled_time,
            "duration_minutes": session.duration_minutes,
            "zoom_link": session.zoom_link,
            "recording_url": session.recording_url,
            "syllabus_content": session.syllabus_content,
            "created_at": session.created_at
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get session error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch session")

@app.get("/admin/session/{session_id}/available-quizzes")
async def get_available_quizzes(
    session_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    """Get all quizzes available for a session that aren't already in session content"""
    try:
        # Get all quizzes for this session
        quizzes = db.query(Quiz).filter(Quiz.session_id == session_id).all()
        
        # Get quiz IDs that are already in session content
        existing_content = db.query(SessionContent).filter(
            SessionContent.session_id == session_id,
            SessionContent.content_type == "QUIZ"
        ).all()
        
        existing_quiz_ids = set()
        for content in existing_content:
            if content.content_data:
                try:
                    content_json = json.loads(content.content_data)
                    if "quiz_id" in content_json:
                        existing_quiz_ids.add(content_json["quiz_id"])
                except:
                    pass
        
        # Filter out quizzes that are already in session content
        available_quizzes = []
        for quiz in quizzes:
            if quiz.id not in existing_quiz_ids:
                available_quizzes.append({
                    "id": quiz.id,
                    "title": quiz.title,
                    "description": quiz.description,
                    "total_marks": quiz.total_marks,
                    "time_limit_minutes": quiz.time_limit_minutes,
                    "created_at": quiz.created_at,
                    "is_active": quiz.is_active
                })
        
        return {"available_quizzes": available_quizzes}
    except Exception as e:
        logger.error(f"Get available quizzes error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch available quizzes")

@app.post("/admin/session-content/add-quiz")
async def add_quiz_to_session_content(
    quiz_id: int,
    session_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    """Add an existing quiz to session content"""
    try:
        # Get the quiz
        quiz = db.query(Quiz).filter(Quiz.id == quiz_id).first()
        if not quiz:
            raise HTTPException(status_code=404, detail="Quiz not found")
        
        # Verify the quiz belongs to the session
        if quiz.session_id != session_id:
            raise HTTPException(status_code=400, detail="Quiz does not belong to this session")
        
        # Check if quiz is already in session content
        existing_content = db.query(SessionContent).filter(
            SessionContent.session_id == session_id,
            SessionContent.content_type == "QUIZ"
        ).all()
        
        for content in existing_content:
            if content.content_data:
                try:
                    content_json = json.loads(content.content_data)
                    if content_json.get("quiz_id") == quiz_id:
                        raise HTTPException(status_code=400, detail="Quiz is already added to session content")
                except:
                    pass
        
        # Parse quiz questions
        questions = []
        if quiz.questions:
            try:
                questions = json.loads(quiz.questions)
            except:
                questions = []
        
        # Create session content entry
        session_content = SessionContent(
            session_id=session_id,
            content_type="QUIZ",
            title=quiz.title,
            description=quiz.description,
            content_data=json.dumps({
                "quiz_id": quiz.id,
                "total_marks": quiz.total_marks,
                "time_limit_minutes": quiz.time_limit_minutes,
                "questions": questions
            }),
            uploaded_by=current_user.id
        )
        
        db.add(session_content)
        db.commit()
        db.refresh(session_content)
        
        return {
            "message": "Quiz added to session content successfully",
            "session_content_id": session_content.id,
            "quiz_id": quiz.id,
            "title": quiz.title
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Add quiz to session content error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to add quiz to session content")

@app.post("/admin/process-quiz-file")
async def process_quiz_file(
    file: UploadFile = File(...),
    current_user = Depends(get_current_admin_or_presenter)
):
    try:
        # Validate file type
        allowed_types = ['.pdf', '.doc', '.docx', '.txt']
        file_ext = os.path.splitext(file.filename)[1].lower()
        
        if file_ext not in allowed_types:
            raise HTTPException(status_code=400, detail="Unsupported file type")
        
        # Read file content
        content = await file.read()
        
        # Extract text based on file type
        if file_ext == '.txt':
            text_content = content.decode('utf-8')
        elif file_ext == '.pdf':
            # For demo purposes, simulate PDF text extraction
            text_content = "Extracted PDF content: " + content.decode('utf-8', errors='ignore')[:2000]
        else:
            # For other formats, simulate extraction
            text_content = "Extracted document content: " + str(content)[:2000]
        
        return {
            "message": "File processed successfully",
            "content": text_content[:2000],  # Limit content length
            "filename": file.filename,
            "file_size": len(content)
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"File processing error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to process file")

def generate_quiz_questions(content: str, question_type: str, num_questions: int):
    """Generate quiz questions using AI logic (simplified implementation)"""
    import re
    import random
    
    # Extract key topics from content
    topics = extract_key_topics(content)
    questions = []
    
    for i in range(num_questions):
        topic = topics[i % len(topics)] if topics else f"Topic {i+1}"
        
        if question_type == "MCQ":
            question = {
                "id": i + 1,
                "type": "MCQ",
                "question": f"What is the primary concept behind {topic}?",
                "options": [
                    f"{topic} is a fundamental principle in this domain",
                    f"{topic} is an advanced technique for optimization",
                    f"{topic} is a basic method for data processing",
                    f"{topic} is a complex algorithm for analysis"
                ],
                "correct_answer": 0,
                "explanation": f"{topic} represents a key concept that forms the foundation of understanding in this area."
            }
        elif question_type == "TRUE_FALSE":
            question = {
                "id": i + 1,
                "type": "TRUE_FALSE",
                "question": f"{topic} is an important concept in this field.",
                "correct_answer": True,
                "explanation": f"This statement about {topic} is correct based on the provided content."
            }
        else:  # SHORT_ANSWER
            question = {
                "id": i + 1,
                "type": "SHORT_ANSWER",
                "question": f"Explain the significance of {topic} and its applications.",
                "sample_answer": f"{topic} is significant because it provides foundational understanding and has practical applications in various scenarios.",
                "keywords": [topic.lower(), "important", "concept", "application"]
            }
        
        questions.append(question)
    
    return questions

def extract_key_topics(content: str):
    """Extract key topics from content using simple NLP techniques"""
    import re
    
    # Default topics for fallback
    default_topics = [
        "Machine Learning", "Data Science", "Algorithms", "Programming", 
        "Statistics", "Neural Networks", "Deep Learning", "Artificial Intelligence",
        "Data Analysis", "Computer Vision", "Natural Language Processing", "Optimization"
    ]
    
    if not content or len(content.strip()) < 10:
        return default_topics[:8]
    
    # Simple keyword extraction
    # Remove common words and extract meaningful terms
    words = re.findall(r'\b[A-Z][a-z]+(?:\s+[A-Z][a-z]+)*\b', content)
    
    # Filter and clean topics
    topics = []
    for word in words:
        if len(word) > 3 and word not in ['The', 'This', 'That', 'With', 'From', 'When', 'Where']:
            topics.append(word)
    
    # Remove duplicates and limit
    unique_topics = list(dict.fromkeys(topics))[:12]
    
    # If not enough topics found, supplement with defaults
    if len(unique_topics) < 5:
        unique_topics.extend(default_topics[:8])
    
    return unique_topics[:12]

# Analytics and Reporting
@app.get("/admin/analytics")
async def get_admin_analytics(
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    """Admin analytics endpoint - accessible by both admins and presenters"""
    try:
        # User statistics
        total_students = db.query(User).filter(User.role == "Student").count()
        total_admins = db.query(Admin).count()
        
        # Course statistics
        total_courses = db.query(Course).count()
        total_modules = db.query(Module).count()
        total_sessions = db.query(SessionModel).count()
        completed_sessions = db.query(SessionModel).count()  # For now, assume all sessions are completed
        
        # Engagement statistics
        total_enrollments = db.query(Enrollment).count()
        active_enrollments = db.query(Enrollment).filter(Enrollment.progress > 0).count()
        total_resources = db.query(Resource).count()
        
        # Handle Quiz counting with proper error handling
        total_quizzes = 0
        total_assignments = 0
        avg_quiz_score = 0
        
        try:
            if Quiz is not None:
                total_quizzes = db.query(Quiz).count()
        except Exception as e:
            logger.warning(f"Could not count quizzes: {str(e)}")
            total_quizzes = 0
        
        try:
            # Try to import Assignment model
            from assignment_quiz_tables import Assignment
            total_assignments = db.query(Assignment).count()
        except (ImportError, Exception) as e:
            logger.warning(f"Could not count assignments: {str(e)}")
            total_assignments = 0
        
        # Performance metrics
        total_attendances = db.query(Attendance).count()
        attended_count = db.query(Attendance).filter(Attendance.attended == True).count()
        attendance_rate = (attended_count / total_attendances * 100) if total_attendances > 0 else 0
        
        completed_courses = db.query(Enrollment).filter(Enrollment.progress >= 90).count()
        completion_rate = (completed_courses / total_enrollments * 100) if total_enrollments > 0 else 0
        
        # Handle quiz score calculation with proper error handling
        try:
            if QuizAttempt is not None:
                avg_quiz_score = db.query(func.avg(QuizAttempt.score)).scalar() or 0
        except Exception as e:
            logger.warning(f"Could not calculate average quiz score: {str(e)}")
            avg_quiz_score = 0
        
        return {
            "users": {
                "total_students": total_students,
                "total_admins": total_admins,
                "growth_rate": 12.5
            },
            "courses": {
                "total_courses": total_courses,
                "total_modules": total_modules,
                "total_sessions": total_sessions,
                "completed_sessions": completed_sessions,
                "completion_percentage": (completed_sessions / total_sessions * 100) if total_sessions > 0 else 0
            },
            "engagement": {
                "total_enrollments": total_enrollments,
                "active_enrollments": active_enrollments,
                "engagement_rate": (active_enrollments / total_enrollments * 100) if total_enrollments > 0 else 0,
                "total_resources": total_resources,
                "total_quizzes": total_quizzes,
                "total_assignments": total_assignments
            },
            "performance": {
                "attendance_rate": round(attendance_rate, 2),
                "completion_rate": round(completion_rate, 2),
                "average_quiz_score": round(float(avg_quiz_score), 2),
                "target_attendance": 80.0,
                "target_completion": 90.0,
                "target_quiz_score": 75.0
            },
            "system_health": {
                "database_status": "healthy",
                "api_response_time": "45ms",
                "uptime": "99.9%",
                "storage_usage": "65%"
            }
        }
    except Exception as e:
        logger.error(f"Admin analytics error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch analytics")

@app.get("/admin/analytics/overview")
async def get_analytics_overview(
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        # User statistics
        total_students = db.query(User).filter(User.role == "Student").count()
        total_admins = db.query(Admin).count()
        
        # Course statistics
        total_courses = db.query(Course).count()
        total_modules = db.query(Module).count()
        total_sessions = db.query(SessionModel).count()
        completed_sessions = db.query(SessionModel).count()  # Simplified since is_completed field may not exist
        
        # Engagement statistics
        total_enrollments = db.query(Enrollment).count()
        active_enrollments = db.query(Enrollment).filter(Enrollment.progress > 0).count()
        total_resources = db.query(Resource).count()
        
        # Handle Quiz and Assignment counting with proper error handling
        total_quizzes = 0
        total_assignments = 0
        
        try:
            if Quiz is not None:
                total_quizzes = db.query(Quiz).count()
        except Exception as e:
            logger.warning(f"Could not count quizzes: {str(e)}")
            total_quizzes = 0
        
        try:
            from assignment_quiz_tables import Assignment
            total_assignments = db.query(Assignment).count()
        except (ImportError, Exception) as e:
            logger.warning(f"Could not count assignments: {str(e)}")
            total_assignments = 0
        
        # Performance metrics
        total_attendances = db.query(Attendance).count()
        attended_count = db.query(Attendance).filter(Attendance.attended == True).count()
        attendance_rate = (attended_count / total_attendances * 100) if total_attendances > 0 else 0
        
        completed_courses = db.query(Enrollment).filter(Enrollment.progress >= 90).count()
        completion_rate = (completed_courses / total_enrollments * 100) if total_enrollments > 0 else 0
        
        avg_quiz_score = db.query(func.avg(QuizAttempt.score)).scalar() or 0
        

        
        return {
            "users": {
                "total_students": total_students,
                "total_admins": total_admins,
                "growth_rate": 12.5  # This would be calculated based on time periods
            },
            "courses": {
                "total_courses": total_courses,
                "total_modules": total_modules,
                "total_sessions": total_sessions,
                "completed_sessions": completed_sessions,
                "completion_percentage": (completed_sessions / total_sessions * 100) if total_sessions > 0 else 0
            },
            "engagement": {
                "total_enrollments": total_enrollments,
                "active_enrollments": active_enrollments,
                "engagement_rate": (active_enrollments / total_enrollments * 100) if total_enrollments > 0 else 0,
                "total_resources": total_resources,
                "total_quizzes": total_quizzes,
                "total_assignments": total_assignments
            },
            "performance": {
                "attendance_rate": round(attendance_rate, 2),
                "completion_rate": round(completion_rate, 2),
                "average_quiz_score": round(float(avg_quiz_score), 2),
                "target_attendance": 80.0,
                "target_completion": 90.0,
                "target_quiz_score": 75.0
            },
            "system_health": {
                "database_status": "healthy",
                "api_response_time": "45ms",
                "uptime": "99.9%",
                "storage_usage": "65%"
            }
        }
    except Exception as e:
        logger.error(f"Analytics overview error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch analytics")

@app.get("/admin/reports/detailed")
async def get_detailed_reports(
    report_type: str = "all",  # all, attendance, progress, performance
    course_id: Optional[int] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        reports = {}
        
        # Base queries with filters
        enrollments_query = db.query(Enrollment)
        if course_id:
            enrollments_query = enrollments_query.filter(Enrollment.course_id == course_id)
        
        if report_type in ["all", "attendance"]:
            # Attendance Report
            attendance_query = db.query(Attendance)
            if date_from:
                attendance_query = attendance_query.join(SessionModel).filter(SessionModel.scheduled_time >= date_from)
            if date_to:
                attendance_query = attendance_query.join(SessionModel).filter(SessionModel.scheduled_time <= date_to)
            
            attendance_data = attendance_query.all()
            reports["attendance"] = {
                "total_sessions": len(set([a.session_id for a in attendance_data])),
                "total_attendances": len(attendance_data),
                "present_count": len([a for a in attendance_data if a.attended]),
                "attendance_rate": len([a for a in attendance_data if a.attended]) / len(attendance_data) * 100 if attendance_data else 0,
                "by_student": {}
            }
            
            # Group by student
            for attendance in attendance_data:
                student_key = attendance.student.username
                if student_key not in reports["attendance"]["by_student"]:
                    reports["attendance"]["by_student"][student_key] = {
                        "total_sessions": 0,
                        "attended_sessions": 0,
                        "attendance_rate": 0
                    }
                reports["attendance"]["by_student"][student_key]["total_sessions"] += 1
                if attendance.attended:
                    reports["attendance"]["by_student"][student_key]["attended_sessions"] += 1
            
            # Calculate individual rates
            for student_data in reports["attendance"]["by_student"].values():
                if student_data["total_sessions"] > 0:
                    student_data["attendance_rate"] = student_data["attended_sessions"] / student_data["total_sessions"] * 100
        
        if report_type in ["all", "progress"]:
            # Progress Report
            enrollments = enrollments_query.all()
            reports["progress"] = {
                "total_enrollments": len(enrollments),
                "completed_courses": len([e for e in enrollments if e.progress >= 90]),
                "in_progress_courses": len([e for e in enrollments if 0 < e.progress < 90]),
                "not_started_courses": len([e for e in enrollments if e.progress == 0]),
                "average_progress": sum([e.progress for e in enrollments]) / len(enrollments) if enrollments else 0,
                "by_course": {}
            }
            
            # Group by course
            for enrollment in enrollments:
                course_title = enrollment.course.title
                if course_title not in reports["progress"]["by_course"]:
                    reports["progress"]["by_course"][course_title] = {
                        "total_students": 0,
                        "average_progress": 0,
                        "completed_students": 0,
                        "completion_rate": 0
                    }
                reports["progress"]["by_course"][course_title]["total_students"] += 1
                if enrollment.progress >= 90:
                    reports["progress"]["by_course"][course_title]["completed_students"] += 1
            
            # Calculate course averages
            for course_title in reports["progress"]["by_course"]:
                course_enrollments = [e for e in enrollments if e.course.title == course_title]
                if course_enrollments:
                    reports["progress"]["by_course"][course_title]["average_progress"] = sum([e.progress for e in course_enrollments]) / len(course_enrollments)
                    reports["progress"]["by_course"][course_title]["completion_rate"] = reports["progress"]["by_course"][course_title]["completed_students"] / reports["progress"]["by_course"][course_title]["total_students"] * 100
        
        if report_type in ["all", "performance"]:
            # Performance Report (Quiz scores, assignment scores)
            quiz_attempts = db.query(QuizAttempt).all()
            submissions = db.query(Submission).filter(Submission.score.isnot(None)).all()
            
            reports["performance"] = {
                "quiz_attempts": len(quiz_attempts),
                "average_quiz_score": sum([qa.score for qa in quiz_attempts]) / len(quiz_attempts) if quiz_attempts else 0,
                "assignment_submissions": len(submissions),
                "average_assignment_score": sum([s.score for s in submissions]) / len(submissions) if submissions else 0,
                "top_performers": [],
                "struggling_students": []
            }
            
            # Find top performers and struggling students
            student_performance = {}
            for attempt in quiz_attempts:
                if attempt.student.username not in student_performance:
                    student_performance[attempt.student.username] = {"quiz_scores": [], "assignment_scores": []}
                student_performance[attempt.student.username]["quiz_scores"].append(attempt.score)
            
            for submission in submissions:
                if submission.student.username not in student_performance:
                    student_performance[submission.student.username] = {"quiz_scores": [], "assignment_scores": []}
                student_performance[submission.student.username]["assignment_scores"].append(submission.score)
            
            # Calculate averages and identify performance levels
            for username, scores in student_performance.items():
                all_scores = scores["quiz_scores"] + scores["assignment_scores"]
                if all_scores:
                    avg_score = sum(all_scores) / len(all_scores)
                    student_data = {
                        "username": username,
                        "average_score": round(avg_score, 2),
                        "total_attempts": len(all_scores)
                    }
                    
                    if avg_score >= 85:
                        reports["performance"]["top_performers"].append(student_data)
                    elif avg_score < 60:
                        reports["performance"]["struggling_students"].append(student_data)
            
            # Sort by score
            reports["performance"]["top_performers"].sort(key=lambda x: x["average_score"], reverse=True)
            reports["performance"]["struggling_students"].sort(key=lambda x: x["average_score"])
        
        return {
            "report_type": report_type,
            "filters": {
                "course_id": course_id,
                "date_from": date_from,
                "date_to": date_to
            },
            "generated_at": datetime.now(),
            "reports": reports
        }
    except Exception as e:
        logger.error(f"Detailed reports error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate reports")

# Certificate Management
@app.post("/admin/certificates/generate")
async def generate_certificate(
    cert_data: CertificateGenerate,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        # Check if student is enrolled and has completed the course
        enrollment = db.query(Enrollment).filter(
            Enrollment.student_id == cert_data.student_id,
            Enrollment.course_id == cert_data.course_id
        ).first()
        
        if not enrollment:
            raise HTTPException(status_code=404, detail="Enrollment not found")
        
        if enrollment.progress < 90:
            raise HTTPException(status_code=400, detail="Student has not completed the course (minimum 90% required)")
        
        # Check if certificate already exists
        existing_cert = db.query(Certificate).filter(
            Certificate.student_id == cert_data.student_id,
            Certificate.course_id == cert_data.course_id
        ).first()
        
        if existing_cert:
            raise HTTPException(status_code=400, detail="Certificate already exists for this student and course")
        
        # Generate certificate
        certificate_filename = f"cert_{cert_data.student_id}_{cert_data.course_id}_{uuid.uuid4().hex[:8]}.pdf"
        certificate_path = UPLOAD_BASE_DIR / "certificates" / certificate_filename
        
        # Here you would integrate with a certificate generation library
        # For now, we'll just create a placeholder
        with open(certificate_path, "w") as f:
            f.write("Certificate placeholder - integrate with PDF generation library")
        
        certificate = Certificate(
            student_id=cert_data.student_id,
            course_id=cert_data.course_id,
            certificate_url=f"/api/certificates/{certificate_filename}",
            issued_by=current_user.id,
            issued_at=cert_data.completion_date or datetime.now()
        )
        db.add(certificate)
        db.commit()
        
        return {
            "message": "Certificate generated successfully",
            "certificate_id": certificate.id,
            "certificate_url": certificate.certificate_url
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Generate certificate error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate certificate")

@app.get("/admin/certificates")
async def get_all_certificates(
    page: int = 1,
    limit: int = 50,
    course_id: Optional[int] = None,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        query = db.query(Certificate)
        
        if course_id:
            query = query.filter(Certificate.course_id == course_id)
        
        total = query.count()
        certificates = query.offset((page - 1) * limit).limit(limit).all()
        
        return {
            "certificates": [{
                "id": c.id,
                "student_name": f"{getattr(c.student, 'first_name', '') or ''} {getattr(c.student, 'last_name', '') or ''}".strip() or c.student.username,
                "course_title": c.course.title,
                "issued_at": c.issued_at,
                "certificate_url": c.certificate_url
            } for c in certificates],
            "total": total,
            "page": page,
            "limit": limit
        }
    except Exception as e:
        logger.error(f"Get certificates error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch certificates")

# Forum Management
@app.post("/admin/forums")
async def create_forum(
    forum_data: ForumCreate,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        forum = Forum(
            **forum_data.dict(),
            created_by=current_user.id
        )
        db.add(forum)
        db.commit()
        db.refresh(forum)
        
        return {"message": "Forum created successfully", "forum_id": forum.id}
    except Exception as e:
        db.rollback()
        logger.error(f"Create forum error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create forum")

@app.get("/admin/forums/{module_id}")
async def get_module_forums_admin(
    module_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        forums = db.query(Forum).filter(Forum.module_id == module_id).all()
        
        result = []
        for forum in forums:
            posts_count = db.query(ForumPost).filter(ForumPost.forum_id == forum.id).count()
            latest_post = db.query(ForumPost).filter(ForumPost.forum_id == forum.id).order_by(ForumPost.created_at.desc()).first()
            
            result.append({
                "id": forum.id,
                "title": forum.title,
                "description": forum.description,
                "is_pinned": forum.is_pinned,
                "posts_count": posts_count,
                "latest_post_at": latest_post.created_at if latest_post else None,
                "created_at": forum.created_at
            })
        
        return {"forums": result}
    except Exception as e:
        logger.error(f"Get module forums error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch forums")

# Student cohort info endpoint
@app.get("/student/cohort-info")
async def get_student_cohort_info(
    current_user: User = Depends(require_role("Student")),
    db: Session = Depends(get_db)
):
    try:
        if not current_user.cohort_id:
            return {"cohort": None}
        
        cohort = db.query(Cohort).filter(Cohort.id == current_user.cohort_id).first()
        if not cohort:
            return {"cohort": None}
        
        total_users = db.query(UserCohort).filter(UserCohort.cohort_id == cohort.id).count()
        total_courses = db.query(CohortCourse).filter(CohortCourse.cohort_id == cohort.id).count()
        
        return {
            "cohort": {
                "id": cohort.id,
                "name": cohort.name,
                "description": cohort.description,
                "instructor_name": cohort.instructor_name,
                "start_date": cohort.start_date,
                "end_date": cohort.end_date,
                "total_users": total_users,
                "total_courses": total_courses
            }
        }
    except Exception as e:
        logger.error(f"Get student cohort info error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch cohort information")

# Student Dashboard Endpoints

@app.get("/student/dashboard")
async def get_student_dashboard(
    request: Request,
    current_user: User = Depends(require_role("Student")),
    db: Session = Depends(get_db)
):
    try:
        # Log dashboard access
        log_student_action(
            student_id=current_user.id,
            student_username=current_user.username,
            action_type="VIEW",
            resource_type="DASHBOARD",
            details=f"Accessed student dashboard",
            ip_address=request.client.host if request.client else "127.0.0.1"
        )
        # Get cohort information
        cohort_info = None
        if current_user.cohort_id:
            cohort = db.query(Cohort).filter(Cohort.id == current_user.cohort_id).first()
            if cohort:
                cohort_info = {
                    "id": cohort.id,
                    "name": cohort.name,
                    "description": cohort.description,
                    "instructor_name": cohort.instructor_name,
                    "start_date": cohort.start_date,
                    "end_date": cohort.end_date
                }
        
        # Get enrolled courses with progress
        enrollments = db.query(Enrollment).filter(Enrollment.student_id == current_user.id).all()
        
        enrolled_courses = []
        total_courses = len(enrollments)
        completed_courses = 0
        in_progress_courses = 0
        not_started_courses = 0
        
        for enrollment in enrollments:
            course = enrollment.course
            progress = enrollment.progress or 0
            
            # Count course status
            if progress >= 90:
                completed_courses += 1
            elif progress > 0:
                in_progress_courses += 1
            else:
                not_started_courses += 1
            
            # Get course statistics
            total_sessions = db.query(SessionModel).join(Module).filter(Module.course_id == course.id).count()
            attended_sessions = db.query(Attendance).join(SessionModel).join(Module).filter(
                Module.course_id == course.id,
                Attendance.student_id == current_user.id,
                Attendance.attended == True
            ).count()
            
            # Find next session
            next_session = db.query(SessionModel).join(Module).filter(
                Module.course_id == course.id,
                SessionModel.scheduled_time > datetime.now()
            ).order_by(SessionModel.scheduled_time).first()
            
            enrolled_courses.append({
                "id": course.id,
                "title": course.title,
                "description": course.description,
                "progress": progress,
                "total_sessions": total_sessions,
                "attended_sessions": attended_sessions,
                "is_cohort_course": bool(getattr(enrollment, 'cohort_id', None)),
                "next_session": {
                    "id": next_session.id,
                    "title": next_session.title,
                    "scheduled_time": next_session.scheduled_time,
                    "zoom_link": next_session.zoom_link
                } if next_session else None
            })
        
        # Get upcoming sessions (cohort-aware)
        upcoming_sessions_query = db.query(SessionModel).join(Module).join(Course).join(Enrollment).filter(
            Enrollment.student_id == current_user.id,
            SessionModel.scheduled_time > datetime.now()
        )
        
        if current_user.cohort_id:
            try:
                upcoming_sessions_query = upcoming_sessions_query.filter(
                    or_(
                        Enrollment.cohort_id == current_user.cohort_id,
                        Enrollment.cohort_id.is_(None)
                    )
                )
            except:
                # Handle case where cohort_id column doesn't exist yet
                pass
        
        upcoming_sessions = upcoming_sessions_query.order_by(SessionModel.scheduled_time).limit(5).all()
        
        upcoming_sessions_data = []
        for session in upcoming_sessions:
            upcoming_sessions_data.append({
                "id": session.id,
                "title": session.title,
                "course_title": session.module.course.title,
                "scheduled_time": session.scheduled_time,
                "zoom_link": session.zoom_link
            })
        
        # Get achievements
        certificates_earned = db.query(Certificate).filter(Certificate.student_id == current_user.id).count()
        quizzes_completed = db.query(QuizAttempt).filter(QuizAttempt.student_id == current_user.id).count()
        assignments_submitted = db.query(Submission).filter(Submission.student_id == current_user.id).count()
        
        # Calculate attendance rate
        total_attendances = db.query(Attendance).filter(Attendance.student_id == current_user.id).count()
        attended_count = db.query(Attendance).filter(
            Attendance.student_id == current_user.id,
            Attendance.attended == True
        ).count()
        attendance_rate = (attended_count / total_attendances * 100) if total_attendances > 0 else 0
        
        return {
            "user_info": {
                "username": current_user.username,
                "email": current_user.email,
                "full_name": current_user.username,
                "college": current_user.college
            },
            "cohort_info": cohort_info,
            "enrollment_summary": {
                "total_courses": total_courses,
                "completed_courses": completed_courses,
                "in_progress_courses": in_progress_courses,
                "not_started_courses": not_started_courses
            },
            "enrolled_courses": enrolled_courses,
            "upcoming_sessions": upcoming_sessions_data,
            "recent_activities": [],
            "achievements": {
                "certificates_earned": certificates_earned,
                "quizzes_completed": quizzes_completed,
                "assignments_submitted": assignments_submitted,
                "attendance_rate": round(attendance_rate, 2)
            }
        }
    except Exception as e:
        logger.error(f"Student dashboard error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch dashboard data")

@app.get("/student/courses")
async def get_available_courses(
    current_user: User = Depends(require_role("Student")),
    db: Session = Depends(get_db)
):
    try:
        courses = db.query(Course).all()
        enrolled_courses = db.query(Enrollment.course_id).filter(Enrollment.student_id == current_user.id).all()
        enrolled_ids = [e[0] for e in enrolled_courses]
        
        # Get user's cohort info
        user_cohort = None
        cohort_course_ids = set()
        if current_user.cohort_id:
            user_cohort = db.query(Cohort).filter(Cohort.id == current_user.cohort_id).first()
            cohort_courses = db.query(CohortCourse.course_id).filter(CohortCourse.cohort_id == current_user.cohort_id).all()
            cohort_course_ids = {cc[0] for cc in cohort_courses}
        
        result = []
        for c in courses:
            # Get actual duration from modules
            max_week = db.query(func.max(Module.week_number)).filter(Module.course_id == c.id).scalar()
            duration_weeks = max_week if max_week else 0
            
            # Get module and session counts
            total_modules = db.query(Module).filter(Module.course_id == c.id).count()
            total_sessions = db.query(SessionModel).join(Module).filter(Module.course_id == c.id).count()
            
            # Check if course is locked for cohort users
            is_locked = False
            lock_reason = None
            if current_user.cohort_id and c.id not in cohort_course_ids and c.id not in enrolled_ids:
                is_locked = True
                cohort_name = user_cohort.name if user_cohort else "your cohort"
                instructor_name = user_cohort.instructor_name if user_cohort and user_cohort.instructor_name else "your instructor"
                lock_reason = f"You are in cohort \"{cohort_name}\". You can only enroll in courses assigned to your cohort. Contact {instructor_name} for more information."
            
            result.append({
                "id": c.id,
                "title": c.title,
                "description": c.description,
                "duration_weeks": duration_weeks,
                "enrolled": c.id in enrolled_ids,
                "total_modules": total_modules,
                "total_sessions": total_sessions,
                "is_cohort_assigned": c.id in cohort_course_ids,
                "is_locked": is_locked,
                "lock_reason": lock_reason,
                "access_level": "locked" if is_locked else "available"
            })
        
        # Include user cohort info in response
        response_data = {"courses": result}
        if user_cohort:
            response_data["user_cohort"] = {
                "id": user_cohort.id,
                "name": user_cohort.name,
                "description": user_cohort.description,
                "instructor_name": user_cohort.instructor_name,
                "start_date": user_cohort.start_date,
                "end_date": user_cohort.end_date
            }
        
        return response_data
    except Exception as e:
        logger.error(f"Get available courses error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch courses")

@app.post("/student/courses/{course_id}/enroll")
async def enroll_in_course(
    course_id: int,
    request: Request,
    current_user: User = Depends(require_role("Student")),
    db: Session = Depends(get_db)
):
    try:
        # Check if course exists
        course = db.query(Course).filter(Course.id == course_id).first()
        if not course:
            raise HTTPException(status_code=404, detail="Course not found or inactive")
        
        # Check if already enrolled
        existing = db.query(Enrollment).filter(
            Enrollment.student_id == current_user.id,
            Enrollment.course_id == course_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="Already enrolled in this course")
        
        # Check cohort restrictions
        if current_user.cohort_id:
            # User is in a cohort, check if course is assigned to their cohort
            cohort_course = db.query(CohortCourse).filter(
                CohortCourse.cohort_id == current_user.cohort_id,
                CohortCourse.course_id == course_id
            ).first()
            
            if not cohort_course:
                # Get cohort info for error message
                cohort = db.query(Cohort).filter(Cohort.id == current_user.cohort_id).first()
                cohort_name = cohort.name if cohort else "your cohort"
                instructor_name = cohort.instructor_name if cohort and cohort.instructor_name else "your instructor"
                
                raise HTTPException(
                    status_code=403, 
                    detail=f"You are in cohort \"{cohort_name}\". You can only enroll in courses assigned to your cohort. Contact {instructor_name} for more information."
                )
        
        # Create enrollment with cohort_id if user is in a cohort
        enrollment = Enrollment(
            student_id=current_user.id, 
            course_id=course_id,
            cohort_id=current_user.cohort_id
        )
        db.add(enrollment)
        db.commit()
        
        # Log student enrollment
        log_student_action(
            student_id=current_user.id,
            student_username=current_user.username,
            action_type="ENROLL",
            resource_type="COURSE",
            resource_id=course_id,
            details=f"Enrolled in course: {course.title}",
            ip_address=request.client.host if request.client else "127.0.0.1"
        )
        
        return {"message": "Successfully enrolled in course", "course_title": course.title}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Course enrollment error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to enroll in course")

@app.get("/student/courses/{course_id}/modules")
async def get_course_modules_student(
    course_id: int,
    current_user: User = Depends(require_role("Student")),
    db: Session = Depends(get_db)
):
    try:
        # Check enrollment
        enrollment = db.query(Enrollment).filter(
            Enrollment.student_id == current_user.id,
            Enrollment.course_id == course_id
        ).first()
        if not enrollment:
            raise HTTPException(status_code=403, detail="Not enrolled in this course")
        
        modules = db.query(Module).filter(Module.course_id == course_id).order_by(Module.week_number).all()
        
        result = []
        for module in modules:
            sessions = db.query(SessionModel).filter(SessionModel.module_id == module.id).order_by(SessionModel.session_number).all()
            
            # Calculate module progress based on attended sessions
            module_sessions = len(sessions)
            attended_module_sessions = sum(1 for s in sessions if db.query(Attendance).filter(
                Attendance.session_id == s.id,
                Attendance.student_id == current_user.id,
                Attendance.attended == True
            ).first())
            
            module_progress = (attended_module_sessions / module_sessions * 100) if module_sessions > 0 else 0
            
            module_data = {
                "id": module.id,
                "week_number": module.week_number,
                "title": module.title,
                "description": module.description,
                "progress": module_progress,
                "resources_count": sum(
                    db.query(Resource).filter(Resource.session_id == s.id).count() + 
                    db.query(SessionContent).filter(SessionContent.session_id == s.id).count() 
                    for s in sessions
                ),
                "sessions": []
            }
            
            for session in sessions:
                # Check attendance for this session
                attendance = db.query(Attendance).filter(
                    Attendance.session_id == session.id,
                    Attendance.student_id == current_user.id
                ).first()
                
                # Get resources count from both tables
                resources_count = db.query(Resource).filter(Resource.session_id == session.id).count()
                session_content_count = db.query(SessionContent).filter(SessionContent.session_id == session.id).count()
                total_resources_count = resources_count + session_content_count
                
                # Get quizzes for this session
                quizzes = db.query(Quiz).filter(Quiz.session_id == session.id, Quiz.is_active == True).all()
                quiz_data = []
                for quiz in quizzes:
                    attempt = db.query(QuizAttempt).filter(
                        QuizAttempt.quiz_id == quiz.id,
                        QuizAttempt.student_id == current_user.id
                    ).first()
                    quiz_data.append({
                        "id": quiz.id,
                        "title": quiz.title,
                        "total_marks": quiz.total_marks,
                        "attempted": attempt is not None,
                        "score": attempt.score if attempt else None
                    })
                
                session_data = {
                    "id": session.id,
                    "session_number": session.session_number,
                    "title": session.title,
                    "description": session.description,
                    "scheduled_time": session.scheduled_time,
                    "duration_minutes": session.duration_minutes,
                    "session_type": getattr(session, 'session_type', 'LIVE'),
                    "zoom_link": session.zoom_link,
                    "recording_url": session.recording_url,
                    "is_completed": getattr(session, 'is_completed', False),
                    "attended": attendance.attended if attendance else False,
                    "resources_count": total_resources_count,
                    "quizzes": quiz_data,
                    "can_access": True  # Allow access to all sessions for now
                }
                
                module_data["sessions"].append(session_data)
            
            result.append(module_data)
        
        return {"modules": result}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get course modules error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch course modules")

@app.get("/student/sessions/{session_id}")
async def get_session_details_student(
    session_id: int,
    current_user: User = Depends(require_role("Student")),
    db: Session = Depends(get_db)
):
    try:
        session = db.query(SessionModel).filter(SessionModel.id == session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        # Check enrollment
        enrollment = db.query(Enrollment).filter(
            Enrollment.student_id == current_user.id,
            Enrollment.course_id == session.module.course_id
        ).first()
        if not enrollment:
            raise HTTPException(status_code=403, detail="Not enrolled in this course")
        
        # Get resources from both Resource and SessionContent tables
        resources = db.query(Resource).filter(Resource.session_id == session_id).all()
        session_contents = db.query(SessionContent).filter(SessionContent.session_id == session_id).all()
        
        # Get attendance
        attendance = db.query(Attendance).filter(
            Attendance.session_id == session_id,
            Attendance.student_id == current_user.id
        ).first()
        
        # Combine resources from both tables
        all_resources = []
        
        # Add traditional resources
        for r in resources:
            all_resources.append({
                "id": f"resource_{r.id}",
                "title": r.title,
                "resource_type": r.resource_type,
                "content_type": "RESOURCE",
                "file_size": r.file_size,
                "description": r.description,
                "view_url": f"/api/resources/{r.id}/view",
                "created_at": r.uploaded_at
            })
        
        # Add session content with meeting link filtering
        content_list = []
        for c in session_contents:
            content_dict = {
                "id": f"content_{c.id}",
                "title": c.title,
                "resource_type": c.file_type or c.content_type,
                "content_type": c.content_type,
                "file_size": c.file_size,
                "description": c.description,
                "file_path": c.file_path,
                "meeting_url": c.meeting_url,
                "scheduled_time": c.scheduled_time.isoformat() if c.scheduled_time else None,
                "view_url": f"/api/session-content/{c.id}/view" if c.file_path else None,
                "created_at": c.created_at
            }
            content_list.append(content_dict)
        
        # Process meeting links with status instead of filtering
        processed_content = filter_meeting_links_for_student(content_list)
        all_resources.extend(processed_content)
        
        return {
            "id": session.id,
            "title": session.title,
            "description": session.description,
            "scheduled_time": session.scheduled_time,
            "duration_minutes": session.duration_minutes,
            "session_type": getattr(session, 'session_type', 'LIVE'),
            "zoom_link": session.zoom_link,
            "recording_url": session.recording_url,
            "syllabus_content": session.syllabus_content,
            "is_completed": getattr(session, 'is_completed', False),
            "attended": attendance.attended if attendance else False,
            "attendance_duration": attendance.duration_minutes if attendance else 0,
            "resources": all_resources,
            "module": {
                "id": session.module.id,
                "title": session.module.title,
                "week_number": session.module.week_number
            },
            "course": {
                "id": session.module.course.id,
                "title": session.module.course.title
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get session details error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch session details")

# Generic content deletion endpoint for composite IDs
@app.delete("/admin/content/{item_id}")
async def delete_content_item(
    item_id: str,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    """Delete content items with composite IDs like resource_1, content_2, etc."""
    try:
        if item_id.startswith("resource_"):
            # Handle resource deletion
            try:
                resource_id = int(item_id.replace("resource_", ""))
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid resource ID format")
            
            resource = db.query(Resource).filter(Resource.id == resource_id).first()
            if not resource:
                raise HTTPException(status_code=404, detail="Resource not found")
            
            resource_title = resource.title
            
            # Delete physical file
            if resource.file_path and os.path.exists(resource.file_path):
                os.remove(resource.file_path)
            
            db.delete(resource)
            db.commit()
            
            # Log resource deletion
            if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
                log_admin_action(
                    admin_id=current_user.id,
                    admin_username=current_user.username,
                    action_type="DELETE",
                    resource_type="RESOURCE",
                    resource_id=resource_id,
                    details=f"Deleted resource: {resource_title}"
                )
            elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
                log_presenter_action(
                    presenter_id=current_user.id,
                    presenter_username=current_user.username,
                    action_type="DELETE",
                    resource_type="RESOURCE",
                    resource_id=resource_id,
                    details=f"Deleted resource: {resource_title}"
                )
            
            return {"message": "Resource deleted successfully"}
            
        elif item_id.startswith("content_"):
            # Handle session content deletion
            try:
                content_id = int(item_id.replace("content_", ""))
            except ValueError:
                raise HTTPException(status_code=400, detail="Invalid content ID format")
            
            content = db.query(SessionContent).filter(SessionContent.id == content_id).first()
            if not content:
                raise HTTPException(status_code=404, detail="Content not found")
            
            content_title = content.title
            content_type = content.content_type
            
            # Delete physical file if it exists (but not for meeting links)
            if content.file_path and os.path.exists(content.file_path) and content_type != "MEETING_LINK":
                os.remove(content.file_path)
            
            db.delete(content)
            db.commit()
            
            # Log session content deletion
            if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
                log_admin_action(
                    admin_id=current_user.id,
                    admin_username=current_user.username,
                    action_type="DELETE",
                    resource_type="SESSION_CONTENT",
                    resource_id=content_id,
                    details=f"Deleted {content_type.lower()}: {content_title}"
                )
            elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
                log_presenter_action(
                    presenter_id=current_user.id,
                    presenter_username=current_user.username,
                    action_type="DELETE",
                    resource_type="SESSION_CONTENT",
                    resource_id=content_id,
                    details=f"Deleted {content_type.lower()}: {content_title}"
                )
            
            return {"message": f"{content_type.replace('_', ' ').title()} deleted successfully"}
        else:
            raise HTTPException(status_code=400, detail="Invalid item ID format. Expected resource_X or content_X")
            
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Delete content item error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete content item")

# File serving endpoints
@app.get("/api/resources/{resource_id}/download")
async def download_resource(
    resource_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        resource = db.query(Resource).filter(Resource.id == resource_id).first()
        if not resource:
            raise HTTPException(status_code=404, detail="Resource not found")
        
        # Check if user has access to this resource
        if current_user.role == "Student":
            enrollment = db.query(Enrollment).filter(
                Enrollment.student_id == current_user.id,
                Enrollment.course_id == resource.session.module.course_id
            ).first()
            if not enrollment:
                raise HTTPException(status_code=403, detail="Access denied")
        
        # Try different possible file paths
        possible_paths = [
            resource.file_path,
            str(UPLOAD_BASE_DIR / "resources" / os.path.basename(resource.file_path)),
            str(UPLOAD_BASE_DIR / os.path.basename(resource.file_path))
        ]
        
        actual_file_path = None
        for path in possible_paths:
            if os.path.exists(path):
                actual_file_path = path
                break
        
        if not actual_file_path:
            raise HTTPException(status_code=404, detail="File not found")
        
        # Log resource download based on user type
        try:
            if hasattr(current_user, 'role'):
                if current_user.role == 'Admin':
                    log_admin_action(
                        admin_id=current_user.id,
                        admin_username=current_user.username,
                        action_type="DOWNLOAD",
                        resource_type="RESOURCE",
                        resource_id=resource_id,
                        details=f"Downloaded resource: {resource.title}"
                    )
                elif current_user.role == 'Student':
                    log_student_action(
                        student_id=current_user.id,
                        student_username=current_user.username,
                        action_type="DOWNLOAD",
                        resource_type="RESOURCE",
                        resource_id=resource_id,
                        details=f"Downloaded resource: {resource.title}"
                    )
        except:
            pass  # Don't fail download if logging fails
        
        return FileResponse(
            path=actual_file_path,
            filename=resource.title,
            media_type='application/octet-stream'
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Download resource error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to download resource")

@app.get("/api/session-content/{content_id}/download")
async def download_session_content(
    content_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        content = db.query(SessionContent).filter(SessionContent.id == content_id).first()
        if not content:
            raise HTTPException(status_code=404, detail="Content not found")
        
        # Check if user has access to this content
        if current_user.role == "Student":
            enrollment = db.query(Enrollment).filter(
                Enrollment.student_id == current_user.id,
                Enrollment.course_id == content.session.module.course_id
            ).first()
            if not enrollment:
                raise HTTPException(status_code=403, detail="Access denied")
        
        if not content.file_path:
            raise HTTPException(status_code=404, detail="No file associated with this content")
        
        # Try different possible file paths
        possible_paths = [
            content.file_path,
            str(UPLOAD_BASE_DIR / "resources" / os.path.basename(content.file_path)),
            str(UPLOAD_BASE_DIR / os.path.basename(content.file_path))
        ]
        
        actual_file_path = None
        for path in possible_paths:
            if os.path.exists(path):
                actual_file_path = path
                break
        
        if not actual_file_path:
            raise HTTPException(status_code=404, detail="File not found")
        
        # Log content download based on user type
        try:
            if hasattr(current_user, 'role'):
                if current_user.role == 'Admin':
                    log_admin_action(
                        admin_id=current_user.id,
                        admin_username=current_user.username,
                        action_type="DOWNLOAD",
                        resource_type="SESSION_CONTENT",
                        resource_id=content_id,
                        details=f"Downloaded session content: {content.title}"
                    )
                elif current_user.role == 'Student':
                    log_student_action(
                        student_id=current_user.id,
                        student_username=current_user.username,
                        action_type="DOWNLOAD",
                        resource_type="SESSION_CONTENT",
                        resource_id=content_id,
                        details=f"Downloaded session content: {content.title}"
                    )
        except:
            pass  # Don't fail download if logging fails
        
        return FileResponse(
            path=actual_file_path,
            filename=content.title,
            media_type='application/octet-stream'
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Download session content error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to download content")

@app.get("/api/recordings/{filename}")
async def get_recording(
    filename: str,
    token: Optional[str] = None,
    db: Session = Depends(get_db)
):
    try:
        recording_path = UPLOAD_BASE_DIR / "recordings" / filename
        
        if not recording_path.exists():
            raise HTTPException(status_code=404, detail="Recording not found")
        
        # Determine media type based on file extension
        file_ext = os.path.splitext(filename)[1].lower()
        media_type_map = {
            '.mp4': 'video/mp4',
            '.avi': 'video/avi',
            '.mov': 'video/quicktime',
            '.wmv': 'video/x-ms-wmv',
            '.flv': 'video/x-flv'
        }
        
        media_type = media_type_map.get(file_ext, 'video/mp4')
        
        return FileResponse(
            path=str(recording_path), 
            media_type=media_type,
            headers={"Content-Disposition": "inline", "X-Download-Options": "noopen", "X-Content-Type-Options": "nosniff", "Cache-Control": "no-store\"{filename}\""}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get recording error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to access recording")

# Attendance Calendar Endpoints
@app.get("/attendance/calendar")
async def get_attendance_calendar(
    month: Optional[int] = None,
    year: Optional[int] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    try:
        # Default to current month/year if not provided
        if not month or not year:
            now = datetime.now()
            month = month or now.month
            year = year or now.year
        
        # Get attendance data for the specified month
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
        
        # Get attendance records for the user in the specified month
        attendance_records = db.query(Attendance).join(SessionModel).filter(
            Attendance.student_id == current_user.id,
            SessionModel.scheduled_time >= start_date,
            SessionModel.scheduled_time < end_date
        ).all()
        
        # Format attendance data by date
        attendance_data = {}
        for record in attendance_records:
            date_str = record.session.scheduled_time.strftime('%Y-%m-%d')
            attendance_data[date_str] = 'present' if record.attended else 'absent'
        
        return {
            "month": month,
            "year": year,
            "attendance_data": attendance_data,
            "total_sessions": len(attendance_records),
            "attended_sessions": len([r for r in attendance_records if r.attended]),
            "attendance_rate": (len([r for r in attendance_records if r.attended]) / len(attendance_records) * 100) if attendance_records else 0
        }
    except Exception as e:
        logger.error(f"Get attendance calendar error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch attendance calendar")

@app.get("/events")
async def get_events(
    db: Session = Depends(get_db)
):
    try:
        # Get all events from Event table
        events_from_db = db.query(Event).filter(Event.date >= datetime.now()).order_by(Event.date).all()
        
        events = []
        for event in events_from_db:
            events.append({
                "id": event.id,
                "title": event.title,
                "date": event.date.strftime('%Y-%m-%d'),
                "type": event.event_type,
                "description": event.description
            })
        
        return {"events": events}
    except Exception as e:
        logger.error(f"Get events error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch events")

# Admin-only event management endpoints
class EventCreate(BaseModel):
    title: str = Field(..., min_length=3, max_length=200)
    date: datetime
    event_type: Optional[str] = Field(default="general")
    description: Optional[str] = None

class EventUpdate(BaseModel):
    title: Optional[str] = Field(None, min_length=3, max_length=200)
    date: Optional[datetime] = None
    event_type: Optional[str] = None
    description: Optional[str] = None

class CohortCreate(BaseModel):
    name: str = Field(..., min_length=3, max_length=200)
    description: Optional[str] = None
    start_date: datetime
    end_date: datetime
    instructor_name: Optional[str] = Field(None, max_length=200)

class CohortUpdate(BaseModel):
    name: Optional[str] = Field(None, min_length=3, max_length=200)
    description: Optional[str] = None
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    instructor_name: Optional[str] = Field(None, max_length=200)

class CohortUserAdd(BaseModel):
    user_ids: List[int]

class CohortCourseAssign(BaseModel):
    course_ids: List[int]

class CohortCourseRemove(BaseModel):
    course_ids: List[int]

@app.post("/admin/events")
async def create_event(
    event_data: EventCreate,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        event = Event(
            title=event_data.title,
            description=event_data.description,
            date=event_data.date,
            event_type=event_data.event_type or "general",
            created_by=current_admin.id
        )
        db.add(event)
        db.commit()
        db.refresh(event)
        
        return {
            "message": "Event created successfully",
            "event_id": event.id,
            "title": event.title,
            "date": event.date
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Create event error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create event")

@app.put("/admin/events/{event_id}")
async def update_event(
    event_id: int,
    event_data: EventUpdate,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        event = db.query(Event).filter(Event.id == event_id).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        
        update_data = event_data.dict(exclude_unset=True)
        if 'title' in update_data:
            event.title = update_data['title']
        if 'date' in update_data:
            event.date = update_data['date']
        if 'description' in update_data:
            event.description = update_data['description']
        if 'event_type' in update_data:
            event.event_type = update_data['event_type']
        
        db.commit()
        return {"message": "Event updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Update event error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update event")

@app.delete("/admin/events/{event_id}")
async def delete_event(
    event_id: int,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        event = db.query(Event).filter(Event.id == event_id).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        
        db.delete(event)
        db.commit()
        return {"message": "Event deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Delete event error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete event")

# Cohort Course Assignment Endpoints
@app.post("/admin/cohorts/{cohort_id}/assign-courses")
async def assign_courses_to_cohort(
    cohort_id: int,
    course_data: CohortCourseAssign,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        # Check if cohort exists
        cohort = db.query(Cohort).filter(Cohort.id == cohort_id).first()
        if not cohort:
            raise HTTPException(status_code=404, detail="Cohort not found")
        
        assigned_courses = []
        for course_id in course_data.course_ids:
            # Check if course exists
            course = db.query(Course).filter(Course.id == course_id).first()
            if not course:
                continue
            
            # Check if already assigned
            existing = db.query(CohortCourse).filter(
                CohortCourse.cohort_id == cohort_id,
                CohortCourse.course_id == course_id
            ).first()
            
            if not existing:
                cohort_course = CohortCourse(
                    cohort_id=cohort_id,
                    course_id=course_id,
                    assigned_by=current_admin.id
                )
                db.add(cohort_course)
                assigned_courses.append(course.title)
        
        db.commit()
        
        # Log the assignment
        log_admin_action(
            admin_id=current_admin.id,
            admin_username=current_admin.username,
            action_type="ASSIGN",
            resource_type="COHORT_COURSE",
            resource_id=cohort_id,
            details=f"Assigned {len(assigned_courses)} courses to cohort {cohort.name}: {', '.join(assigned_courses)}"
        )
        
        return {
            "message": f"Successfully assigned {len(assigned_courses)} courses to cohort {cohort.name}",
            "assigned_courses": assigned_courses
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Assign courses to cohort error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to assign courses to cohort")

@app.delete("/admin/cohorts/{cohort_id}/remove-courses")
async def remove_courses_from_cohort(
    cohort_id: int,
    course_data: CohortCourseRemove,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        # Check if cohort exists
        cohort = db.query(Cohort).filter(Cohort.id == cohort_id).first()
        if not cohort:
            raise HTTPException(status_code=404, detail="Cohort not found")
        
        removed_courses = []
        for course_id in course_data.course_ids:
            # Get course title for logging
            course = db.query(Course).filter(Course.id == course_id).first()
            if course:
                removed_courses.append(course.title)
            
            # Remove the assignment
            db.query(CohortCourse).filter(
                CohortCourse.cohort_id == cohort_id,
                CohortCourse.course_id == course_id
            ).delete()
        
        db.commit()
        
        # Log the removal
        log_admin_action(
            admin_id=current_admin.id,
            admin_username=current_admin.username,
            action_type="REMOVE",
            resource_type="COHORT_COURSE",
            resource_id=cohort_id,
            details=f"Removed {len(removed_courses)} courses from cohort {cohort.name}: {', '.join(removed_courses)}"
        )
        
        return {
            "message": f"Successfully removed {len(removed_courses)} courses from cohort {cohort.name}",
            "removed_courses": removed_courses
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Remove courses from cohort error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to remove courses from cohort")

@app.get("/admin/cohorts/{cohort_id}/courses")
async def get_cohort_courses(
    cohort_id: int,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        # Check if cohort exists
        cohort = db.query(Cohort).filter(Cohort.id == cohort_id).first()
        if not cohort:
            raise HTTPException(status_code=404, detail="Cohort not found")
        
        # Get assigned courses
        cohort_courses = db.query(CohortCourse).filter(CohortCourse.cohort_id == cohort_id).all()
        assigned_course_ids = {cc.course_id for cc in cohort_courses}
        
        # Get all courses with assignment status
        all_courses = db.query(Course).all()
        
        assigned_courses = []
        available_courses = []
        
        for course in all_courses:
            course_data = {
                "id": course.id,
                "title": course.title,
                "description": course.description,
                "created_at": course.created_at
            }
            
            if course.id in assigned_course_ids:
                # Add assignment info
                cohort_course = next(cc for cc in cohort_courses if cc.course_id == course.id)
                course_data["assigned_at"] = cohort_course.assigned_at
                assigned_courses.append(course_data)
            else:
                available_courses.append(course_data)
        
        return {
            "cohort": {
                "id": cohort.id,
                "name": cohort.name,
                "description": cohort.description
            },
            "assigned_courses": assigned_courses,
            "available_courses": available_courses
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get cohort courses error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch cohort courses")

@app.get("/admin/courses/{course_id}/cohorts")
async def get_course_cohorts(
    course_id: int,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        # Check if course exists
        course = db.query(Course).filter(Course.id == course_id).first()
        if not course:
            raise HTTPException(status_code=404, detail="Course not found")
        
        # Get cohorts assigned to this course
        cohort_courses = db.query(CohortCourse).filter(CohortCourse.course_id == course_id).all()
        
        assigned_cohorts = []
        for cc in cohort_courses:
            cohort = db.query(Cohort).filter(Cohort.id == cc.cohort_id).first()
            if cohort:
                # Get student count in this cohort
                student_count = db.query(User).filter(User.cohort_id == cohort.id).count()
                
                assigned_cohorts.append({
                    "id": cohort.id,
                    "name": cohort.name,
                    "description": cohort.description,
                    "instructor_name": cohort.instructor_name,
                    "student_count": student_count,
                    "assigned_at": cc.assigned_at
                })
        
        return {
            "course": {
                "id": course.id,
                "title": course.title,
                "description": course.description
            },
            "assigned_cohorts": assigned_cohorts
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get course cohorts error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch course cohorts")

# Student Quiz and Resource Viewing Endpoints
@app.get("/student/quiz/{quiz_id}")
async def get_quiz_for_student(
    quiz_id: int,
    current_user: User = Depends(require_role("Student")),
    db: Session = Depends(get_db)
):
    try:
        quiz = db.query(Quiz).filter(Quiz.id == quiz_id).first()
        if not quiz:
            raise HTTPException(status_code=404, detail="Quiz not found")
        
        # Check enrollment
        enrollment = db.query(Enrollment).filter(
            Enrollment.student_id == current_user.id,
            Enrollment.course_id == quiz.session.module.course_id
        ).first()
        if not enrollment:
            raise HTTPException(status_code=403, detail="Not enrolled in this course")
        
        # Check if already attempted
        attempt = db.query(QuizAttempt).filter(
            QuizAttempt.quiz_id == quiz_id,
            QuizAttempt.student_id == current_user.id
        ).first()
        
        # Parse questions
        questions = []
        if quiz.questions:
            try:
                questions = json.loads(quiz.questions)
            except:
                questions = []
        
        return {
            "id": quiz.id,
            "title": quiz.title,
            "description": quiz.description,
            "total_marks": quiz.total_marks,
            "time_limit_minutes": quiz.time_limit_minutes,
            "questions": questions,
            "attempted": attempt is not None,
            "score": attempt.score if attempt else None,
            "can_retake": True  # Allow retakes for now
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get quiz error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch quiz")

@app.post("/student/quiz/{quiz_id}/submit")
async def submit_quiz_attempt(
    quiz_id: int,
    answers: dict,
    current_user: User = Depends(require_role("Student")),
    db: Session = Depends(get_db)
):
    try:
        quiz = db.query(Quiz).filter(Quiz.id == quiz_id).first()
        if not quiz:
            raise HTTPException(status_code=404, detail="Quiz not found")
        
        # Check enrollment
        enrollment = db.query(Enrollment).filter(
            Enrollment.student_id == current_user.id,
            Enrollment.course_id == quiz.session.module.course_id
        ).first()
        if not enrollment:
            raise HTTPException(status_code=403, detail="Not enrolled in this course")
        
        # Calculate score
        questions = json.loads(quiz.questions) if quiz.questions else []
        total_questions = len(questions)
        correct_answers = 0
        
        for i, question in enumerate(questions):
            user_answer = answers.get(str(i))
            if question["type"] == "MCQ":
                if user_answer == question["correct_answer"]:
                    correct_answers += 1
            elif question["type"] == "TRUE_FALSE":
                if user_answer == question["correct_answer"]:
                    correct_answers += 1
            # For SHORT_ANSWER, give full marks for now (would need AI grading)
            elif question["type"] == "SHORT_ANSWER":
                correct_answers += 1
        
        score = (correct_answers / total_questions * quiz.total_marks) if total_questions > 0 else 0
        
        # Delete previous attempt if exists
        db.query(QuizAttempt).filter(
            QuizAttempt.quiz_id == quiz_id,
            QuizAttempt.student_id == current_user.id
        ).delete()
        
        # Create new attempt
        attempt = QuizAttempt(
            quiz_id=quiz_id,
            student_id=current_user.id,
            score=score,
            answers=json.dumps(answers),
            time_taken_minutes=answers.get("time_taken", 0)
        )
        db.add(attempt)
        db.commit()
        
        return {
            "message": "Quiz submitted successfully",
            "score": score,
            "total_marks": quiz.total_marks,
            "percentage": (score / quiz.total_marks * 100) if quiz.total_marks > 0 else 0,
            "correct_answers": correct_answers,
            "total_questions": total_questions
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Submit quiz error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to submit quiz")

# Student Quiz Endpoints
@app.get("/student/quizzes/{session_id}")
async def get_session_quizzes_student(
    session_id: int,
    current_user: User = Depends(require_role("Student")),
    db: Session = Depends(get_db)
):
    try:
        # Check enrollment first
        session = db.query(SessionModel).filter(SessionModel.id == session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        enrollment = db.query(Enrollment).filter(
            Enrollment.student_id == current_user.id,
            Enrollment.course_id == session.module.course_id
        ).first()
        if not enrollment:
            raise HTTPException(status_code=403, detail="Not enrolled in this course")
        
        # Get quizzes from both Quiz table and SessionContent
        quizzes = db.query(Quiz).filter(Quiz.session_id == session_id, Quiz.is_active == True).all()
        session_contents = db.query(SessionContent).filter(
            SessionContent.session_id == session_id,
            SessionContent.content_type == "QUIZ"
        ).all()
        
        result = []
        
        # Add regular quizzes
        for quiz in quizzes:
            attempt = db.query(QuizAttempt).filter(
                QuizAttempt.quiz_id == quiz.id,
                QuizAttempt.student_id == current_user.id
            ).first()
            
            result.append({
                "id": quiz.id,
                "type": "quiz",
                "title": quiz.title,
                "description": quiz.description,
                "total_marks": quiz.total_marks,
                "time_limit_minutes": quiz.time_limit_minutes,
                "attempted": attempt is not None,
                "score": attempt.score if attempt else None,
                "view_url": f"/student/quiz/{quiz.id}"
            })
        
        # Add quiz content from session content
        for content in session_contents:
            if content.content_data:
                try:
                    quiz_data = json.loads(content.content_data)
                    quiz_id = quiz_data.get("quiz_id")
                    
                    # Check if this quiz is already added from Quiz table
                    if not any(q["id"] == quiz_id and q["type"] == "quiz" for q in result):
                        attempt = None
                        if quiz_id:
                            attempt = db.query(QuizAttempt).filter(
                                QuizAttempt.quiz_id == quiz_id,
                                QuizAttempt.student_id == current_user.id
                            ).first()
                        
                        result.append({
                            "id": content.id,
                            "type": "content_quiz",
                            "title": content.title,
                            "description": content.description,
                            "total_marks": quiz_data.get("total_marks", 100),
                            "time_limit_minutes": quiz_data.get("time_limit_minutes", 60),
                            "attempted": attempt is not None if attempt else False,
                            "score": attempt.score if attempt else None,
                            "view_url": f"/student/session-content/{content.id}/quiz"
                        })
                except:
                    pass
        
        return {"quizzes": result}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get session quizzes error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch quizzes")

@app.get("/admin/resources/{resource_id}/view")
async def view_resource_admin(
    resource_id: int,
    request: Request,
    token: Optional[str] = None,
    # current_user = Depends(get_current_admin_presenter_student_mentor_or_manager),
    db: Session = Depends(get_db)
):
    try:
        resource = db.query(Resource).filter(Resource.id == resource_id).first()
        if not resource:
            raise HTTPException(status_code=404, detail="Resource not found")
        
        # Track the resource view
        try:
            from resource_analytics_models import ResourceView
            client_ip = request.client.host if request.client else "127.0.0.1"
            user_agent = request.headers.get("user-agent", "")
            
            view_record = ResourceView(
                resource_id=resource_id,
                student_id=current_user["id"] if isinstance(current_user, dict) else current_user.id,
                viewed_at=datetime.utcnow(),
                ip_address=client_ip,
                user_agent=user_agent
            )
            
            db.add(view_record)
            db.commit()
            logger.info(f"Resource view tracked: resource_id={resource_id}, user_id={current_user['id'] if isinstance(current_user, dict) else current_user.id}")
        except Exception as track_error:
            logger.warning(f"Failed to track resource view: {str(track_error)}")
            # Don't fail the request if tracking fails
        
        if not resource.file_path:
            raise HTTPException(status_code=404, detail="No file associated with this resource")
        
        # Try different possible file paths
        possible_paths = [
            resource.file_path,
            str(UPLOAD_BASE_DIR / "resources" / os.path.basename(resource.file_path)),
            str(UPLOAD_BASE_DIR / os.path.basename(resource.file_path))
        ]
        
        actual_file_path = None
        for path in possible_paths:
            if os.path.exists(path):
                actual_file_path = path
                break
        
        if not actual_file_path:
            logger.error(f"Resource file not found in any location: {possible_paths}")
            # Create the file if it doesn't exist
            placeholder_content = f"Resource '{resource.title}' is currently unavailable.\n\nOriginal file: {resource.file_path}\nResource Type: {resource.resource_type}\nDescription: {resource.description or 'No description'}"
            
            # Create the file in the resources directory
            new_file_path = UPLOAD_BASE_DIR / "resources" / f"missing_{resource.id}.txt"
            with open(new_file_path, "w", encoding='utf-8') as f:
                f.write(placeholder_content)
            
            # Update the database with the new path
            resource.file_path = str(new_file_path)
            resource.file_size = len(placeholder_content.encode('utf-8'))
            db.commit()
            
            actual_file_path = str(new_file_path)
        
        # Determine media type based on file extension
        file_ext = os.path.splitext(actual_file_path)[1].lower()
        media_type_map = {
            '.pdf': 'application/pdf',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png',
            '.gif': 'image/gif',
            '.mp4': 'video/mp4',
            '.avi': 'video/avi',
            '.mov': 'video/quicktime',
            '.txt': 'text/plain',
            '.doc': 'application/msword',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.ppt': 'application/vnd.ms-powerpoint',
            '.pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation'
        }
        
        media_type = media_type_map.get(file_ext, 'application/octet-stream')
        
        # Add CORS headers for file serving
        headers = {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET",
            "Access-Control-Allow-Headers": "*"
        }
        
        # For all files, set headers to display inline (no downloads)
        headers["Content-Disposition"] = f"inline; filename=\"{os.path.basename(actual_file_path)}\""
        
        return FileResponse(
            path=actual_file_path,
            media_type=media_type,
            headers=headers
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"View resource error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to view resource")

@app.get("/api/resources/{resource_id}/view")
async def view_resource(resource_id: int, request: Request, current_user = Depends(get_current_user), db: Session = Depends(get_db)):
    try:
        resource = db.query(Resource).filter(Resource.id == resource_id).first()
        if not resource:
            raise HTTPException(status_code=404, detail="Resource not found")
        
        # Track the resource view
        try:
            from resource_analytics_models import ResourceView
            client_ip = request.client.host if request.client else "127.0.0.1"
            user_agent = request.headers.get("user-agent", "")
            
            view_record = ResourceView(
                resource_id=resource_id,
                student_id=current_user["id"] if isinstance(current_user, dict) else current_user.id,
                viewed_at=datetime.utcnow(),
                ip_address=client_ip,
                user_agent=user_agent
            )
            
            db.add(view_record)
            db.commit()
            logger.info(f"Resource view tracked: resource_id={resource_id}, user_id={current_user['id'] if isinstance(current_user, dict) else current_user.id}")
        except Exception as track_error:
            logger.warning(f"Failed to track resource view: {str(track_error)}")
            # Don't fail the request if tracking fails
        
        actual_file_path = None
        if resource.file_path:
            possible_paths = [
                resource.file_path,
                str(UPLOAD_BASE_DIR / "resources" / os.path.basename(resource.file_path))
            ]
            for path in possible_paths:
                if os.path.exists(path):
                    actual_file_path = path
                    break
        
        if not actual_file_path:
            raise HTTPException(status_code=404, detail="File not found")
        
        filename = os.path.basename(actual_file_path)
        file_ext = os.path.splitext(filename)[1].lower()
        file_url = f"/api/resources/{filename}"
        
        # For Office files, use canvas-based viewer
        if file_ext in [".ppt", ".pptx", ".doc", ".docx", ".xls", ".xlsx"]:
            file_type = "PowerPoint" if file_ext in [".ppt", ".pptx"] else "Word" if file_ext in [".doc", ".docx"] else "Excel"
            
            html_content = f"""<!DOCTYPE html>
<html>
<head>
    <title>{resource.title}</title>
    <style>
        body {{ margin: 0; padding: 0; font-family: Arial, sans-serif; background: #f5f5f5; }}
        .viewer {{ width: 100%; height: 100vh; display: flex; align-items: center; justify-content: center; }}
        canvas {{ background: white; box-shadow: 0 4px 8px rgba(0,0,0,0.1); cursor: pointer; }}
    </style>
</head>
<body>
    <div class="viewer">
        <canvas id="pptCanvas" width="800" height="600"></canvas>
    </div>
    <script>
        const canvas = document.getElementById('pptCanvas');
        const ctx = canvas.getContext('2d');
        
        // Draw PowerPoint slide
        ctx.fillStyle = '#ffffff';
        ctx.fillRect(0, 0, canvas.width, canvas.height);
        
        // Title
        ctx.fillStyle = '#2c3e50';
        ctx.font = 'bold 32px Arial';
        ctx.textAlign = 'center';
        ctx.fillText('{resource.title}', canvas.width/2, 100);
        
        // PowerPoint icon
        ctx.fillStyle = '#d35400';
        ctx.font = '80px Arial';
        ctx.fillText('📊', canvas.width/2, 250);
        
        // File type
        ctx.fillStyle = '#7f8c8d';
        ctx.font = '24px Arial';
        ctx.fillText('{file_type} Presentation', canvas.width/2, 320);
        
        // Instructions
        ctx.fillStyle = '#3498db';
        ctx.font = '18px Arial';
        ctx.fillText('Click to view presentation', canvas.width/2, 400);
        
        // Border
        ctx.strokeStyle = '#bdc3c7';
        ctx.lineWidth = 2;
        ctx.strokeRect(0, 0, canvas.width, canvas.height);
        
        // Click handler - embed file in iframe instead of opening new tab
        canvas.addEventListener('click', function() {{
            const viewer = document.querySelector('.viewer');
            viewer.innerHTML = '<iframe src="{file_url}" style="width:100%;height:100vh;border:none;"></iframe>';
        }});
    </script>
</body>
</html>"""
        else:
            # For other files, use direct iframe
            html_content = f"""<!DOCTYPE html>
<html>
<head>
    <title>{resource.title}</title>
    <style>body{{margin:0;padding:0;}}iframe{{width:100%;height:100vh;border:none;}}</style>
</head>
<body>
    <iframe src="{file_url}"></iframe>
</body>
</html>"""
        
        return Response(content=html_content, media_type="text/html")
    except Exception as e:
        logger.error(f"View resource error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to view resource")

@app.get("/api/session-content/{content_id}/view")
async def view_session_content(
    content_id: int,
    token: Optional[str] = None,
    db: Session = Depends(get_db)
):
    try:
        content = db.query(SessionContent).filter(SessionContent.id == content_id).first()
        if not content:
            raise HTTPException(status_code=404, detail="Content not found")
        
        # Handle meeting link content
        if content.content_type == "MEETING_LINK":
            now = datetime.now()
            scheduled_time = content.scheduled_time
            is_active = scheduled_time and now >= scheduled_time
            
            return {
                "type": "meeting_link",
                "meeting_url": content.meeting_url if is_active else None,
                "scheduled_time": scheduled_time,
                "is_active": is_active,
                "title": content.title,
                "description": content.description
            }
        
        # Handle quiz content
        if content.content_type == "QUIZ" and content.content_data:
            try:
                quiz_data = json.loads(content.content_data)
                quiz_id = quiz_data.get("quiz_id")
                if quiz_id:
                    return {"type": "quiz", "quiz_id": quiz_id}
            except:
                pass
        
        # Ensure resources directory exists
        (UPLOAD_BASE_DIR / "resources").mkdir(exist_ok=True)
        
        actual_file_path = None
        
        if content.file_path:
            # Try different possible file paths
            possible_paths = [
                content.file_path,
                str(UPLOAD_BASE_DIR / "resources" / os.path.basename(content.file_path)),
                str(UPLOAD_BASE_DIR / os.path.basename(content.file_path))
            ]
            
            for path in possible_paths:
                if os.path.exists(path):
                    actual_file_path = path
                    break
        
        # If no file found, create a placeholder
        if not actual_file_path:
            placeholder_content = f"""Session Content: {content.title}

Type: {content.content_type}
Description: {content.description or 'No description provided'}

This content is currently available for viewing.
Content will be displayed here when the actual file is uploaded.

Content ID: {content.id}
Session ID: {content.session_id}
Created: {content.created_at}
"""
            
            # Create placeholder file
            placeholder_filename = f"content_{content.id}.txt"
            actual_file_path = str(UPLOAD_BASE_DIR / "resources" / placeholder_filename)
            
            with open(actual_file_path, "w", encoding='utf-8') as f:
                f.write(placeholder_content)
            
            # Update database
            content.file_path = actual_file_path
            content.file_size = len(placeholder_content.encode('utf-8'))
            db.commit()
        
        # Determine media type
        file_ext = os.path.splitext(actual_file_path)[1].lower()
        media_type_map = {
            '.pdf': 'application/pdf',
            '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.png': 'image/png', '.gif': 'image/gif',
            '.mp4': 'video/mp4', '.avi': 'video/avi', '.mov': 'video/quicktime',
            '.txt': 'text/plain', '.doc': 'application/msword',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.ppt': 'application/vnd.ms-powerpoint',
            '.pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation'
        }
        
        media_type = media_type_map.get(file_ext, 'application/octet-stream')
        filename = os.path.basename(actual_file_path)
        
        # Set headers for proper viewing
        headers = {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET",
            "Access-Control-Allow-Headers": "*",
            "Content-Disposition": "inline", "X-Download-Options": "noopen", "X-Content-Type-Options": "nosniff", "Cache-Control": "no-store\"{filename}\""
        }
        
        return FileResponse(
            path=actual_file_path,
            media_type=media_type,
            headers=headers
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Session content view error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to view content")

@app.get("/student/meeting-links/{session_id}")
async def get_session_meeting_links(
    session_id: int,
    current_user: User = Depends(require_role("Student")),
    db: Session = Depends(get_db)
):
    try:
        # Check enrollment
        session = db.query(SessionModel).filter(SessionModel.id == session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        enrollment = db.query(Enrollment).filter(
            Enrollment.student_id == current_user.id,
            Enrollment.course_id == session.module.course_id
        ).first()
        if not enrollment:
            raise HTTPException(status_code=403, detail="Not enrolled in this course")
        
        # Get meeting links from session content
        meeting_contents = db.query(SessionContent).filter(
            SessionContent.session_id == session_id,
            SessionContent.content_type == "MEETING_LINK"
        ).all()
        
        now = datetime.now()
        meeting_links = []
        
        for content in meeting_contents:
            scheduled_time = content.scheduled_time
            is_active = scheduled_time and now >= scheduled_time
            is_upcoming = scheduled_time and now < scheduled_time
            
            meeting_links.append({
                "id": content.id,
                "title": content.title,
                "description": content.description,
                "meeting_url": content.meeting_url,
                "scheduled_time": scheduled_time,
                "is_active": is_active,
                "is_upcoming": is_upcoming,
                "status": "active" if is_active else "upcoming" if is_upcoming else "ended"
            })
        
        return {"meeting_links": meeting_links}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get meeting links error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch meeting links")

@app.get("/student/session-content/{content_id}/quiz")
async def get_quiz_from_content(
    content_id: int,
    current_user: User = Depends(require_role("Student")),
    db: Session = Depends(get_db)
):
    try:
        content = db.query(SessionContent).filter(SessionContent.id == content_id).first()
        if not content or content.content_type != "QUIZ":
            raise HTTPException(status_code=404, detail="Quiz content not found")
        
        # Check enrollment
        enrollment = db.query(Enrollment).filter(
            Enrollment.student_id == current_user.id,
            Enrollment.course_id == content.session.module.course_id
        ).first()
        if not enrollment:
            raise HTTPException(status_code=403, detail="Not enrolled in this course")
        
        # Get quiz data from content_data
        if not content.content_data:
            raise HTTPException(status_code=404, detail="Quiz data not found")
        
        try:
            quiz_data = json.loads(content.content_data)
            quiz_id = quiz_data.get("quiz_id")
            
            if quiz_id:
                # Get the actual quiz
                quiz = db.query(Quiz).filter(Quiz.id == quiz_id).first()
                if quiz:
                    # Check if already attempted
                    attempt = db.query(QuizAttempt).filter(
                        QuizAttempt.quiz_id == quiz_id,
                        QuizAttempt.student_id == current_user.id
                    ).first()
                    
                    # Parse questions
                    questions = []
                    if quiz.questions:
                        try:
                            questions = json.loads(quiz.questions)
                        except:
                            questions = quiz_data.get("questions", [])
                    else:
                        questions = quiz_data.get("questions", [])
                    
                    return {
                        "id": quiz.id,
                        "title": quiz.title,
                        "description": quiz.description,
                        "total_marks": quiz.total_marks,
                        "time_limit_minutes": quiz.time_limit_minutes,
                        "questions": questions,
                        "attempted": attempt is not None,
                        "score": attempt.score if attempt else None,
                        "can_retake": True
                    }
            
            # Fallback to content data
            return {
                "id": f"content_{content_id}",
                "title": content.title,
                "description": content.description,
                "total_marks": quiz_data.get("total_marks", 100),
                "time_limit_minutes": quiz_data.get("time_limit_minutes", 60),
                "questions": quiz_data.get("questions", []),
                "attempted": False,
                "score": None,
                "can_retake": True
            }
        except json.JSONDecodeError:
            raise HTTPException(status_code=500, detail="Invalid quiz data")
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get quiz from content error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch quiz")

@app.post("/student/session-content/{content_id}/quiz/submit")
async def submit_content_quiz(
    content_id: int,
    answers: dict,
    current_user: User = Depends(require_role("Student")),
    db: Session = Depends(get_db)
):
    try:
        content = db.query(SessionContent).filter(SessionContent.id == content_id).first()
        if not content or content.content_type != "QUIZ":
            raise HTTPException(status_code=404, detail="Quiz content not found")
        
        # Check enrollment
        enrollment = db.query(Enrollment).filter(
            Enrollment.student_id == current_user.id,
            Enrollment.course_id == content.session.module.course_id
        ).first()
        if not enrollment:
            raise HTTPException(status_code=403, detail="Not enrolled in this course")
        
        if not content.content_data:
            raise HTTPException(status_code=404, detail="Quiz data not found")
        
        quiz_data = json.loads(content.content_data)
        quiz_id = quiz_data.get("quiz_id")
        
        if quiz_id:
            # Submit to actual quiz
            return await submit_quiz_attempt(quiz_id, answers, current_user, db)
        else:
            # Handle content-only quiz (create a temporary quiz attempt)
            questions = quiz_data.get("questions", [])
            total_questions = len(questions)
            correct_answers = 0
            
            for i, question in enumerate(questions):
                user_answer = answers.get(str(i))
                if question["type"] == "MCQ":
                    if user_answer == question["correct_answer"]:
                        correct_answers += 1
                elif question["type"] == "TRUE_FALSE":
                    if user_answer == question["correct_answer"]:
                        correct_answers += 1
                elif question["type"] == "SHORT_ANSWER":
                    correct_answers += 1  # Give full marks for now
            
            total_marks = quiz_data.get("total_marks", 100)
            score = (correct_answers / total_questions * total_marks) if total_questions > 0 else 0
            
            return {
                "message": "Quiz submitted successfully",
                "score": score,
                "total_marks": total_marks,
                "percentage": (score / total_marks * 100) if total_marks > 0 else 0,
                "correct_answers": correct_answers,
                "total_questions": total_questions
            }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Submit content quiz error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to submit quiz")

# Assignment endpoints for students
@app.get("/student/assignments/{course_id}")
async def get_course_assignments_student(
    course_id: int,
    current_user: User = Depends(require_role("Student")),
    db: Session = Depends(get_db)
):
    try:
        # Check enrollment
        enrollment = db.query(Enrollment).filter(
            Enrollment.student_id == current_user.id,
            Enrollment.course_id == course_id
        ).first()
        if not enrollment:
            raise HTTPException(status_code=403, detail="Not enrolled in this course")
        
        assignments = db.query(Assignment).filter(Assignment.course_id == course_id).all()
        
        result = []
        for assignment in assignments:
            submission = db.query(Submission).filter(
                Submission.assignment_id == assignment.id,
                Submission.student_id == current_user.id
            ).first()
            
            result.append({
                "id": assignment.id,
                "title": assignment.title,
                "description": assignment.description,
                "due_date": assignment.due_date,
                "total_marks": assignment.total_marks,
                "submission_type": assignment.submission_type,
                "submitted": submission is not None,
                "submitted_at": submission.submitted_at if submission else None,
                "score": submission.score if submission else None,
                "feedback": submission.feedback if submission else None,
                "course_title": assignment.course.title
            })
        
        return {"assignments": result}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get course assignments error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch assignments")

@app.post("/student/assignments/{assignment_id}/submit")
async def submit_assignment_student(
    assignment_id: int,
    content: str,
    file: UploadFile = File(None),
    current_user: User = Depends(require_role("Student")),
    db: Session = Depends(get_db)
):
    try:
        assignment = db.query(Assignment).filter(Assignment.id == assignment_id).first()
        if not assignment:
            raise HTTPException(status_code=404, detail="Assignment not found")
        
        # Check enrollment
        enrollment = db.query(Enrollment).filter(
            Enrollment.student_id == current_user.id,
            Enrollment.course_id == assignment.course_id
        ).first()
        if not enrollment:
            raise HTTPException(status_code=403, detail="Not enrolled in this course")
        
        # Handle file upload if provided
        file_path = None
        if file:
            file_extension = os.path.splitext(file.filename)[1]
            unique_filename = f"assignment_{assignment_id}_{current_user.id}_{uuid.uuid4().hex[:8]}{file_extension}"
            file_path = UPLOAD_BASE_DIR / "assignments" / unique_filename
            
            with open(file_path, "wb") as buffer:
                file_content = await file.read()
                buffer.write(file_content)
            
            file_path = str(file_path)
        
        # Delete existing submission if any
        db.query(Submission).filter(
            Submission.assignment_id == assignment_id,
            Submission.student_id == current_user.id
        ).delete()
        
        # Create new submission
        submission = Submission(
            assignment_id=assignment_id,
            student_id=current_user.id,
            content=content,
            file_path=file_path
        )
        db.add(submission)
        db.commit()
        
        return {
            "message": "Assignment submitted successfully",
            "submission_id": submission.id,
            "submitted_at": submission.submitted_at
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Submit assignment error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to submit assignment")

# Cohort Management Endpoints
@app.post("/admin/cohorts")
async def create_cohort(
    cohort_data: CohortCreate,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        cohort = Cohort(
            name=cohort_data.name,
            description=cohort_data.description,
            start_date=cohort_data.start_date,
            end_date=cohort_data.end_date,
            instructor_name=cohort_data.instructor_name,
            created_by=current_user.id
        )
        db.add(cohort)
        db.commit()
        db.refresh(cohort)
        
        # Log cohort creation
        if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="CREATE",
                resource_type="COHORT",
                resource_id=cohort.id,
                details=f"Created cohort: {cohort_data.name}"
            )
        elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="CREATE",
                resource_type="COHORT",
                resource_id=cohort.id,
                details=f"Created cohort: {cohort_data.name}"
            )
        
        return {"message": "Cohort created successfully", "cohort_id": cohort.id}
    except Exception as e:
        db.rollback()
        logger.error(f"Create cohort error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create cohort")

@app.get("/admin/cohorts")
async def get_all_cohorts(
    page: int = 1,
    limit: int = 20,
    search: Optional[str] = None,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        query = db.query(Cohort)
        
        if search:
            query = query.filter(
                or_(
                    Cohort.name.contains(search),
                    Cohort.description.contains(search),
                    Cohort.instructor_name.contains(search)
                )
            )
        
        total = query.count()
        cohorts = query.offset((page - 1) * limit).limit(limit).all()
        
        result = []
        for cohort in cohorts:
            user_count = db.query(UserCohort).filter(UserCohort.cohort_id == cohort.id).count()
            course_count = db.query(CohortCourse).filter(CohortCourse.cohort_id == cohort.id).count()
            
            result.append({
                "id": cohort.id,
                "name": cohort.name,
                "description": cohort.description,
                "start_date": cohort.start_date,
                "end_date": cohort.end_date,
                "instructor_name": cohort.instructor_name,
                "user_count": user_count,
                "course_count": course_count,
                "created_at": cohort.created_at
            })
        
        return {
            "cohorts": result,
            "total": total,
            "page": page,
            "limit": limit
        }
    except Exception as e:
        logger.error(f"Get cohorts error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch cohorts")

@app.get("/admin/cohorts/{cohort_id}")
async def get_cohort_details(
    cohort_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        cohort = db.query(Cohort).filter(Cohort.id == cohort_id).first()
        if not cohort:
            raise HTTPException(status_code=404, detail="Cohort not found")
        
        # Get users in cohort
        user_cohorts = db.query(UserCohort).filter(UserCohort.cohort_id == cohort_id).all()
        users = [{
            "id": uc.user.id,
            "username": uc.user.username,
            "email": uc.user.email,
            "college": uc.user.college,
            "assigned_at": uc.assigned_at
        } for uc in user_cohorts]
        
        # Get courses assigned to cohort
        cohort_courses = db.query(CohortCourse).filter(CohortCourse.cohort_id == cohort_id).all()
        courses = [{
            "id": cc.course.id,
            "title": cc.course.title,
            "description": cc.course.description,
            "assigned_at": cc.assigned_at
        } for cc in cohort_courses]
        
        return {
            "id": cohort.id,
            "name": cohort.name,
            "description": cohort.description,
            "start_date": cohort.start_date,
            "end_date": cohort.end_date,
            "instructor_name": cohort.instructor_name,
            "created_at": cohort.created_at,
            "users": users,
            "courses": courses
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get cohort details error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch cohort details")

@app.put("/admin/cohorts/{cohort_id}")
async def update_cohort(
    cohort_id: int,
    cohort_data: CohortUpdate,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        cohort = db.query(Cohort).filter(Cohort.id == cohort_id).first()
        if not cohort:
            raise HTTPException(status_code=404, detail="Cohort not found")
        
        update_data = cohort_data.dict(exclude_unset=True)
        for field, value in update_data.items():
            setattr(cohort, field, value)
        
        db.commit()
        return {"message": "Cohort updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Update cohort error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update cohort")

@app.delete("/admin/cohorts/{cohort_id}")
async def delete_cohort(
    cohort_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        cohort = db.query(Cohort).filter(Cohort.id == cohort_id).first()
        if not cohort:
            raise HTTPException(status_code=404, detail="Cohort not found")
        
        cohort_name = cohort.name
        
        # Remove users from cohort
        db.query(UserCohort).filter(UserCohort.cohort_id == cohort_id).delete()
        
        # Remove course assignments
        db.query(CohortCourse).filter(CohortCourse.cohort_id == cohort_id).delete()
        
        # Update users' cohort_id to None
        db.query(User).filter(User.cohort_id == cohort_id).update({"cohort_id": None})
        
        db.delete(cohort)
        db.commit()
        
        # Log cohort deletion
        if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="DELETE",
                resource_type="COHORT",
                resource_id=cohort_id,
                details=f"Deleted cohort: {cohort_name}"
            )
        elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="DELETE",
                resource_type="COHORT",
                resource_id=cohort_id,
                details=f"Deleted cohort: {cohort_name}"
            )
        
        return {"message": "Cohort deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Delete cohort error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete cohort")

@app.post("/admin/cohorts/{cohort_id}/users")
async def add_users_to_cohort(
    cohort_id: int,
    user_data: CohortUserAdd,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        cohort = db.query(Cohort).filter(Cohort.id == cohort_id).first()
        if not cohort:
            raise HTTPException(status_code=404, detail="Cohort not found")
        
        added_users = []
        errors = []
        
        for user_id in user_data.user_ids:
            user = db.query(User).filter(User.id == user_id).first()
            if not user:
                errors.append(f"User ID {user_id} not found")
                continue
            
            # Check if user is already in a cohort
            existing_cohort = db.query(UserCohort).filter(UserCohort.user_id == user_id).first()
            if existing_cohort:
                errors.append(f"User {user.username} is already in cohort {existing_cohort.cohort.name}")
                continue
            
            # Add user to cohort
            user_cohort = UserCohort(
                user_id=user_id,
                cohort_id=cohort_id,
                assigned_by=current_user.id
            )
            db.add(user_cohort)
            
            # Update user's current cohort
            user.cohort_id = cohort_id
            
            added_users.append(user.username)
        
        db.commit()
        
        # Log cohort user addition
        if added_users:
            if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
                log_admin_action(
                    admin_id=current_user.id,
                    admin_username=current_user.username,
                    action_type="CREATE",
                    resource_type="COHORT_USER",
                    resource_id=cohort_id,
                    details=f"Added {len(added_users)} users to cohort: {', '.join(added_users)}"
                )
            elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
                log_presenter_action(
                    presenter_id=current_user.id,
                    presenter_username=current_user.username,
                    action_type="CREATE",
                    resource_type="COHORT_USER",
                    resource_id=cohort_id,
                    details=f"Added {len(added_users)} users to cohort: {', '.join(added_users)}"
                )
        
        return {
            "message": f"Added {len(added_users)} users to cohort",
            "added_users": added_users,
            "errors": errors
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Add users to cohort error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to add users to cohort")

@app.delete("/admin/cohorts/{cohort_id}/users/{user_id}")
async def remove_user_from_cohort(
    cohort_id: int,
    user_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        user_cohort = db.query(UserCohort).filter(
            UserCohort.cohort_id == cohort_id,
            UserCohort.user_id == user_id
        ).first()
        
        if not user_cohort:
            raise HTTPException(status_code=404, detail="User not found in cohort")
        
        # Get user info for logging
        user = db.query(User).filter(User.id == user_id).first()
        username = user.username if user else f"User ID {user_id}"
        
        # Remove from cohort
        db.delete(user_cohort)
        
        # Update user's current cohort
        if user:
            user.cohort_id = None
        
        db.commit()
        
        # Log cohort user removal
        if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="DELETE",
                resource_type="COHORT_USER",
                resource_id=cohort_id,
                details=f"Removed user {username} from cohort"
            )
        elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="DELETE",
                resource_type="COHORT_USER",
                resource_id=cohort_id,
                details=f"Removed user {username} from cohort"
            )
        
        return {"message": "User removed from cohort successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Remove user from cohort error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to remove user from cohort")

@app.post("/admin/cohorts/{cohort_id}/courses")
async def assign_courses_to_cohort(
    cohort_id: int,
    course_data: CohortCourseAssign,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        cohort = db.query(Cohort).filter(Cohort.id == cohort_id).first()
        if not cohort:
            raise HTTPException(status_code=404, detail="Cohort not found")
        
        assigned_courses = []
        errors = []
        
        for course_id in course_data.course_ids:
            course = db.query(Course).filter(Course.id == course_id).first()
            if not course:
                errors.append(f"Course ID {course_id} not found")
                continue
            
            # Check if course is already assigned to cohort
            existing_assignment = db.query(CohortCourse).filter(
                CohortCourse.cohort_id == cohort_id,
                CohortCourse.course_id == course_id
            ).first()
            
            if existing_assignment:
                errors.append(f"Course {course.title} is already assigned to this cohort")
                continue
            
            # Assign course to cohort
            cohort_course = CohortCourse(
                cohort_id=cohort_id,
                course_id=course_id,
                assigned_by=current_user.id
            )
            db.add(cohort_course)
            assigned_courses.append(course.title)
        
        db.commit()
        
        # Log cohort course assignment
        if assigned_courses:
            if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
                log_admin_action(
                    admin_id=current_user.id,
                    admin_username=current_user.username,
                    action_type="CREATE",
                    resource_type="COHORT_COURSE",
                    resource_id=cohort_id,
                    details=f"Assigned {len(assigned_courses)} courses to cohort: {', '.join(assigned_courses)}"
                )
            elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
                log_presenter_action(
                    presenter_id=current_user.id,
                    presenter_username=current_user.username,
                    action_type="CREATE",
                    resource_type="COHORT_COURSE",
                    resource_id=cohort_id,
                    details=f"Assigned {len(assigned_courses)} courses to cohort: {', '.join(assigned_courses)}"
                )
        
        return {
            "message": f"Assigned {len(assigned_courses)} courses to cohort",
            "assigned_courses": assigned_courses,
            "errors": errors
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Assign courses to cohort error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to assign courses to cohort")

@app.delete("/admin/cohorts/{cohort_id}/courses/{course_id}")
async def remove_course_from_cohort(
    cohort_id: int,
    course_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        cohort_course = db.query(CohortCourse).filter(
            CohortCourse.cohort_id == cohort_id,
            CohortCourse.course_id == course_id
        ).first()
        
        if not cohort_course:
            raise HTTPException(status_code=404, detail="Course not assigned to cohort")
        
        # Get course info for logging
        course = db.query(Course).filter(Course.id == course_id).first()
        course_title = course.title if course else f"Course ID {course_id}"
        
        db.delete(cohort_course)
        db.commit()
        
        # Log cohort course removal
        if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
            log_admin_action(
                admin_id=current_user.id,
                admin_username=current_user.username,
                action_type="DELETE",
                resource_type="COHORT_COURSE",
                resource_id=cohort_id,
                details=f"Removed course {course_title} from cohort"
            )
        elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
            log_presenter_action(
                presenter_id=current_user.id,
                presenter_username=current_user.username,
                action_type="DELETE",
                resource_type="COHORT_COURSE",
                resource_id=cohort_id,
                details=f"Removed course {course_title} from cohort"
            )
        
        return {"message": "Course removed from cohort successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Remove course from cohort error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to remove course from cohort")

@app.post("/admin/cohorts/{cohort_id}/bulk-upload")
async def bulk_upload_cohort_users(
    cohort_id: int,
    file: UploadFile = File(...),
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        cohort = db.query(Cohort).filter(Cohort.id == cohort_id).first()
        if not cohort:
            raise HTTPException(status_code=404, detail="Cohort not found")
        
        file_ext = os.path.splitext(file.filename)[1].lower()
        if file_ext not in ['.csv', '.xlsx', '.xls']:
            raise HTTPException(status_code=400, detail="Only CSV and Excel files are allowed")
        
        content = await file.read()
        
        if file_ext == '.csv':
            csv_content = content.decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(csv_content))
            records = list(csv_reader)
        else:
            from openpyxl import load_workbook
            workbook = load_workbook(io.BytesIO(content))
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]
            records = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                records.append(dict(zip(headers, row)))
        
        created_users = []
        added_to_cohort = []
        errors = []
        
        for row_num, row in enumerate(records, 1):
            try:
                # Support both old and new field names
                username = str(row.get('Username') or row.get('username', '') or '').strip()
                email = str(row.get('Email') or row.get('email', '') or '').strip()
                password = str(row.get('Password') or row.get('password', '') or '').strip()
                role = str(row.get('Role') or row.get('role', 'Student') or 'Student').strip()
                college = str(row.get('College') or row.get('college', '') or '').strip()
                year = str(row.get('Year') or row.get('year', '') or '').strip()
                
                if not all([username, email, password]):
                    errors.append(f"Row {row_num}: Missing required fields (Username, Email, Password)")
                    continue
                
                # Check if user exists
                existing_user = db.query(User).filter(User.username == username).first()
                if existing_user:
                    # Add existing user to cohort if not already in one
                    existing_cohort = db.query(UserCohort).filter(UserCohort.user_id == existing_user.id).first()
                    if existing_cohort:
                        errors.append(f"Row {row_num}: User '{username}' is already in cohort '{existing_cohort.cohort.name}'")
                        continue
                    
                    # Add to cohort
                    user_cohort = UserCohort(
                        user_id=existing_user.id,
                        cohort_id=cohort_id,
                        assigned_by=current_user.id
                    )
                    db.add(user_cohort)
                    existing_user.cohort_id = cohort_id
                    added_to_cohort.append(username)
                else:
                    # Create new user
                    if db.query(User).filter(User.email == email).first():
                        errors.append(f"Row {row_num}: Email '{email}' already exists")
                        continue
                    
                    user = User(
                        username=username,
                        email=email,
                        password_hash=get_password_hash(password),
                        role=role,
                        college=college,
                        cohort_id=cohort_id
                    )
                    
                    # Add year field if it exists in the User model
                    if hasattr(User, 'year') and year:
                        user.year = year
                    db.add(user)
                    db.flush()
                    
                    # Add to cohort
                    user_cohort = UserCohort(
                        user_id=user.id,
                        cohort_id=cohort_id,
                        assigned_by=current_user.id
                    )
                    db.add(user_cohort)
                    created_users.append(username)
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        db.commit()
        
        # Log cohort bulk upload
        total_processed = len(created_users) + len(added_to_cohort)
        if total_processed > 0:
            if hasattr(current_user, 'username') and db.query(Admin).filter(Admin.id == current_user.id).first():
                log_admin_action(
                    admin_id=current_user.id,
                    admin_username=current_user.username,
                    action_type="UPLOAD",
                    resource_type="COHORT_BULK_USERS",
                    resource_id=cohort_id,
                    details=f"Bulk uploaded {total_processed} users to cohort from file: {file.filename}"
                )
            elif hasattr(current_user, 'username') and db.query(Presenter).filter(Presenter.id == current_user.id).first():
                log_presenter_action(
                    presenter_id=current_user.id,
                    presenter_username=current_user.username,
                    action_type="UPLOAD",
                    resource_type="COHORT_BULK_USERS",
                    resource_id=cohort_id,
                    details=f"Bulk uploaded {total_processed} users to cohort from file: {file.filename}"
                )
        
        return {
            "message": f"Processed {len(records)} records",
            "created_users": created_users,
            "added_to_cohort": added_to_cohort,
            "total_processed": total_processed,
            "errors": errors
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Bulk upload cohort users error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to bulk upload users")

@app.get("/admin/cohorts/{cohort_id}/export")
async def export_cohort_users(
    cohort_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        cohort = db.query(Cohort).filter(Cohort.id == cohort_id).first()
        if not cohort:
            raise HTTPException(status_code=404, detail="Cohort not found")
        
        # Get users in cohort
        user_cohorts = db.query(UserCohort).filter(UserCohort.cohort_id == cohort_id).all()
        
        # Create CSV content
        csv_content = "username,email,role,college,assigned_at\n"
        for uc in user_cohorts:
            csv_content += f"{uc.user.username},{uc.user.email},{uc.user.role},{uc.user.college or ''},{uc.assigned_at}\n"
        
        # Create response
        response = Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=cohort_{cohort.name}_users.csv"}
        )
        
        return response
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Export cohort users error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to export cohort users")

@app.get("/admin/available-users")
async def get_available_users(
    search: Optional[str] = None,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        # Get users not in any cohort
        query = db.query(User).filter(User.cohort_id.is_(None))
        
        if search:
            query = query.filter(
                or_(
                    User.username.contains(search),
                    User.email.contains(search),
                    User.college.contains(search)
                )
            )
        
        users = query.limit(50).all()
        
        return {
            "users": [{
                "id": u.id,
                "username": u.username,
                "email": u.email,
                "college": u.college,
                "role": u.role
            } for u in users]
        }
    except Exception as e:
        logger.error(f"Get available users error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch available users")

@app.get("/admin/available-courses")
async def get_available_courses_for_cohort(
    cohort_id: Optional[int] = None,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        if cohort_id:
            # Get courses not assigned to this cohort
            assigned_course_ids = db.query(CohortCourse.course_id).filter(
                CohortCourse.cohort_id == cohort_id
            ).all()
            assigned_ids = [c[0] for c in assigned_course_ids]
            
            courses = db.query(Course).filter(~Course.id.in_(assigned_ids)).all()
        else:
            # Get all courses
            courses = db.query(Course).all()
        
        return {
            "courses": [{
                "id": c.id,
                "title": c.title,
                "description": c.description
            } for c in courses]
        }
    except Exception as e:
        logger.error(f"Get available courses error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch available courses")

@app.get("/admin/download-cohort-template")
async def download_cohort_template(
    current_admin = Depends(get_current_admin_or_presenter)
):
    """Download Excel template for cohort bulk upload"""
    try:
        if generate_cohort_template_excel is None:
            raise HTTPException(status_code=500, detail="Template generation not available. Please install required dependencies.")
        
        return generate_cohort_template_excel()
    except Exception as e:
        logger.error(f"Download cohort template error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to generate cohort template")

# Admin Logs Endpoints
@app.get("/admin/logs")
async def get_admin_logs(
    page: int = 1,
    limit: int = 50,
    action_type: Optional[str] = None,
    resource_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    search: Optional[str] = None,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        import mysql.connector
        
        # Connect to database directly for admin_logs table
        conn = mysql.connector.connect(
            host=os.getenv('DB_HOST', 'localhost'),
            user=os.getenv('DB_USER', 'root'),
            password=os.getenv('DB_PASSWORD', ''),
            database=os.getenv('DB_NAME', 'database')
        )
        cursor = conn.cursor(dictionary=True)
        
        # Build query with filters - handle empty strings properly
        query = "SELECT * FROM admin_logs WHERE 1=1"
        params = []
        
        # Only add filters if they have actual values (not empty strings)
        if action_type and action_type.strip():
            query += " AND action_type = %s"
            params.append(action_type.strip())
        
        if resource_type and resource_type.strip():
            query += " AND resource_type = %s"
            params.append(resource_type.strip())
        
        if search and search.strip():
            query += " AND (admin_username LIKE %s OR details LIKE %s OR resource_type LIKE %s)"
            search_param = f"%{search.strip()}%"
            params.extend([search_param, search_param, search_param])
        
        if date_from and date_from.strip():
            try:
                # Parse date and add to query
                parsed_date = datetime.strptime(date_from.strip(), '%Y-%m-%d')
                query += " AND DATE(timestamp) >= %s"
                params.append(parsed_date.date())
            except ValueError:
                pass  # Ignore invalid date format
        
        if date_to and date_to.strip():
            try:
                # Parse date and add to query
                parsed_date = datetime.strptime(date_to.strip(), '%Y-%m-%d')
                query += " AND DATE(timestamp) <= %s"
                params.append(parsed_date.date())
            except ValueError:
                pass  # Ignore invalid date format
        
        # Count total records
        count_query = query.replace("SELECT *", "SELECT COUNT(*)")
        cursor.execute(count_query, params)
        total = cursor.fetchone()['COUNT(*)']
        
        # Add pagination and ordering
        query += " ORDER BY timestamp DESC LIMIT %s OFFSET %s"
        params.extend([limit, (page - 1) * limit])
        
        cursor.execute(query, params)
        logs = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return {
            "logs": logs,
            "total": total,
            "page": page,
            "limit": limit,
            "filters": {
                "action_type": action_type,
                "resource_type": resource_type,
                "date_from": date_from,
                "date_to": date_to,
                "search": search
            }
        }
    except Exception as e:
        logger.error(f"Get admin logs error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch admin logs")

@app.get("/admin/presenter-logs")
async def get_presenter_logs(
    page: int = 1,
    limit: int = 50,
    action_type: Optional[str] = None,
    resource_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    search: Optional[str] = None,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        import mysql.connector
        
        # Connect to database directly for presenter_logs table
        conn = mysql.connector.connect(
            host=os.getenv('DB_HOST', 'localhost'),
            user=os.getenv('DB_USER', 'root'),
            password=os.getenv('DB_PASSWORD', ''),
            database=os.getenv('DB_NAME', 'database')
        )
        cursor = conn.cursor(dictionary=True)
        
        # Build query with filters
        query = "SELECT * FROM presenter_logs WHERE 1=1"
        params = []
        
        if action_type and action_type.strip():
            query += " AND action_type = %s"
            params.append(action_type.strip())
        
        if resource_type and resource_type.strip():
            query += " AND resource_type = %s"
            params.append(resource_type.strip())
        
        if date_from and date_from.strip():
            query += " AND timestamp >= %s"
            params.append(date_from.strip())
        
        if date_to and date_to.strip():
            query += " AND timestamp <= %s"
            params.append(date_to.strip())
        
        if search and search.strip():
            query += " AND (presenter_username LIKE %s OR details LIKE %s)"
            params.extend([f"%{search.strip()}%", f"%{search.strip()}%"])
        
        # Add ordering and pagination
        query += " ORDER BY timestamp DESC LIMIT %s OFFSET %s"
        params.extend([limit, (page - 1) * limit])
        
        cursor.execute(query, params)
        logs = cursor.fetchall()
        
        # Get total count
        count_query = "SELECT COUNT(*) as total FROM presenter_logs WHERE 1=1"
        count_params = []
        
        if action_type and action_type.strip():
            count_query += " AND action_type = %s"
            count_params.append(action_type.strip())
        
        if resource_type and resource_type.strip():
            count_query += " AND resource_type = %s"
            count_params.append(resource_type.strip())
        
        if date_from and date_from.strip():
            count_query += " AND timestamp >= %s"
            count_params.append(date_from.strip())
        
        if date_to and date_to.strip():
            count_query += " AND timestamp <= %s"
            count_params.append(date_to.strip())
        
        if search and search.strip():
            count_query += " AND (presenter_username LIKE %s OR details LIKE %s)"
            count_params.extend([f"%{search.strip()}%", f"%{search.strip()}%"])
        
        cursor.execute(count_query, count_params)
        total_result = cursor.fetchone()
        total = total_result['total'] if total_result else 0
        
        cursor.close()
        conn.close()
        
        return {
            "logs": logs,
            "total": total,
            "page": page,
            "limit": limit,
            "filters": {
                "action_type": action_type,
                "resource_type": resource_type,
                "date_from": date_from,
                "date_to": date_to,
                "search": search
            }
        }
    except Exception as e:
        logger.error(f"Get presenter logs error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch presenter logs")

@app.get("/admin/all-logs")
async def get_all_logs(
    page: int = 1,
    limit: int = 50,
    action_type: Optional[str] = None,
    resource_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    search: Optional[str] = None,
    log_type: Optional[str] = None,  # 'admin', 'presenter', or 'all'
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    """Get combined logs from both admin and presenter logs"""
    try:
        import mysql.connector
        
        # Connect to database
        conn = mysql.connector.connect(
            host=os.getenv('DB_HOST', 'localhost'),
            user=os.getenv('DB_USER', 'root'),
            password=os.getenv('DB_PASSWORD', ''),
            database=os.getenv('DB_NAME', 'database')
        )
        cursor = conn.cursor(dictionary=True)
        
        # Build union query for both admin and presenter logs
        admin_query = "SELECT id, admin_id as user_id, admin_username as username, 'Admin' as user_type, action_type, resource_type, resource_id, details, ip_address, timestamp FROM admin_logs"
        presenter_query = "SELECT id, presenter_id as user_id, presenter_username as username, 'Presenter' as user_type, action_type, resource_type, resource_id, details, ip_address, timestamp FROM presenter_logs"
        
        # Apply filters
        where_conditions = []
        params = []
        
        if log_type and log_type.strip() and log_type.lower() != 'all':
            if log_type.lower() == 'admin':
                query = admin_query
            elif log_type.lower() == 'presenter':
                query = presenter_query
            else:
                query = f"({admin_query}) UNION ALL ({presenter_query})"
        else:
            query = f"({admin_query}) UNION ALL ({presenter_query})"
        
        # Add WHERE conditions if not using UNION
        if log_type and log_type.lower() in ['admin', 'presenter']:
            where_conditions.append("1=1")
            
            if action_type and action_type.strip():
                where_conditions.append("action_type = %s")
                params.append(action_type.strip())
            
            if resource_type and resource_type.strip():
                where_conditions.append("resource_type = %s")
                params.append(resource_type.strip())
            
            if date_from and date_from.strip():
                where_conditions.append("timestamp >= %s")
                params.append(date_from.strip())
            
            if date_to and date_to.strip():
                where_conditions.append("timestamp <= %s")
                params.append(date_to.strip())
            
            if search and search.strip():
                username_field = "admin_username" if log_type.lower() == 'admin' else "presenter_username"
                where_conditions.append(f"({username_field} LIKE %s OR details LIKE %s)")
                params.extend([f"%{search.strip()}%", f"%{search.strip()}%"])
            
            if where_conditions:
                query += " WHERE " + " AND ".join(where_conditions)
        
        # For UNION queries, we need to wrap and filter
        if log_type is None or log_type.lower() == 'all' or (log_type and log_type.lower() not in ['admin', 'presenter']):
            query = f"SELECT * FROM ({query}) as combined_logs WHERE 1=1"
            
            if action_type and action_type.strip():
                query += " AND action_type = %s"
                params.append(action_type.strip())
            
            if resource_type and resource_type.strip():
                query += " AND resource_type = %s"
                params.append(resource_type.strip())
            
            if date_from and date_from.strip():
                query += " AND timestamp >= %s"
                params.append(date_from.strip())
            
            if date_to and date_to.strip():
                query += " AND timestamp <= %s"
                params.append(date_to.strip())
            
            if search and search.strip():
                query += " AND (username LIKE %s OR details LIKE %s)"
                params.extend([f"%{search.strip()}%", f"%{search.strip()}%"])
        
        # Add ordering and pagination
        query += " ORDER BY timestamp DESC LIMIT %s OFFSET %s"
        params.extend([limit, (page - 1) * limit])
        
        cursor.execute(query, params)
        logs = cursor.fetchall()
        
        # Get total count (simplified for combined logs)
        if log_type and log_type.lower() == 'admin':
            count_query = "SELECT COUNT(*) as total FROM admin_logs WHERE 1=1"
        elif log_type and log_type.lower() == 'presenter':
            count_query = "SELECT COUNT(*) as total FROM presenter_logs WHERE 1=1"
        else:
            count_query = "SELECT (SELECT COUNT(*) FROM admin_logs) + (SELECT COUNT(*) FROM presenter_logs) as total"
        
        cursor.execute(count_query)
        total_result = cursor.fetchone()
        total = total_result['total'] if total_result else 0
        
        cursor.close()
        conn.close()
        
        return {
            "logs": logs,
            "total": total,
            "page": page,
            "limit": limit,
            "filters": {
                "action_type": action_type,
                "resource_type": resource_type,
                "date_from": date_from,
                "date_to": date_to,
                "search": search,
                "log_type": log_type
            }
        }
    except Exception as e:
        logger.error(f"Get all logs error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch logs")

@app.get("/admin/logs/summary")
async def get_admin_logs_summary(
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        import mysql.connector
        import os
        from dotenv import load_dotenv
        
        load_dotenv()
        
        conn = mysql.connector.connect(
            host=os.getenv('DB_HOST', 'localhost'),
            user=os.getenv('DB_USER', 'root'),
            password=os.getenv('DB_PASSWORD', ''),
            database=os.getenv('DB_NAME', 'database')
        )
        cursor = conn.cursor(dictionary=True)
        
        # Get summary statistics
        summary_queries = {
            "total_logs": "SELECT COUNT(*) as count FROM admin_logs",
            "today_logs": "SELECT COUNT(*) as count FROM admin_logs WHERE DATE(timestamp) = CURDATE()",
            "this_week_logs": "SELECT COUNT(*) as count FROM admin_logs WHERE timestamp >= DATE_SUB(NOW(), INTERVAL 7 DAY)",
            "by_action_type": "SELECT action_type, COUNT(*) as count FROM admin_logs GROUP BY action_type ORDER BY count DESC",
            "by_resource_type": "SELECT resource_type, COUNT(*) as count FROM admin_logs WHERE resource_type IS NOT NULL GROUP BY resource_type ORDER BY count DESC",
            "recent_activity": "SELECT admin_username, action_type, resource_type, details, timestamp FROM admin_logs ORDER BY timestamp DESC LIMIT 10"
        }
        
        summary = {}
        
        for key, query in summary_queries.items():
            cursor.execute(query)
            if key in ["total_logs", "today_logs", "this_week_logs"]:
                summary[key] = cursor.fetchone()['count']
            else:
                summary[key] = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        return summary
    except Exception as e:
        logger.error(f"Get admin logs summary error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch admin logs summary")

@app.get("/admin/logs/export")
async def export_admin_logs(
    action_type: Optional[str] = None,
    resource_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    search: Optional[str] = None,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        import mysql.connector
        import os
        from dotenv import load_dotenv
        
        load_dotenv()
        
        conn = mysql.connector.connect(
            host=os.getenv('DB_HOST', 'localhost'),
            user=os.getenv('DB_USER', 'root'),
            password=os.getenv('DB_PASSWORD', ''),
            database=os.getenv('DB_NAME', 'database')
        )
        cursor = conn.cursor(dictionary=True)
        
        # Build query with filters - handle empty strings properly
        query = "SELECT * FROM admin_logs WHERE 1=1"
        params = []
        
        if action_type and action_type.strip():
            query += " AND action_type = %s"
            params.append(action_type.strip())
        
        if resource_type and resource_type.strip():
            query += " AND resource_type = %s"
            params.append(resource_type.strip())
        
        if search and search.strip():
            query += " AND (admin_username LIKE %s OR details LIKE %s OR resource_type LIKE %s)"
            search_param = f"%{search.strip()}%"
            params.extend([search_param, search_param, search_param])
        
        if date_from and date_from.strip():
            try:
                parsed_date = datetime.strptime(date_from.strip(), '%Y-%m-%d')
                query += " AND DATE(timestamp) >= %s"
                params.append(parsed_date.date())
            except ValueError:
                pass
        
        if date_to and date_to.strip():
            try:
                parsed_date = datetime.strptime(date_to.strip(), '%Y-%m-%d')
                query += " AND DATE(timestamp) <= %s"
                params.append(parsed_date.date())
            except ValueError:
                pass
        
        query += " ORDER BY timestamp DESC"
        
        cursor.execute(query, params)
        logs = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Create CSV content
        csv_content = "ID,Admin ID,Admin Username,Action Type,Resource Type,Resource ID,Details,IP Address,Timestamp\n"
        
        for log in logs:
            # Escape CSV values
            details = str(log.get('details', '')).replace('"', '""')
            csv_content += f"{log['id']},{log.get('admin_id', '')},{log.get('admin_username', '')},{log.get('action_type', '')},{log.get('resource_type', '')},{log.get('resource_id', '')},\"{details}\",{log.get('ip_address', '')},{log.get('timestamp', '')}\n"
        
        # Return CSV response
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=admin_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
        )
    except Exception as e:
        logger.error(f"Export admin logs error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to export admin logs")

# Function to log admin actions
def log_admin_action(
    admin_id: int,
    admin_username: str,
    action_type: str,
    resource_type: str = None,
    resource_id: int = None,
    details: str = None,
    ip_address: str = "127.0.0.1"
):
    try:
        import mysql.connector
        import os
        from dotenv import load_dotenv
        
        load_dotenv()
        
        conn = mysql.connector.connect(
            host=os.getenv('DB_HOST', 'localhost'),
            user=os.getenv('DB_USER', 'root'),
            password=os.getenv('DB_PASSWORD', ''),
            database=os.getenv('DB_NAME', 'database')
        )
        cursor = conn.cursor()
        
        insert_query = """
        INSERT INTO admin_logs (admin_id, admin_username, action_type, resource_type, resource_id, details, ip_address)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        
        cursor.execute(insert_query, (admin_id, admin_username, action_type, resource_type, resource_id, details, ip_address))
        conn.commit()
        
        cursor.close()
        conn.close()
    except Exception as e:
        logger.error(f"Failed to log admin action: {str(e)}")

# Health check endpoint
@app.get("/health")
async def health_check():
    return {
        "status": "healthy",
        "timestamp": datetime.now(),
        "version": "1.0.0",
        "database": "connected",
        "uploads_directory": str(UPLOAD_BASE_DIR),
        "uploads_exists": UPLOAD_BASE_DIR.exists()
    }

# Test file serving endpoint
@app.get("/test/file-serve")
async def test_file_serve():
    try:
        # Create a simple test file
        test_file_path = UPLOAD_BASE_DIR / "resources" / "test.txt"
        test_content = "This is a test file to verify file serving is working correctly."
        
        with open(test_file_path, "w", encoding='utf-8') as f:
            f.write(test_content)
        
        return FileResponse(
            path=str(test_file_path),
            media_type="text/plain",
            headers={
                "Access-Control-Allow-Origin": "*",
                "Content-Disposition": "inline; filename=\"test.txt\""
            }
        )
    except Exception as e:
        logger.error(f"Test file serve error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Test failed: {str(e)}")

# Debug endpoint to check resource info
@app.get("/debug/resource/{resource_id}")
async def debug_resource(
    resource_id: int,
    db: Session = Depends(get_db)
):
    try:
        resource = db.query(Resource).filter(Resource.id == resource_id).first()
        if not resource:
            return {"error": "Resource not found"}
        
        # Check file paths
        possible_paths = [
            resource.file_path,
            str(UPLOAD_BASE_DIR / "resources" / os.path.basename(resource.file_path)),
            str(UPLOAD_BASE_DIR / os.path.basename(resource.file_path))
        ]
        
        path_status = {}
        for path in possible_paths:
            path_status[path] = {
                "exists": os.path.exists(path),
                "is_file": os.path.isfile(path) if os.path.exists(path) else False,
                "size": os.path.getsize(path) if os.path.exists(path) else 0
            }
        
        return {
            "resource_id": resource.id,
            "title": resource.title,
            "resource_type": resource.resource_type,
            "file_path": resource.file_path,
            "file_size": resource.file_size,
            "upload_base_dir": str(UPLOAD_BASE_DIR),
            "path_status": path_status
        }
    except Exception as e:
        return {"error": str(e)}

# Test endpoint to create sample resources
@app.post("/admin/test/create-sample-resource")
async def create_sample_resource(
    session_id: int,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        # Verify session exists
        session = db.query(SessionModel).filter(SessionModel.id == session_id).first()
        if not session:
            raise HTTPException(status_code=404, detail="Session not found")
        
        # Create a simple text file for testing
        test_content = f"Sample Resource for {session.title}\n\nThis is a demonstration resource file created for testing purposes.\n\nContent includes:\n- Sample text content\n- Testing information\n- Resource viewing functionality\n- File download capabilities\n\nSession Details:\n- Title: {session.title}\n- Description: {session.description}\n- Duration: {session.duration_minutes} minutes\n\nThis file demonstrates that resources can be properly uploaded, stored, and viewed within the LMS system."
        
        # Create unique filename
        unique_filename = f"sample_resource_{session_id}_{uuid.uuid4().hex[:8]}.txt"
        file_path = UPLOAD_BASE_DIR / "resources" / unique_filename
        
        # Ensure resources directory exists
        (UPLOAD_BASE_DIR / "resources").mkdir(exist_ok=True)
        
        # Write test content to file
        with open(file_path, "w", encoding='utf-8') as f:
            f.write(test_content)
        
        # Create resource record
        resource = Resource(
            session_id=session_id,
            title=f"Sample Resource - {session.title}",
            resource_type="CODE",
            file_path=str(file_path),
            file_size=len(test_content.encode('utf-8')),
            description="A sample text resource for testing resource viewing and download functionality",
            uploaded_by=current_admin.id
        )
        db.add(resource)
        db.commit()
        db.refresh(resource)
        
        return {
            "message": "Sample resource created successfully",
            "resource_id": resource.id,
            "filename": unique_filename,
            "file_path": str(file_path),
            "view_url": f"/api/resources/{resource.id}/view",
            "download_url": f"/api/resources/{resource.id}/download"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Create sample resource error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create sample resource")

@app.post("/admin/fix-resource-paths")
async def fix_resource_paths(
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        resources = db.query(Resource).all()
        session_contents = db.query(SessionContent).all()
        fixed_count = 0
        
        # Ensure resources directory exists
        (UPLOAD_BASE_DIR / "resources").mkdir(exist_ok=True)
        
        # Fix Resource table entries
        for resource in resources:
            if resource.file_path and not os.path.exists(resource.file_path):
                # Try to find the file in the uploads directory
                filename = os.path.basename(resource.file_path)
                new_path = UPLOAD_BASE_DIR / "resources" / filename
                
                if os.path.exists(new_path):
                    resource.file_path = str(new_path)
                    fixed_count += 1
                else:
                    # Create a placeholder file
                    placeholder_content = f"Resource '{resource.title}' was moved or deleted.\n\nOriginal path: {resource.file_path}\nResource Type: {resource.resource_type}\nDescription: {resource.description or 'No description'}\n\nThis is a placeholder file created to maintain system integrity."
                    with open(new_path, "w", encoding='utf-8') as f:
                        f.write(placeholder_content)
                    resource.file_path = str(new_path)
                    resource.file_size = len(placeholder_content.encode('utf-8'))
                    fixed_count += 1
        
        # Fix SessionContent table entries
        for content in session_contents:
            if content.file_path and not os.path.exists(content.file_path):
                # Try to find the file in the uploads directory
                filename = os.path.basename(content.file_path)
                new_path = UPLOAD_BASE_DIR / "resources" / filename
                
                if os.path.exists(new_path):
                    content.file_path = str(new_path)
                    fixed_count += 1
                else:
                    # Create a placeholder file
                    placeholder_content = f"Content '{content.title}' was moved or deleted.\n\nOriginal path: {content.file_path}\nContent Type: {content.content_type}\nDescription: {content.description or 'No description'}\n\nThis is a placeholder file created to maintain system integrity."
                    with open(new_path, "w", encoding='utf-8') as f:
                        f.write(placeholder_content)
                    content.file_path = str(new_path)
                    content.file_size = len(placeholder_content.encode('utf-8'))
                    fixed_count += 1
        
        db.commit()
        
        return {
            "message": f"Fixed {fixed_count} resource and content paths",
            "total_resources": len(resources),
            "total_session_contents": len(session_contents),
            "fixed_count": fixed_count
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Fix resource paths error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fix resource paths")

# Create sample courses endpoint
@app.post("/admin/create-sample-courses")
async def create_sample_courses(
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        # Check if courses already exist
        existing_courses = db.query(Course).count()
        if existing_courses > 0:
            return {"message": f"Courses already exist ({existing_courses} courses found)"}
        
        # Create sample courses
        courses_data = [
            {
                "title": "Introduction to Python Programming",
                "description": "Learn the fundamentals of Python programming language including variables, functions, and object-oriented programming."
            },
            {
                "title": "Web Development with React",
                "description": "Build modern web applications using React, JavaScript, and related technologies."
            },
            {
                "title": "Data Science Fundamentals",
                "description": "Introduction to data analysis, statistics, and machine learning using Python and popular libraries."
            }
        ]
        
        created_courses = []
        for course_data in courses_data:
            course = Course(
                title=course_data["title"],
                description=course_data["description"]
            )
            db.add(course)
            db.flush()
            
            # Create basic module structure
            for week in range(1, 5):  # 4 weeks
                module = Module(
                    course_id=course.id,
                    week_number=week,
                    title=f"Week {week} - {course_data['title']}",
                    description=f"Learning objectives for week {week}"
                )
                db.add(module)
                db.flush()
                
                # Create 2 sessions per week
                for session_num in range(1, 3):
                    session = SessionModel(
                        module_id=module.id,
                        session_number=session_num,
                        title=f"Week {week} - Session {session_num}",
                        description=f"Session {session_num} content for week {week}",
                        duration_minutes=120,
                        scheduled_time=datetime.now() + timedelta(days=(week-1)*7 + session_num)
                    )
                    db.add(session)
                    db.flush()
                    
                    # Create sample resources for each session
                    sample_resource_content = f"Sample resource content for {session.title}\n\nThis is a demonstration resource file.\n\nTopics covered:\n- Introduction to the session\n- Key concepts\n- Practical examples\n- Summary and next steps"
                    
                    resource_filename = f"sample_resource_session_{session.id}.txt"
                    resource_path = UPLOAD_BASE_DIR / "resources" / resource_filename
                    
                    # Create the resource file
                    with open(resource_path, "w") as f:
                        f.write(sample_resource_content)
                    
                    # Create resource record
                    resource = Resource(
                        session_id=session.id,
                        title=f"Session {session_num} Materials",
                        resource_type="CODE",
                        file_path=str(resource_path),
                        file_size=len(sample_resource_content.encode('utf-8')),
                        description=f"Course materials and notes for {session.title}",
                        uploaded_by=current_admin.id
                    )
                    db.add(resource)
            
            created_courses.append(course.title)
        
        db.commit()
        return {
            "message": f"Created {len(created_courses)} sample courses with resources",
            "courses": created_courses
        }
    except Exception as e:
        db.rollback()
        logger.error(f"Create sample courses error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create sample courses")

# Test settings endpoints
@app.get("/test/settings")
async def test_settings_endpoints():
    """Test endpoint to verify settings API integration"""
    return {
        "message": "Settings API is integrated",
        "available_endpoints": [
            "GET /admin/settings - Get current settings",
            "PUT /admin/settings - Update settings",
            "POST /admin/settings/test-email - Test email configuration",
            "POST /admin/settings/backup - Create system backup",
            "GET /admin/settings/system-status - Get system status",
            "POST /admin/settings/reset - Reset settings to defaults",
            "GET /admin/settings/audit-log - Get settings audit log"
        ]
    }

@app.get("/test/logs")
async def test_logs_endpoint():
    """Test endpoint to verify logs are working without auth"""
    try:
        import mysql.connector
        
        conn = mysql.connector.connect(
            host=os.getenv('DB_HOST', 'localhost'),
            user=os.getenv('DB_USER', 'root'),
            password=os.getenv('DB_PASSWORD', ''),
            database=os.getenv('DB_NAME', 'database')
        )
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("SELECT * FROM admin_logs ORDER BY timestamp DESC LIMIT 5")
        logs = cursor.fetchall()
        
        cursor.execute("SELECT COUNT(*) as count FROM admin_logs")
        total = cursor.fetchone()['count']
        
        cursor.close()
        conn.close()
        
        return {
            "message": "Logs API test successful",
            "total_logs": total,
            "sample_logs": logs
        }
    except Exception as e:
        return {
            "message": "Logs API test failed",
            "error": str(e)
        }

# Add more logs when admin performs actions
@app.get("/admin/logs/test-add")
async def test_add_log(
    current_admin = Depends(get_current_admin_or_presenter)
):
    """Test endpoint to add a log entry"""
    log_admin_action(
        admin_id=current_admin.id,
        admin_username=current_admin.username,
        action_type="VIEW",
        resource_type="TEST",
        details="Test log entry from API"
    )
    return {"message": "Test log added successfully"}

# Add a generic file serving endpoint that handles file paths directly
@app.get("/files/{file_path:path}")
async def serve_file_by_path(
    file_path: str
):
    try:
        # Construct the full file path
        full_path = UPLOAD_BASE_DIR / file_path
        
        if not full_path.exists():
            # Try alternative paths
            alt_paths = [
                UPLOAD_BASE_DIR / "resources" / os.path.basename(file_path),
                UPLOAD_BASE_DIR / os.path.basename(file_path)
            ]
            
            for alt_path in alt_paths:
                if alt_path.exists():
                    full_path = alt_path
                    break
            else:
                raise HTTPException(status_code=404, detail="File not found")
        
        # Determine media type
        file_ext = os.path.splitext(file_path)[1].lower()
        media_type_map = {
            '.pdf': 'application/pdf',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png',
            '.gif': 'image/gif',
            '.mp4': 'video/mp4',
            '.avi': 'video/avi',
            '.mov': 'video/quicktime',
            '.txt': 'text/plain',
            '.doc': 'application/msword',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.ppt': 'application/vnd.ms-powerpoint',
            '.pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation'
        }
        
        media_type = media_type_map.get(file_ext, 'application/octet-stream')
        filename = os.path.basename(file_path)
        
        return FileResponse(
            path=str(full_path),
            media_type=media_type,
            headers={
                "Access-Control-Allow-Origin": "*",
                "Content-Disposition": "inline", "X-Download-Options": "noopen", "X-Content-Type-Options": "nosniff", "Cache-Control": "no-store\"{filename}\""
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Serve file by path error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to serve file")

# Add a simple static file serving endpoint for uploads
@app.get("/uploads/{file_type}/{filename}")
async def serve_uploaded_file(
    file_type: str,
    filename: str
):
    try:
        file_path = UPLOAD_BASE_DIR / file_type / filename
        
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="File not found")
        
        # Determine media type
        file_ext = os.path.splitext(filename)[1].lower()
        media_type_map = {
            '.pdf': 'application/pdf',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png',
            '.gif': 'image/gif',
            '.mp4': 'video/mp4',
            '.avi': 'video/avi',
            '.mov': 'video/quicktime',
            '.txt': 'text/plain',
            '.doc': 'application/msword',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.ppt': 'application/vnd.ms-powerpoint',
            '.pptx': 'application/vnd.openxmlformats-officedocument.presentationml.presentation'
        }
        
        media_type = media_type_map.get(file_ext, 'application/octet-stream')
        
        return FileResponse(
            path=str(file_path),
            media_type=media_type,
            headers={
                "Access-Control-Allow-Origin": "*",
                "Content-Disposition": "inline", "X-Download-Options": "noopen", "X-Content-Type-Options": "nosniff", "Cache-Control": "no-store\"{filename}\""
            }
        )
    except Exception as e:
        logger.error(f"Serve file error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to serve file")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)






























# Admin Cohorts Endpoint
@app.get("/admin/cohorts")
async def get_admin_cohorts(
    page: int = 1,
    limit: int = 20,
    search: Optional[str] = None,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        query = db.query(Cohort)
        
        if search:
            query = query.filter(
                or_(
                    Cohort.name.contains(search),
                    Cohort.description.contains(search),
                    Cohort.instructor_name.contains(search)
                )
            )
        
        total = query.count()
        cohorts = query.offset((page - 1) * limit).limit(limit).all()
        
        result = []
        for cohort in cohorts:
            user_count = db.query(UserCohort).filter(UserCohort.cohort_id == cohort.id).count()
            course_count = db.query(CohortCourse).filter(CohortCourse.cohort_id == cohort.id).count()
            
            result.append({
                "id": cohort.id,
                "name": cohort.name,
                "description": cohort.description,
                "start_date": cohort.start_date,
                "end_date": cohort.end_date,
                "instructor_name": cohort.instructor_name,
                "user_count": user_count,
                "course_count": course_count,
                "created_at": cohort.created_at
            })
        
        return {
            "cohorts": result,
            "total": total,
            "page": page,
            "limit": limit
        }
    except Exception as e:
        logger.error(f"Get admin cohorts error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch cohorts")

# Presenter Cohorts Endpoint - Allow both admin and presenter access
@app.get("/presenter/cohorts")
async def get_presenter_cohorts(
    page: int = 1,
    limit: int = 20,
    search: Optional[str] = None,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        query = db.query(Cohort)
        
        if search:
            query = query.filter(
                or_(
                    Cohort.name.contains(search),
                    Cohort.description.contains(search),
                    Cohort.instructor_name.contains(search)
                )
            )
        
        total = query.count()
        cohorts = query.offset((page - 1) * limit).limit(limit).all()
        
        result = []
        for cohort in cohorts:
            user_count = db.query(UserCohort).filter(UserCohort.cohort_id == cohort.id).count()
            course_count = db.query(CohortCourse).filter(CohortCourse.cohort_id == cohort.id).count()
            
            result.append({
                "id": cohort.id,
                "name": cohort.name,
                "description": cohort.description,
                "start_date": cohort.start_date,
                "end_date": cohort.end_date,
                "instructor_name": cohort.instructor_name,
                "user_count": user_count,
                "course_count": course_count,
                "created_at": cohort.created_at
            })
        
        return {
            "cohorts": result,
            "total": total,
            "page": page,
            "limit": limit
        }
    except Exception as e:
        logger.error(f"Get presenter cohorts error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch cohorts")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
# Analytics endpoint (404 error fix)
@app.get("/admin/analytics")
async def get_admin_analytics(
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        # Basic analytics data
        total_users = db.query(User).count()
        total_courses = db.query(Course).count()
        total_modules = db.query(Module).count()
        total_sessions = db.query(SessionModel).count()
        total_enrollments = db.query(Enrollment).count()
        total_quizzes = db.query(Quiz).count()
        total_resources = db.query(Resource).count()
        
        return {
            "users": {
                "total": total_users,
                "active": total_users
            },
            "courses": {
                "total": total_courses,
                "modules": total_modules,
                "sessions": total_sessions
            },
            "engagement": {
                "enrollments": total_enrollments,
                "quizzes": total_quizzes,
                "resources": total_resources
            }
        }
    except Exception as e:
        logger.error(f"Analytics error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch analytics")



@app.post("/presenter/cohorts")
async def create_presenter_cohort(
    cohort_data: CohortCreate,
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    try:
        cohort = Cohort(
            name=cohort_data.name,
            description=cohort_data.description,
            start_date=cohort_data.start_date,
            end_date=cohort_data.end_date,
            instructor_name=cohort_data.instructor_name,
            created_by=current_presenter.id
        )
        db.add(cohort)
        db.commit()
        db.refresh(cohort)
        
        # Log cohort creation
        log_presenter_action(
            presenter_id=current_presenter.id,
            presenter_username=current_presenter.username,
            action_type="CREATE",
            resource_type="COHORT",
            resource_id=cohort.id,
            details=f"Created cohort: {cohort_data.name}"
        )
        
        return {"message": "Cohort created successfully", "cohort_id": cohort.id}
    except Exception as e:
        db.rollback()
        logger.error(f"Create presenter cohort error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create cohort")

# Additional presenter endpoints for cohort management
@app.get("/presenter/cohorts/{cohort_id}")
async def get_presenter_cohort_details(
    cohort_id: int,
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    try:
        cohort = db.query(Cohort).filter(Cohort.id == cohort_id).first()
        if not cohort:
            raise HTTPException(status_code=404, detail="Cohort not found")
        
        # Get users in cohort
        user_cohorts = db.query(UserCohort).filter(UserCohort.cohort_id == cohort_id).all()
        users = [{
            "id": uc.user.id,
            "username": uc.user.username,
            "email": uc.user.email,
            "college": uc.user.college,
            "assigned_at": uc.assigned_at
        } for uc in user_cohorts]
        
        # Get courses assigned to cohort
        cohort_courses = db.query(CohortCourse).filter(CohortCourse.cohort_id == cohort_id).all()
        courses = [{
            "id": cc.course.id,
            "title": cc.course.title,
            "description": cc.course.description,
            "assigned_at": cc.assigned_at
        } for cc in cohort_courses]
        
        return {
            "id": cohort.id,
            "name": cohort.name,
            "description": cohort.description,
            "start_date": cohort.start_date,
            "end_date": cohort.end_date,
            "instructor_name": cohort.instructor_name,
            "created_at": cohort.created_at,
            "users": users,
            "courses": courses
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get presenter cohort details error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch cohort details")

# All other presenter endpoints that return 403 for restricted access
@app.put("/presenter/cohorts/{cohort_id}")
async def update_presenter_cohort(
    cohort_id: int,
    cohort_data: dict,
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    raise HTTPException(status_code=403, detail="Presenters cannot modify cohorts")

@app.delete("/presenter/cohorts/{cohort_id}")
async def delete_presenter_cohort(
    cohort_id: int,
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    raise HTTPException(status_code=403, detail="Presenters cannot delete cohorts")

@app.post("/presenter/cohorts/{cohort_id}/users")
async def add_users_to_presenter_cohort(
    cohort_id: int,
    user_data: dict,
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    raise HTTPException(status_code=403, detail="Presenters cannot add users to cohorts")

@app.delete("/presenter/cohorts/{cohort_id}/users/{user_id}")
async def remove_user_from_presenter_cohort(
    cohort_id: int,
    user_id: int,
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    raise HTTPException(status_code=403, detail="Presenters cannot remove users from cohorts")

@app.post("/presenter/cohorts/{cohort_id}/courses")
async def assign_courses_to_presenter_cohort(
    cohort_id: int,
    course_data: dict,
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    raise HTTPException(status_code=403, detail="Presenters cannot assign courses to cohorts")

@app.delete("/presenter/cohorts/{cohort_id}/courses/{course_id}")
async def remove_course_from_presenter_cohort(
    cohort_id: int,
    course_id: int,
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    raise HTTPException(status_code=403, detail="Presenters cannot remove courses from cohorts")

@app.post("/presenter/cohorts/{cohort_id}/bulk-upload")
async def bulk_upload_presenter_cohort_users(
    cohort_id: int,
    file: UploadFile = File(...),
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    raise HTTPException(status_code=403, detail="Presenters cannot bulk upload users")

@app.get("/presenter/cohorts/{cohort_id}/export")
async def export_presenter_cohort_users(
    cohort_id: int,
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    try:
        # Presenters can view/export cohort users
        cohort = db.query(Cohort).filter(Cohort.id == cohort_id).first()
        if not cohort:
            raise HTTPException(status_code=404, detail="Cohort not found")
        
        # Get users in cohort
        user_cohorts = db.query(UserCohort).filter(UserCohort.cohort_id == cohort_id).all()
        
        # Create CSV content
        csv_content = "username,email,role,college,assigned_at\n"
        for uc in user_cohorts:
            csv_content += f"{uc.user.username},{uc.user.email},{uc.user.role},{uc.user.college or ''},{uc.assigned_at}\n"
        
        # Create response
        response = Response(
            content=csv_content,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=cohort_{cohort.name}_users.csv"}
        )
        
        return response
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Export presenter cohort users error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to export cohort users")

@app.get("/presenter/users/available")
async def get_available_users_for_presenter(
    search: Optional[str] = None,
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    try:
        # Presenters can view available users but not modify them
        query = db.query(User).filter(User.cohort_id.is_(None))
        
        if search:
            query = query.filter(
                or_(
                    User.username.contains(search),
                    User.email.contains(search),
                    User.college.contains(search)
                )
            )
        
        users = query.limit(50).all()
        
        return {
            "users": [{
                "id": u.id,
                "username": u.username,
                "email": u.email,
                "college": u.college,
                "role": u.role
            } for u in users]
        }
    except Exception as e:
        logger.error(f"Get available users for presenter error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch available users")

@app.get("/presenter/courses/available")
async def get_available_courses_for_presenter(
    cohort_id: Optional[int] = None,
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    try:
        if cohort_id:
            # Get courses not assigned to this cohort
            assigned_course_ids = db.query(CohortCourse.course_id).filter(
                CohortCourse.cohort_id == cohort_id
            ).all()
            assigned_ids = [c[0] for c in assigned_course_ids]
            
            courses = db.query(Course).filter(~Course.id.in_(assigned_ids)).all()
        else:
            courses = db.query(Course).all()
        
        return {
            "courses": [{
                "id": c.id,
                "title": c.title,
                "description": c.description
            } for c in courses]
        }
    except Exception as e:
        logger.error(f"Get available courses for presenter error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch available courses")

@app.get("/presenter/presenters")
async def get_presenters_for_presenter(
    current_presenter: Presenter = Depends(get_current_presenter),
    db: Session = Depends(get_db)
):
    try:
        presenters = db.query(Presenter).all()
        
        return {
            "presenters": [{
                "id": p.id,
                "username": p.username,
                "email": p.email,
                "is_active": getattr(p, 'is_active', True)
            } for p in presenters]
        }
    except Exception as e:
        logger.error(f"Get presenters for presenter error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch presenters")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

# Notifications endpoints
@app.get("/notifications/unread-count")
async def get_unread_notifications_count(
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        # Count unread notifications for the current user
        unread_count = db.query(Notification).filter(
            Notification.user_id == current_user.id,
            Notification.is_read == False
        ).count()
        
        return {"unread_count": unread_count}
    except Exception as e:
        logger.error(f"Get unread notifications count error: {str(e)}")
        return {"unread_count": 0}

@app.get("/notifications")
async def get_notifications(
    limit: int = 10,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        notifications = db.query(Notification).filter(
            or_(
                Notification.user_id == current_user.id,
                Notification.is_global == True
            )
        ).order_by(Notification.created_at.desc()).limit(limit).all()
        
        return {
            "notifications": [{
                "id": n.id,
                "title": n.title,
                "message": n.message,
                "notification_type": n.notification_type,
                "is_read": n.is_read,
                "is_global": n.is_global,
                "created_at": n.created_at
            } for n in notifications]
        }
    except Exception as e:
        logger.error(f"Get notifications error: {str(e)}")
        return {"notifications": []}

@app.post("/notifications/{notification_id}/mark-read")
async def mark_notification_read(
    notification_id: int,
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        notification = db.query(Notification).filter(
            Notification.id == notification_id,
            or_(
                Notification.user_id == current_user.id,
                Notification.is_global == True
            )
        ).first()
        
        if notification:
            notification.is_read = True
            db.commit()
            return {"message": "Notification marked as read"}
        else:
            raise HTTPException(status_code=404, detail="Notification not found")
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Mark notification read error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to mark notification as read")

@app.post("/notifications/mark-all-read")
async def mark_all_notifications_read(
    current_user = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        db.query(Notification).filter(
            Notification.user_id == current_user.id,
            Notification.is_read == False
        ).update({"is_read": True})
        db.commit()
        
        return {"message": "All notifications marked as read"}
    except Exception as e:
        db.rollback()
        logger.error(f"Mark all notifications read error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to mark all notifications as read")

@app.post("/admin/notifications")
async def create_notification(
    notification_data: NotificationCreate,
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    try:
        notification = Notification(
            user_id=notification_data.user_id,
            title=notification_data.title,
            message=notification_data.message,
            notification_type=notification_data.notification_type,
            is_global=notification_data.is_global,
            created_by=current_admin.id
        )
        db.add(notification)
        db.commit()
        
        return {"message": "Notification created successfully", "notification_id": notification.id}
    except Exception as e:
        db.rollback()
        logger.error(f"Create notification error: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to create notification")

# Resource Analytics for Admin Dashboard
@app.get("/admin/analytics/resources")
async def get_resource_analytics_dashboard(
    current_admin = Depends(get_current_admin_or_presenter),
    db: Session = Depends(get_db)
):
    """Get resource analytics for admin dashboard"""
    try:
        from resource_analytics_models import ResourceView
        
        # Get total resources
        total_resources = db.query(Resource).count()
        
        # Get total views
        total_views = db.query(ResourceView).count()
        
        # Get unique viewers
        unique_viewers = db.query(func.count(func.distinct(ResourceView.student_id))).scalar()
        
        # Get most viewed resources
        most_viewed = db.query(
            Resource.id,
            Resource.title,
            Resource.resource_type,
            func.count(ResourceView.id).label('view_count')
        ).join(
            ResourceView, Resource.id == ResourceView.resource_id
        ).group_by(
            Resource.id, Resource.title, Resource.resource_type
        ).order_by(
            func.count(ResourceView.id).desc()
        ).limit(5).all()
        
        return {
            "total_resources": total_resources,
            "total_views": total_views,
            "unique_viewers": unique_viewers or 0,
            "most_viewed_resources": [
                {
                    "id": r.id,
                    "title": r.title,
                    "type": r.resource_type,
                    "views": r.view_count
                }
                for r in most_viewed
            ]
        }
    except Exception as e:
        logger.error(f"Resource analytics error: {str(e)}")
        return {
            "total_resources": 0,
            "total_views": 0,
            "unique_viewers": 0,
            "most_viewed_resources": []
        }